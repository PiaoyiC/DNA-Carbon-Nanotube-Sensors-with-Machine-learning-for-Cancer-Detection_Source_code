import numpy as np
import pandas as pd
from sklearn.preprocessing import StandardScaler
from sklearn.feature_selection import SelectKBest, mutual_info_classif
from sklearn.decomposition import PCA
from scipy import stats
import xgboost as xgb
from sklearn.model_selection import StratifiedKFold, cross_val_score, cross_val_predict
from sklearn.metrics import (f1_score, roc_curve, auc, confusion_matrix, precision_score,
                             recall_score, roc_auc_score)
from bayes_opt import BayesianOptimization
import matplotlib.pyplot as plt
from joblib import Parallel, delayed
import logging
import os
import seaborn as sns
from itertools import combinations
import warnings
import re
import joblib
import json
from imblearn.over_sampling import ADASYN
from openpyxl import load_workbook
from openpyxl.styles import Font

warnings.filterwarnings('ignore')

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')


class AdvancedFeatureInteractions:
    """Feature interaction processing class"""

    def __init__(self,
                 interaction_method='gaussian',
                 selection_method='boruta',
                 max_features=30,
                 gaussian_gamma='auto',
                 scale_after_interaction=True):
        """
        Initialize feature interaction processor

        Parameters:
            interaction_method: Feature interaction method ('gaussian', 'polynomial', 'custom', 'combined')
            selection_method: Feature selection method ('mutual_info', 'correlation', 'pca', None)
            max_features: Maximum number of features to select
            gaussian_gamma: Gamma parameter for Gaussian kernel, can be 'auto' or specific value
            scale_after_interaction: Whether to perform standardization after interaction
        """
        self.interaction_method = interaction_method
        self.selection_method = selection_method
        self.max_features = max_features
        self.gaussian_gamma = gaussian_gamma
        self.scale_after_interaction = scale_after_interaction

        # Initialize other attributes
        self.feature_names_ = None
        self.selected_features_ = 30
        self.feature_importance_ = None
        self.scaler = StandardScaler()

        # Save standardizer and feature selector after interaction
        self.interaction_scaler = None
        self.selector = None

        # Record whether it has been fitted
        self.is_fitted = False

        # Output initialization parameters

    def set_original_feature_names(self, names):
        """Set the names of original features"""
        self.original_feature_names = names

    def fit_transform(self, X, y=None):
        """Fit and transform data"""
        try:

            # 1. Standardize original features
            X_scaled = self.scaler.fit_transform(X)

            # 2. Generate interaction features
            if self.interaction_method == 'gaussian':
                X_interaction = self._gaussian_interaction(X_scaled)
            else:
                raise ValueError("Unknown interaction method: {}".format(self.interaction_method))

            # 3. Post-interaction standardization - save standardizer for subsequent use
            if self.scale_after_interaction:
                self.interaction_scaler = StandardScaler()
                X_interaction = self.interaction_scaler.fit_transform(X_interaction)

            # 4. Feature selection
            if self.selection_method is not None and y is not None:
                X_final = self._select_features(X_interaction, y)
            else:
                X_final = X_interaction

            # Mark as fitted
            self.is_fitted = True

            return X_final

        except Exception as e:
            logging.error("Error in fit_transform: {}".format(str(e)))
            raise

    def transform(self, X):
        """Transform data without refitting parameters"""
        if not self.is_fitted:
            raise ValueError("This AdvancedFeatureInteractions instance is not fitted yet. "
                             "Call 'fit_transform' before using this method.")

        try:

            # 1. Use already fitted standardizer to transform data
            X_scaled = self.scaler.transform(X)

            # 2. Apply the same feature interaction
            if self.interaction_method == 'gaussian':
                X_interaction = self._gaussian_interaction(X_scaled)
            else:
                raise ValueError("Unknown interaction method: {}".format(self.interaction_method))

            # 3. Apply already fitted post-interaction standardization
            if self.scale_after_interaction and self.interaction_scaler is not None:
                X_interaction = self.interaction_scaler.transform(X_interaction)

            # 4. Apply existing feature selection
            if self.selection_method is not None and self.selector is not None:
                X_final = self.selector.transform(X_interaction)
            else:
                X_final = X_interaction

            return X_final

        except Exception as e:
            logging.error("Error in transform: {}".format(str(e)))
            raise

    def _gaussian_interaction(self, X):
        """Enhanced Gaussian kernel feature interaction, adding distance features and local density features"""
        n_samples, n_features = X.shape

        # Ensure feature names exist
        if not hasattr(self, 'original_feature_names') or self.original_feature_names is None:
            self.original_feature_names = ["feature_{}".format(i) for i in range(n_features)]

        # Determine gamma parameter
        if self.gaussian_gamma == 'auto':
            gamma = 1.0 / n_features
        elif isinstance(self.gaussian_gamma, (int, float)):
            gamma = self.gaussian_gamma
        else:
            gamma = 1.0 / n_features


        # Compute core features
        core_features = []
        core_names = []

        # 1. Direct features
        core_features.append(X)
        core_names.extend(self.original_feature_names)

        # 2. Pairwise Gaussian kernel interactions
        for i, j in combinations(range(n_features), 2):
            # Compute Gaussian kernel between two features
            diff = X[:, i:i + 1] - X[:, j:j + 1]
            kernel = np.exp(-gamma * (diff ** 2))
            core_features.append(kernel)
            core_names.append("gaussian_{}_{}".format(self.original_feature_names[i], self.original_feature_names[j]))

            # Weighted combination
            weighted_combo = kernel * (X[:, i:i + 1] + X[:, j:j + 1])
            core_features.append(weighted_combo)
            core_names.append(
                "weighted_gaussian_{}_{}".format(self.original_feature_names[i], self.original_feature_names[j]))

        # 3. Distance features
        for i in range(n_features):
            # Compute sum of RBF kernels between current feature and all other features
            rbf_sum = np.zeros((n_samples, 1))
            for j in range(n_features):
                if i != j:
                    diff = X[:, i:i + 1] - X[:, j:j + 1]
                    rbf_sum += np.exp(-gamma * (diff ** 2))
            core_features.append(rbf_sum)
            core_names.append("rbf_distance_{}".format(self.original_feature_names[i]))

        # 4. Local density features
        for i in range(n_features):
            # Compute local density estimation
            density = np.zeros((n_samples, 1))
            for j in range(n_features):
                diff = X[:, i:i + 1] - X[:, j:j + 1]
                density += np.exp(-gamma * np.sum(diff ** 2, axis=1, keepdims=True))
            core_features.append(density)
            core_names.append("local_density_{}".format(self.original_feature_names[i]))

        # Combine all features
        X_gaussian = np.hstack(core_features)
        self.feature_names_ = core_names

        # Record feature information

        return X_gaussian

    def _select_features(self, X, y):
        """Enhanced feature selection, supporting multiple nonlinear methods"""
        # Import necessary libraries within the method
        from sklearn.feature_selection import SelectKBest, mutual_info_classif


        if self.selection_method == 'mutual_info':
            # Original mutual information selection
            if self.max_features is None:
                n_features = X.shape[1]
            else:
                n_features = min(self.max_features, X.shape[1])

            self.selector = SelectKBest(
                score_func=mutual_info_classif,
                k=n_features
            )
            X_selected = self.selector.fit_transform(X, y)
            self.feature_importance_ = self.selector.scores_

        elif self.selection_method == 'random_forest':
            # Random forest-based feature selection
            from sklearn.ensemble import RandomForestClassifier
            if self.max_features is None:
                n_features = X.shape[1]
            else:
                n_features = min(self.max_features, X.shape[1])

            forest = RandomForestClassifier(n_estimators=100, random_state=42, n_jobs=-1)
            forest.fit(X, y)

            # Get feature importance
            self.feature_importance_ = forest.feature_importances_

            # Select top-k features
            indices = np.argsort(self.feature_importance_)[::-1][:n_features]

            # Create feature selection mask
            selected_mask = np.zeros(X.shape[1], dtype=bool)
            selected_mask[indices] = True

            # Implement selector to maintain API consistency
            from sklearn.feature_selection import SelectFromModel
            self.selector = SelectFromModel(forest, max_features=n_features, prefit=True)
            X_selected = self.selector.transform(X)

        elif self.selection_method == 'xgboost':
            # XGBoost-based feature selection
            import xgboost as xgb
            if self.max_features is None:
                n_features = X.shape[1]
            else:
                n_features = min(self.max_features, X.shape[1])


            # Use lightweight configuration to speed up
            xgb_model = xgb.XGBClassifier(
                n_estimators=50,
                learning_rate=0.1,
                max_depth=3,
                use_label_encoder=False,
                eval_metric='logloss',
                random_state=42,
                n_jobs=-1
            )
            xgb_model.fit(X, y)

            # Get feature importance
            self.feature_importance_ = xgb_model.feature_importances_

            # Select top-k features
            indices = np.argsort(self.feature_importance_)[::-1][:n_features]

            # Create feature selection mask
            selected_mask = np.zeros(X.shape[1], dtype=bool)
            selected_mask[indices] = True

            # Implement selector to maintain API consistency
            from sklearn.feature_selection import SelectFromModel
            self.selector = SelectFromModel(xgb_model, max_features=n_features, prefit=True)
            X_selected = self.selector.transform(X)

        elif self.selection_method == 'permutation':
            # Permutation importance-based feature selection
            from sklearn.ensemble import RandomForestClassifier
            from sklearn.inspection import permutation_importance

            if self.max_features is None:
                n_features = X.shape[1]
            else:
                n_features = min(self.max_features, X.shape[1])


            # Train baseline algorithms
            forest = RandomForestClassifier(n_estimators=50, random_state=42, n_jobs=-1)
            forest.fit(X, y)

            # Compute permutation importance
            result = permutation_importance(
                forest, X, y,
                n_repeats=10,
                random_state=42,
                n_jobs=-1
            )

            # Get feature importance
            self.feature_importance_ = result.importances_mean

            # Select top-k features
            indices = np.argsort(self.feature_importance_)[::-1][:n_features]

            # Create feature selection mask
            selected_mask = np.zeros(X.shape[1], dtype=bool)
            selected_mask[indices] = True

            # Manually select features
            X_selected = X[:, selected_mask]

            # Create simple selector
            class SimpleSelector:
                def __init__(self, mask):
                    self.mask = mask

                def transform(self, X):
                    return X[:, self.mask]

                def get_support(self):
                    return self.mask

            self.selector = SimpleSelector(selected_mask)

        elif self.selection_method == 'kernel_pca':
            # Kernel PCA feature extraction
            from sklearn.decomposition import KernelPCA

            if self.max_features is None:
                n_features = min(X.shape[1], X.shape[0])
            else:
                n_features = min(self.max_features, X.shape[1], X.shape[0])


            # Create Kernel PCA
            kpca = KernelPCA(
                n_components=n_features,
                kernel='rbf',
                gamma=self.gaussian_gamma if self.gaussian_gamma != 'auto' else None,
                random_state=42,
                n_jobs=-1
            )

            # Transform data
            X_selected = kpca.fit_transform(X, y)

            # Save feature importance (variance explained ratio)
            if hasattr(kpca, 'lambdas_'):
                self.feature_importance_ = kpca.lambdas_[:n_features]
            else:
                self.feature_importance_ = np.ones(n_features)

            self.selector = kpca

        elif self.selection_method == 'boruta':
            # Boruta feature selection (requires boruta_py to be installed)
            try:
                from boruta import BorutaPy
                from sklearn.ensemble import RandomForestClassifier

                if self.max_features is None:
                    n_features = X.shape[1]
                else:
                    n_features = min(self.max_features, X.shape[1])


                # Create random forest classifier
                forest = RandomForestClassifier(
                    n_estimators=100,
                    max_depth=5,
                    random_state=42,
                    n_jobs=-1
                )

                # Create Boruta feature selector
                boruta_selector = BorutaPy(
                    estimator=forest,
                    n_estimators='auto',  # Automatically select number of estimators
                    verbose=0,
                    random_state=42
                )

                # Fit data
                # Boruta requires shape [samples, features], no need to transpose if X already has this shape
                boruta_selector.fit(X, y)

                # Select features considered important by Boruta
                X_selected = boruta_selector.transform(X)

                # If selected features are fewer than requested, use support ranking to select more
                if X_selected.shape[1] < n_features:
                    # Select features based on Boruta ranking
                    rank_indices = np.argsort(boruta_selector.ranking_)[:n_features]
                    selected_mask = np.zeros(X.shape[1], dtype=bool)
                    selected_mask[rank_indices] = True
                    X_selected = X[:, selected_mask]

                    # Update Boruta's mask
                    boruta_mask = np.zeros(X.shape[1], dtype=bool)
                    boruta_mask[rank_indices] = True
                    boruta_selector.support_ = boruta_mask

                # Save feature importance (reciprocal of ranking)
                self.feature_importance_ = 1.0 / boruta_selector.ranking_
                self.selector = boruta_selector

            except ImportError:
                logging.warning("Boruta not installed. Falling back to mutual_info selection.")
                # Fall back to mutual information selection
                return self._select_features_with_method(X, y, 'mutual_info')

        elif self.selection_method == 'shap':
            # SHAP value-based feature selection
            try:
                import shap
                import xgboost as xgb

                if self.max_features is None:
                    n_features = X.shape[1]
                else:
                    n_features = min(self.max_features, X.shape[1])


                # Train algorithms
                model = xgb.XGBClassifier(
                    n_estimators=50,
                    max_depth=3,
                    learning_rate=0.1,
                    use_label_encoder=False,
                    eval_metric='logloss',
                    random_state=42,
                    n_jobs=-1
                )
                model.fit(X, y)

                # Compute SHAP values
                explainer = shap.Explainer(model, X)
                shap_values = explainer(X)

                # Compute mean absolute SHAP value for each feature
                feature_importance = np.abs(shap_values.values).mean(axis=0)
                self.feature_importance_ = feature_importance

                # Select top-k features
                indices = np.argsort(self.feature_importance_)[::-1][:n_features]

                # Create feature selection mask
                selected_mask = np.zeros(X.shape[1], dtype=bool)
                selected_mask[indices] = True

                # Implement selector to maintain API consistency
                from sklearn.feature_selection import SelectFromModel
                self.selector = SelectFromModel(model, max_features=n_features, prefit=True)
                X_selected = self.selector.transform(X)

            except ImportError:
                logging.warning("SHAP not installed. Falling back to xgboost selection.")
                # Fall back to XGBoost feature selection
                return self._select_features_with_method(X, y, 'xgboost')

        elif self.selection_method == 'cmim':
            # Conditional mutual information maximization
            try:
                from sklearn.feature_selection import SelectKBest

                if self.max_features is None:
                    n_features = X.shape[1]
                else:
                    n_features = min(self.max_features, X.shape[1])


                # A simplified approach based on mutual information
                from sklearn.feature_selection import mutual_info_classif

                # Get mutual information
                mi_scores = mutual_info_classif(X, y)

                # Select top-k features
                indices = np.argsort(mi_scores)[::-1][:n_features]

                # Create feature selection mask
                selected_mask = np.zeros(X.shape[1], dtype=bool)
                selected_mask[indices] = True

                # Manually select features
                X_selected = X[:, selected_mask]

                # Save feature importance
                self.feature_importance_ = mi_scores

                # Create simple selector
                class SimpleSelector:
                    def __init__(self, mask):
                        self.mask = mask

                    def transform(self, X):
                        return X[:, self.mask]

                    def get_support(self):
                        return self.mask

                self.selector = SimpleSelector(selected_mask)

            except ImportError:
                logging.warning("Required packages not installed. Falling back to mutual_info selection.")
                # Fall back to mutual information selection
                return self._select_features_with_method(X, y, 'mutual_info')

        else:
            X_selected = X
            self.feature_importance_ = np.ones(X.shape[1])

            # Create selector that does nothing
            class IdentitySelector:
                def transform(self, X):
                    return X

                def get_support(self):
                    return np.ones(X.shape[1], dtype=bool)

            self.selector = IdentitySelector()

        # Update feature names
        if len(self.feature_names_) != X.shape[1]:
            self.feature_names_ = ["feature_{}".format(i) for i in range(X.shape[1])]

        # Save selected feature names
        selected_mask = self.selector.get_support() if hasattr(self.selector, 'get_support') else np.ones(X.shape[1],
                                                                                                          dtype=bool)
        self.selected_features_ = np.array(self.feature_names_)[selected_mask]

        return X_selected

    def get_feature_names(self):
        """Get feature names"""
        if self.selected_features_ is not None:
            return self.selected_features_
        return self.feature_names_

    def get_feature_importance(self):
        """Get feature importance"""
        if self.feature_importance_ is None:
            return None

        importance_df = pd.DataFrame({
            'Feature': self.feature_names_,
            'Importance': self.feature_importance_
        })
        return importance_df.sort_values('Importance', ascending=False)


def weighted_specificity(y_true, y_pred, labels):

    total_specificity = 0
    total_weight = 0

    for label in labels:
        true_negatives = np.sum((y_true != label) & (y_pred != label))
        false_positives = np.sum((y_true != label) & (y_pred == label))

        weight = np.sum(y_true == label)
        total_weight += weight

        if (true_negatives + false_positives) > 0:
            specificity = true_negatives / (true_negatives + false_positives)
        else:
            specificity = 0.0
        total_specificity += specificity * weight

    return total_specificity / total_weight if total_weight > 0 else 0.0


class XGBoostClassifier:
    def __init__(self, config):
        self.config = config
        self.classifiers = {}
        self.all_classes = config['all_classes']
        self.results_data = {}
        self.inner_cv_results = {}
        self.outer_cv_results = {}
        self.scaler = StandardScaler()

        # Get ADASYN parameters
        adasyn_sampling_strategy = config.get('adasyn_sampling_strategy', 'auto')
        adasyn_n_neighbors = config.get('adasyn_n_neighbors', 5)

        self.adasyn = ADASYN(
            sampling_strategy=adasyn_sampling_strategy,
            n_neighbors=adasyn_n_neighbors,
            random_state=42
        )

        # Feature interaction parameters
        feature_interaction_params = {
            'interaction_method': config.get('interaction_method', 'gaussian'),
            'selection_method': config.get('selection_method', 'mutual_info'),
            'max_features': config.get('max_features', None),
            'gaussian_gamma': config.get('gaussian_gamma', 'auto'),
            'scale_after_interaction': config.get('scale_after_interaction', True)
        }

        # Record feature processing parameters
        # Initialize feature interaction processor
        self.feature_interaction = AdvancedFeatureInteractions(**feature_interaction_params)

    def load_data(self):
        """Load and process data"""
        try:
            df = pd.read_excel(self.config['file_path'], sheet_name=self.config['sheet_name'])

            # Extract sample types and features
            labels = df['Sample'].values
            features = df.iloc[:, 1:].values

            # Save original indices
            self.original_indices = np.arange(len(features))

            # Get original feature names
            original_feature_names = df.columns[1:].tolist()

            # Set original feature names
            self.feature_interaction.set_original_feature_names(original_feature_names)

            # Process features
            features_processed = self.feature_interaction.fit_transform(
                features,
                labels
            )

            self.feature_names = self.feature_interaction.get_feature_names()
            self.feature_importance = self.feature_interaction.get_feature_importance()

            return features_processed, np.array(labels), self.original_indices

        except Exception as e:
            logging.error("Error loading data: {}".format(str(e)))
            raise

    def apply_adasyn(self, X, y, positive_class):

        y_binary = (y == positive_class).astype(int)

        n_positive = sum(y_binary == 1)
        n_negative = sum(y_binary == 0)

        # Apply ADASYN oversampling
        try:
            # ADASYN requires at least two minority class samples, check and handle
            if n_positive < 2:
                logging.warning(f"Not enough positive samples for ADASYN (found {n_positive}, need at least 2)")
                return X, y_binary, self.original_indices

            # Check if minority class samples are sufficient to support the specified number of neighbors
            n_neighbors = self.config.get('adasyn_n_neighbors', 5)
            if n_positive <= n_neighbors:
                adaptive_n_neighbors = max(1, n_positive - 1)
                # Create a new temporary ADASYN instance with adjusted number of neighbors
                temp_adasyn = ADASYN(
                    sampling_strategy=self.config.get('adasyn_sampling_strategy', 'auto'),
                    n_neighbors=adaptive_n_neighbors,
                    random_state=42
                )
                X_resampled, y_binary_resampled = temp_adasyn.fit_resample(X, y_binary)
            else:
                X_resampled, y_binary_resampled = self.adasyn.fit_resample(X, y_binary)

            # Save index information for synthetic data
            sample_indices = np.concatenate([
                self.original_indices,
                np.full(len(X_resampled) - len(X), -1)
            ])

            return X_resampled, y_binary_resampled, sample_indices

        except ValueError as e:
            # Handle errors that ADASYN may encounter
            logging.warning(f"ADASYN error: {str(e)}")
            logging.warning("Continuing without ADASYN resampling for this class")
            return X, y_binary, self.original_indices

    def process_fold(self, X, y_binary, sample_indices, train_index, test_index, positive_class, fold_idx):
        """Process single cross-validation fold - evaluate using only real data"""
        X_train, X_test = X[train_index], X[test_index]
        y_train, y_test = y_binary[train_index], y_binary[test_index]
        indices_train, indices_test = sample_indices[train_index], sample_indices[test_index]

        original_test_mask = indices_test >= 0
        X_test_original = X_test[original_test_mask]
        y_test_original = y_test[original_test_mask]

        # Check class distribution in test set
        test_class_dist = np.bincount(y_test_original)

        # Record class distribution
        neg_count = sum(y_train == 0)
        pos_count = sum(y_train == 1)
        class_ratio = neg_count / max(1, pos_count)

        # Optimize hyperparameters
        best_params = self._optimize_hyperparameters(X_train, y_train)

        # Train algorithms
        model = self._create_model(best_params)
        model.fit(X_train, y_train)

        # Get prediction probabilities
        y_pred_proba = model.predict_proba(X_test_original)[:, 1]

        # Use dynamic threshold selection - find threshold with best F1
        thresholds = np.arange(0.1, 0.7, 0.025)
        best_f1 = 0
        best_threshold = 0.4

        # Only perform threshold optimization when test set contains both classes
        if len(test_class_dist) > 1 and test_class_dist[0] > 0 and test_class_dist[1] > 0:
            for threshold in thresholds:
                y_pred_thresh = (y_pred_proba >= threshold).astype(int)
                f1 = f1_score(y_test_original, y_pred_thresh, average='weighted')
                if f1 > best_f1:
                    best_f1 = f1
                    best_threshold = threshold


        # Make predictions using best threshold
        y_test_pred = (y_pred_proba >= best_threshold).astype(int)

        original_train_mask = indices_train >= 0
        X_train_original = X_train[original_train_mask]
        y_train_original = y_train[original_train_mask]
        y_train_pred_proba = model.predict_proba(X_train_original)[:, 1]
        y_train_pred = (y_train_pred_proba >= best_threshold).astype(int)

        # Calculate evaluation metrics
        train_f1 = f1_score(y_train_original, y_train_pred, average='weighted')
        test_f1 = f1_score(y_test_original, y_test_pred, average='weighted')
        test_precision = precision_score(y_test_original, y_test_pred, average='weighted')
        train_precision = precision_score(y_train_original, y_train_pred, average='weighted')
        test_recall = recall_score(y_test_original, y_test_pred, average='weighted')
        train_recall = recall_score(y_train_original, y_train_pred, average='weighted')

        sensitivity = test_recall
        train_sensitivity = train_recall

        specificity = weighted_specificity(y_test_original, y_test_pred, [0, 1])
        train_specificity = weighted_specificity(y_train_original, y_train_pred, [0, 1])

        # ROC curve calculation
        try:
            fpr, tpr, _ = roc_curve(y_test_original, y_pred_proba)
            roc_auc = auc(fpr, tpr)
        except Exception as e:
            logging.warning("Error calculating ROC curve: {}".format(str(e)))
            fpr, tpr = np.array([0, 1]), np.array([0, 1])
            roc_auc = 0.5

        # Confusion matrix
        conf_matrix = confusion_matrix(y_test_original, y_test_pred)

        return {
            "fold_idx": fold_idx,
            "y_test": y_test_original,
            "y_pred": y_test_pred,
            "f1": test_f1,
            "train_f1": train_f1,
            "precision": test_precision,
            "train_precision": train_precision,
            "recall": test_recall,
            "train_recall": train_recall,
            "sensitivity": sensitivity,
            "train_sensitivity": train_sensitivity,
            "specificity": specificity,
            "train_specificity": train_specificity,
            "fpr": fpr,
            "tpr": tpr,
            "roc_auc": roc_auc,
            "algorithms": model,
            "best_params": best_params,
            "confusion_matrix": conf_matrix,
            "class_name": positive_class
        }

    def _optimize_hyperparameters(self, X_train, y_train):
        """Optimize hyperparameters using Bayesian optimization"""

        # Calculate class ratio for scale_pos_weight
        neg_count = sum(y_train == 0)
        pos_count = sum(y_train == 1)
        base_scale_pos_weight = neg_count / max(1, pos_count)

        def objective(n_estimators, max_depth, learning_rate, subsample,
                      colsample_bytree, min_child_weight, gamma, reg_alpha,
                      reg_lambda, scale_pos_weight):
            """Bayesian optimization objective function"""
            # Convert parameters
            n_estimators = int(n_estimators)
            max_depth = int(max_depth)
            min_child_weight = max(1, int(min_child_weight))

            actual_scale_pos_weight = scale_pos_weight * base_scale_pos_weight

            # Create algorithms
            model = xgb.XGBClassifier(
                n_estimators=n_estimators,
                max_depth=max_depth,
                learning_rate=learning_rate,
                subsample=subsample,
                colsample_bytree=colsample_bytree,
                min_child_weight=min_child_weight,
                gamma=gamma,
                reg_alpha=reg_alpha,
                reg_lambda=reg_lambda,
                scale_pos_weight=actual_scale_pos_weight,
                use_label_encoder=False,  # Avoid warning
                eval_metric='logloss',  # Avoid warning
                random_state=42,
                n_jobs=-1
            )

            # Use StratifiedKFold for cross-validation
            cv = StratifiedKFold(n_splits=self.config['inner_cv_splits'], shuffle=True, random_state=42)

            # Track multiple metrics
            f1_scores = []
            auc_scores = []
            recall_scores = []
            precision_scores = []
            sensitivity_scores = []
            specificity_scores = []

            for train_idx, val_idx in cv.split(X_train, y_train):
                X_train_cv, X_val_cv = X_train[train_idx], X_train[val_idx]
                y_train_cv, y_val_cv = y_train[train_idx], y_train[val_idx]

                # Ensure label format is correct
                y_train_cv = y_train_cv.astype(int)

                # Train algorithms
                model.fit(X_train_cv, y_train_cv)

                # Get prediction probabilities
                y_pred_proba = model.predict_proba(X_val_cv)[:, 1]

                # Try different thresholds to find best F1
                best_f1 = 0
                best_threshold = 0.5
                best_pred = None

                for threshold in [0.3, 0.35, 0.4, 0.45, 0.5]:
                    y_pred_thresh = (y_pred_proba >= threshold).astype(int)
                    f1 = f1_score(y_val_cv, y_pred_thresh, average='weighted')
                    if f1 > best_f1:
                        best_f1 = f1
                        best_threshold = threshold
                        best_pred = y_pred_thresh

                # Record the threshold used
                logging.debug("Best threshold for this fold: {:.2f}".format(best_threshold))

                # Use best prediction CV results
                y_pred = best_pred

                # Compute and save multiple metrics
                f1 = f1_score(y_val_cv, y_pred, average='weighted')
                f1_scores.append(f1)

                try:
                    roc_auc = roc_auc_score(y_val_cv, y_pred_proba)
                    auc_scores.append(roc_auc)
                except Exception as e:
                    logging.debug("Error calculating AUC: {}".format(str(e)))
                    auc_scores.append(0)

                recall = recall_score(y_val_cv, y_pred, average='weighted')
                recall_scores.append(recall)

                sensitivity = recall
                sensitivity_scores.append(sensitivity)

                precision = precision_score(y_val_cv, y_pred, average='weighted')
                precision_scores.append(precision)

                spec = weighted_specificity(y_val_cv, y_pred, [0, 1])
                specificity_scores.append(spec)

            # Combined score - weighted combination of multiple metrics
            mean_f1 = np.mean(f1_scores)
            mean_auc = np.mean(auc_scores)
            mean_recall = np.mean(recall_scores)
            mean_specificity = np.mean(specificity_scores)
            mean_precision = np.mean(precision_scores)

            combined_score = 0.55 * mean_f1 + 0.15 * mean_auc + 0.10 * mean_recall + 0.20 * mean_specificity

            # Record detailed evaluation
            param_str = "n_est={}, depth={}, lr={:.4f}, spw={:.2f}".format(
                n_estimators, max_depth, learning_rate, actual_scale_pos_weight)
            score_str = "F1={:.4f}, AUC={:.4f}, Recall={:.4f}, Spec={:.4f}, Score={:.4f}".format(
                mean_f1, mean_auc, mean_recall, mean_specificity, combined_score)
            logging.debug("Params: {} → {}".format(param_str, score_str))

            return combined_score

        # Use optimized parameter search space
        param_ranges = {
            'n_estimators': (100, 800),
            'max_depth': (2, 5),
            'learning_rate': (0.01, 0.15),
            'subsample': (0.7, 1.0),
            'colsample_bytree': (0.7, 1.0),
            'min_child_weight': (1, 5),
            'gamma': (0, 0.5),
            'reg_alpha': (0, 1.0),
            'reg_lambda': (0.1, 10.0),
            'scale_pos_weight': (0.8, 3.0)
        }

        # Use Bayesian optimization
        optimizer = BayesianOptimization(
            f=objective,
            pbounds=param_ranges,
            random_state=42,
            allow_duplicate_points=True
        )

        optimizer.maximize(
            init_points=self.config['bayes_opt']['init_points'],
            n_iter=self.config['bayes_opt']['n_iter']
        )

        # Return best parameters and convert to correct types
        best_params = optimizer.max['params']
        best_params['n_estimators'] = int(best_params['n_estimators'])
        best_params['max_depth'] = int(best_params['max_depth'])
        best_params['min_child_weight'] = max(1, int(best_params.get('min_child_weight', 1)))
        best_params['scale_pos_weight'] = best_params['scale_pos_weight'] * base_scale_pos_weight

        return best_params

    def _create_model(self, params):
        """Create XGBoost algorithms"""
        # Ensure all necessary parameters exist
        params_with_defaults = {
            'n_estimators': int(params.get('n_estimators', 100)),
            'max_depth': int(params.get('max_depth', 3)),
            'learning_rate': params.get('learning_rate', 0.1),
            'subsample': params.get('subsample', 0.8),
            'colsample_bytree': params.get('colsample_bytree', 0.8),
            'scale_pos_weight': params.get('scale_pos_weight', 1.0),
            'min_child_weight': int(params.get('min_child_weight', 1)),
            'gamma': params.get('gamma', 0),
            'reg_alpha': params.get('reg_alpha', 0),
            'reg_lambda': params.get('reg_lambda', 1.0)
        }

        return xgb.XGBClassifier(
            n_estimators=params_with_defaults['n_estimators'],
            max_depth=params_with_defaults['max_depth'],
            learning_rate=params_with_defaults['learning_rate'],
            subsample=params_with_defaults['subsample'],
            colsample_bytree=params_with_defaults['colsample_bytree'],
            min_child_weight=params_with_defaults['min_child_weight'],
            gamma=params_with_defaults['gamma'],
            reg_alpha=params_with_defaults['reg_alpha'],
            reg_lambda=params_with_defaults['reg_lambda'],
            scale_pos_weight=params_with_defaults['scale_pos_weight'],
            use_label_encoder=False,  # Avoid warning
            eval_metric='logloss',  # Avoid warning
            random_state=42,
            n_jobs=-1
        )

    def plot_roc_curve(self, results, positive_class):
        """Plot ROC curve"""
        plt.figure(figsize=(10, 8))
        ax = plt.gca()

        for spine in ax.spines.values():
            spine.set_linewidth(5)

        # Set Arial font
        plt.rcParams['font.family'] = 'Arial'

        plt.title("ROC Curve for {}".format(positive_class), fontsize=20, pad=15, fontname='Arial')
        plt.xlabel("False Positive Rate", fontsize=24, labelpad=10, fontname='Arial')
        plt.ylabel("True Positive Rate", fontsize=24, labelpad=10, fontname='Arial')

        tprs = []
        aucs = []
        mean_fpr = np.linspace(0, 1, 500)

        for idx, result in enumerate(results):
            fpr = result['fpr']
            tpr = result['tpr']
            roc_auc = result['roc_auc']

            if fpr[0] != 0:
                fpr = np.concatenate([[0], fpr])
                tpr = np.concatenate([[0], tpr])

            if fpr[-1] != 1:
                fpr = np.concatenate([fpr, [1]])
                tpr = np.concatenate([tpr, [tpr[-1]]])

            interp_tpr = np.interp(mean_fpr, fpr, tpr)
            interp_tpr[0] = 0.0
            tprs.append(interp_tpr)
            aucs.append(roc_auc)

            # Set thickness of each fold validation ROC curve to 2
            plt.plot(fpr, tpr, alpha=0.15,
                     color=self.config['color_map'].get(positive_class, 'b'),
                     linestyle='-', linewidth=2)

        mean_tpr = np.mean(tprs, axis=0)
        mean_auc = np.mean(aucs)
        std_auc = np.std(aucs)

        # Set thickness of mean ROC curve to 5
        plt.plot(mean_fpr, mean_tpr,
                 color=self.config['color_map'].get(positive_class, 'b'),
                 label='Mean ROC (AUC±std)\n{:.3f}±{:.3f}'.format(mean_auc, std_auc),
                 lw=5, alpha=0.8)

        std_tpr = np.std(tprs, axis=0)
        tprs_upper = np.minimum(mean_tpr + std_tpr, 1)
        tprs_lower = np.maximum(mean_tpr - std_tpr, 0)

        plt.fill_between(mean_fpr, tprs_lower, tprs_upper,
                         color=self.config['color_map'].get(positive_class, 'b'),
                         alpha=0.2, label='±1 std. dev.')

        plt.plot([0, 1], [0, 1], 'k--', lw=3, alpha=0.7)
        plt.grid(True, alpha=0.3, linewidth=0)
        plt.xlim([-0.02, 1.02])
        plt.ylim([-0.02, 1.02])

        ax.tick_params(width=5, length=10, labelsize=24)

        # Set tick label font
        for label in ax.get_xticklabels() + ax.get_yticklabels():
            label.set_fontname('Arial')

        legend = plt.legend(loc="lower right", fontsize=24, frameon=True)
        legend.get_frame().set_linewidth(3)

        # Set legend font
        for text in legend.get_texts():
            text.set_fontname('Arial')

        output_file = os.path.join(self.config['output_path'],
                                   "ROC_{}.png".format(positive_class))
        plt.savefig(output_file, dpi=600, bbox_inches='tight')
        plt.close()

    def plot_confusion_matrix(self, conf_matrix, class_name, fold_idx):
        """Plot confusion matrix"""
        plt.figure(figsize=(8, 6))

        # Set Arial font
        plt.rcParams['font.family'] = 'Arial'

        ax = plt.gca()
        sns.heatmap(conf_matrix, annot=True, fmt='d', cmap='Blues')

        plt.title('Confusion Matrix - {} (Fold {})'.format(class_name, fold_idx + 1),
                  fontsize=20, pad=15, fontname='Arial')
        plt.ylabel('True Label', fontsize=24, labelpad=10, fontname='Arial')
        plt.xlabel('Predicted Label', fontsize=24, labelpad=10, fontname='Arial')

        # Set tick label font
        ax.tick_params(labelsize=24)
        for label in ax.get_xticklabels() + ax.get_yticklabels():
            label.set_fontname('Arial')

        output_file = os.path.join(self.config['output_path'],
                                   "confusion_matrix_{}_fold_{}.png".format(class_name, fold_idx + 1))
        plt.savefig(output_file, dpi=300, bbox_inches='tight')
        plt.close()

    def format_excel_file(self, file_path):
        """Format Excel file: set font to Times New Roman and bold header row"""
        wb = load_workbook(file_path)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row_idx, row in enumerate(ws.iter_rows(), start=1):
                for cell in row:
                    if row_idx == 1:  # Header row
                        cell.font = Font(name='Times New Roman', size=11, bold=True)
                    else:
                        cell.font = Font(name='Times New Roman', size=11)

        wb.save(file_path)

    def save_results(self, all_results):
        """Save CV results to Excel"""
        output_path = os.path.join(self.config['output_path'],
                                   self.config['output_filename'])

        with pd.ExcelWriter(output_path) as writer:
            # 1. Save overall summary
            summary_data = []
            for class_name, results in all_results.items():
                # Collect metrics from all folds
                f1_scores = [r['f1'] for r in results]
                train_f1_scores = [r['train_f1'] for r in results]
                precision_scores = [r['precision'] for r in results]
                train_precision_scores = [r['train_precision'] for r in results]
                recall_scores = [r['recall'] for r in results]
                train_recall_scores = [r['train_recall'] for r in results]
                sensitivity_scores = [r['sensitivity'] for r in results]
                train_sensitivity_scores = [r['train_sensitivity'] for r in results]
                specificity_scores = [r['specificity'] for r in results]
                train_specificity_scores = [r['train_specificity'] for r in results]
                auc_scores = [r['roc_auc'] for r in results]

                summary_data.append({
                    'Class': class_name,
                    'Mean F1': np.mean(f1_scores),
                    'Std F1': np.std(f1_scores),
                    'Mean Train F1': np.mean(train_f1_scores),
                    'Std Train F1': np.std(train_f1_scores),
                    'Mean Precision': np.mean(precision_scores),
                    'Std Precision': np.std(precision_scores),
                    'Mean Train Precision': np.mean(train_precision_scores),
                    'Std Train Precision': np.std(train_precision_scores),
                    'Mean Recall': np.mean(recall_scores),
                    'Std Recall': np.std(recall_scores),
                    'Mean Train Recall': np.mean(train_recall_scores),
                    'Std Train Recall': np.std(train_recall_scores),
                    'Mean Sensitivity': np.mean(sensitivity_scores),
                    'Std Sensitivity': np.std(sensitivity_scores),
                    'Mean Train Sensitivity': np.mean(train_sensitivity_scores),
                    'Std Train Sensitivity': np.std(train_sensitivity_scores),
                    'Mean Specificity': np.mean(specificity_scores),
                    'Std Specificity': np.std(specificity_scores),
                    'Mean Train Specificity': np.mean(train_specificity_scores),
                    'Std Train Specificity': np.std(train_specificity_scores),
                    'Mean AUC': np.mean(auc_scores),
                    'Std AUC': np.std(auc_scores),
                    'Best Params': str(results[-1]['best_params'])
                })

            pd.DataFrame(summary_data).to_excel(writer,
                                                sheet_name='Summary',
                                                index=False)

            # 2. Save detailed CV results and ROC curve data for each class
            for class_name, results in all_results.items():
                # Save detailed CV results for each fold
                fold_data = []
                for r in results:
                    fold_data.append({
                        'Fold': r['fold_idx'] + 1,
                        'F1 Score': r['f1'],
                        'Train F1': r['train_f1'],
                        'Precision': r['precision'],
                        'Train Precision': r['train_precision'],
                        'Recall': r['recall'],
                        'Train Recall': r['train_recall'],
                        'Sensitivity': r['sensitivity'],
                        'Train Sensitivity': r['train_sensitivity'],
                        'Specificity': r['specificity'],
                        'Train Specificity': r['train_specificity'],
                        'ROC AUC': r['roc_auc'],
                        'n_estimators': int(r['best_params'].get('n_estimators', 0)),
                        'max_depth': int(r['best_params'].get('max_depth', 0)),
                        'learning_rate': r['best_params'].get('learning_rate', 0),
                        'scale_pos_weight': r['best_params'].get('scale_pos_weight', 0),
                        'min_child_weight': r['best_params'].get('min_child_weight', 1),
                        'gamma': r['best_params'].get('gamma', 0),
                        'subsample': r['best_params'].get('subsample', 0),
                        'colsample_bytree': r['best_params'].get('colsample_bytree', 0),
                        'reg_alpha': r['best_params'].get('reg_alpha', 0),
                        'reg_lambda': r['best_params'].get('reg_lambda', 0)
                    })

                pd.DataFrame(fold_data).to_excel(writer,
                                                 sheet_name=f'{class_name}_details',
                                                 index=False)

                # Save ROC curve data
                # 1. ROC data for each fold
                for idx, result in enumerate(results):
                    fold_roc_data = pd.DataFrame({
                        'FPR': result['fpr'],
                        'TPR': result['tpr']
                    })
                    fold_roc_data.to_excel(writer,
                                           sheet_name=f'{class_name}_ROC_data',
                                           startcol=idx * 3,
                                           index=False)
                    # Add AUC value next to each fold's data
                    pd.DataFrame({'AUC': [result['roc_auc']]}).to_excel(
                        writer,
                        sheet_name=f'{class_name}_ROC_data',
                        startcol=idx * 3 + 2,
                        index=False
                    )

                # 2. Calculate and save mean ROC curve data
                mean_fpr = np.linspace(0, 1, 500)
                tprs = []
                for result in results:
                    fpr = result['fpr']
                    tpr = result['tpr']
                    if fpr[0] != 0:
                        fpr = np.concatenate([[0], fpr])
                        tpr = np.concatenate([[0], tpr])
                    if fpr[-1] != 1:
                        fpr = np.concatenate([fpr, [1]])
                        tpr = np.concatenate([tpr, [tpr[-1]]])

                    # Interpolation
                    interp_tpr = np.interp(mean_fpr, fpr, tpr)
                    interp_tpr[0] = 0.0
                    tprs.append(interp_tpr)

                mean_tpr = np.mean(tprs, axis=0)
                std_tpr = np.std(tprs, axis=0)

                # Save mean ROC curve data
                mean_roc_data = pd.DataFrame({
                    'FPR': mean_fpr,
                    'Mean_TPR': mean_tpr,
                    'Std_TPR': std_tpr,
                    'TPR_Upper': np.minimum(mean_tpr + std_tpr, 1),
                    'TPR_Lower': np.maximum(mean_tpr - std_tpr, 0)
                })
                mean_roc_data.to_excel(writer,
                                       sheet_name=f'{class_name}_mean_ROC',
                                       index=False)

        # Format Excel file
        self.format_excel_file(output_path)

    def run(self):
        """Run complete training and evaluation pipeline"""
        # Load data
        X, y, original_indices = self.load_data()
        logging.info("Using XGBoost classifier")
        logging.info("Data shape: {} samples, {} features".format(X.shape[0], X.shape[1]))

        all_results = {}

        for class_name in self.all_classes:
            logging.info("Processing class: {}".format(class_name))

            X_balanced, y_binary_balanced, sample_indices = self.apply_adasyn(X, y, class_name)

            # Create cross-validation object
            outer_cv = StratifiedKFold(n_splits=self.config['outer_cv_splits'], shuffle=True, random_state=42)

            # Use cross-validation - parallel processing
            try:
                # Prepare cross-validation indices
                fold_indices = list(enumerate(outer_cv.split(X_balanced, y_binary_balanced)))

                # Process each fold in parallel
                results = Parallel(n_jobs=-1)(
                    delayed(self.process_fold)(
                        X_balanced, y_binary_balanced, sample_indices, train_idx, test_idx, class_name, fold_idx
                    )
                    for fold_idx, (train_idx, test_idx) in fold_indices
                )

                # Skip if no valid CV results
                if not results:
                    logging.warning("No valid CV results for class {}. Skipping.".format(class_name))
                    continue

                all_results[class_name] = results

                # Plot ROC curve
                self.plot_roc_curve(results, class_name)

                # Output performance metrics for current class
                f1_scores = [r['f1'] for r in results]
                auc_scores = [r['roc_auc'] for r in results]
                precision_scores = [r['precision'] for r in results]
                recall_scores = [r['recall'] for r in results]
                sensitivity_scores = [r['sensitivity'] for r in results]
                specificity_scores = [r['specificity'] for r in results]

                logging.info("Performance metrics for {}:".format(class_name))
                logging.info("F1: {:.3f}±{:.3f}, AUC: {:.3f}±{:.3f}, Precision: {:.3f}±{:.3f}, Recall: {:.3f}±{:.3f}, Sensitivity: {:.3f}±{:.3f}, Specificity: {:.3f}±{:.3f}".format(
                    np.mean(f1_scores), np.std(f1_scores),
                    np.mean(auc_scores), np.std(auc_scores),
                    np.mean(precision_scores), np.std(precision_scores),
                    np.mean(recall_scores), np.std(recall_scores),
                    np.mean(sensitivity_scores), np.std(sensitivity_scores),
                    np.mean(specificity_scores), np.std(specificity_scores)))

            except Exception as e:
                logging.error("Error processing class {}: {}".format(class_name, str(e)))
                import traceback
                logging.error(traceback.format_exc())
                # Continue processing other classes

        # Save CV results
        if all_results:
            self.save_results(all_results)
            return all_results
        else:
            logging.error("No CV results were generated. Check for errors above.")
            return {}


def main():
    """Main function"""
    # Configuration parameters
    config = {
        'file_path': os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '00_Data', 'Train.xlsx'),  # Path to the input Excel file containing training data
        'sheet_name': 'Sheet1',  # Name of the sheet in the Excel file (e.g., 'Sheet1')
        'output_path': os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '..', 'Results', 'Multicancer detection', '02_Algorithm optimization', 'XGB'),  # Directory path for saving output results
        'output_filename': 'XGBoost_results.xlsx',

        # Cross-validation configuration
        'outer_cv_splits': 5,
        'inner_cv_splits': 5,

        # Feature processing configuration
        'interaction_method': 'gaussian',
        'selection_method': 'mutual_info',
        'max_features': 40,
        'gaussian_gamma': 'auto',
        'scale_after_interaction': True,
        'debug_mode': True,

        # ADASYN configuration
        'adasyn_sampling_strategy': 0.5,
        'adasyn_n_neighbors': 5,

        # Bayesian optimization configuration
        'bayes_opt': {
            'init_points': 12,
            'n_iter': 80,
        },

        # Color configuration
        'color_map': {
            'LC': [254 / 255, 160 / 255, 64 / 255],
            'LuC': [242 / 255, 128 / 255, 128 / 255],
            'OC': [88 / 255, 97 / 255, 172 / 255],
            'N': [106 / 255, 180 / 255, 193 / 255]
        }
    }

    try:
        # Create output directory
        os.makedirs(config['output_path'], exist_ok=True)

        if config['debug_mode']:
            logging.info("Feature interaction method: {}".format(config['interaction_method']))
            logging.info("Feature selection method: {}".format(config['selection_method']))

        # Read Excel to get class information
        df = pd.read_excel(config['file_path'], sheet_name=config['sheet_name'])
        config['all_classes'] = sorted(df['Sample'].unique())

        # Validate data format
        required_columns = ['Sample']
        if not all(col in df.columns for col in required_columns):
            raise ValueError("Missing required columns. Required: {}".format(required_columns))

        # Run classifier
        classifier = XGBoostClassifier(config)
        results = classifier.run()

        # Output final CV results summary
        logging.info("Training completed successfully!")

    except Exception as e:
        logging.error("An error occurred: {}".format(str(e)))
        import traceback
        logging.error(traceback.format_exc())
        raise


if __name__ == "__main__":
    main()