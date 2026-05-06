import numpy as np
import pandas as pd
from sklearn.preprocessing import StandardScaler
from sklearn.feature_selection import SelectKBest, mutual_info_classif
from sklearn.decomposition import PCA
from scipy import stats
from sklearn.neural_network import MLPClassifier
from sklearn.model_selection import StratifiedKFold, cross_val_score, cross_val_predict
from sklearn.metrics import (f1_score, roc_curve, auc, confusion_matrix, precision_score,
                             recall_score, precision_recall_curve, average_precision_score,
                             roc_auc_score)
from bayes_opt import BayesianOptimization
import matplotlib.pyplot as plt
from joblib import Parallel, delayed
import logging
import os
import seaborn as sns
from itertools import combinations
import warnings
import re
import joblib
import os
import json
from imblearn.over_sampling import SMOTE  # Import SMOTE
from openpyxl import load_workbook
from openpyxl.styles import Font

warnings.filterwarnings('ignore')

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')


class AdvancedFeatureInteractions:
    """Feature interaction processing class"""

    def __init__(self,
                 interaction_method='gaussian',
                 selection_method='boruta',
                 max_features=30,
                 gaussian_gamma='auto',
                 scale_after_interaction=True):
        """
        Initialize feature interaction processor

        Parameters:
            interaction_method: Feature interaction method ('gaussian', 'polynomial', 'custom', 'combined')
            selection_method: Feature selection method ('mutual_info', 'correlation', 'pca', None)
            max_features: Maximum number of features to select
            gaussian_gamma: Gamma parameter for Gaussian kernel, can be 'auto' or specific value
            scale_after_interaction: Whether to perform standardization after interaction
        """
        self.interaction_method = interaction_method
        self.selection_method = selection_method
        self.max_features = max_features
        self.gaussian_gamma = gaussian_gamma
        self.scale_after_interaction = scale_after_interaction

        # Initialize other attributes
        self.feature_names_ = None
        self.selected_features_ = 30
        self.feature_importance_ = None
        self.scaler = StandardScaler()

        # Save scaler and feature selector after interaction
        self.interaction_scaler = None
        self.selector = None

        # Record whether fitting has been done
        self.is_fitted = False

        # Output initialization parameters

    def fit_transform(self, X, y=None):
        """Fit and transform data"""
        try:

            # 1. Standardize original features
            X_scaled = self.scaler.fit_transform(X)

            # 2. Generate interaction features
            if self.interaction_method == 'gaussian':
                X_interaction = self._gaussian_interaction(X_scaled)
            else:
                raise ValueError("Unknown interaction method: {}".format(self.interaction_method))

            # 3. Post-interaction standardization - Save scaler for later use
            if self.scale_after_interaction:
                self.interaction_scaler = StandardScaler()
                X_interaction = self.interaction_scaler.fit_transform(X_interaction)

            # 4. Feature selection
            if self.selection_method is not None and y is not None:
                X_final = self._select_features(X_interaction, y)
            else:
                X_final = X_interaction

            # Mark as fitted
            self.is_fitted = True

            return X_final

        except Exception as e:
            logging.error("Error in fit_transform: {}".format(str(e)))
            raise

    # New: transform method, apply same transformation to new data without refitting
    def transform(self, X):
        """Transform data without refitting parameters"""
        if not self.is_fitted:
            raise ValueError("This AdvancedFeatureInteractions instance is not fitted yet. "
                             "Call 'fit_transform' before using this method.")

        try:

            # 1. Transform data using already fitted scaler
            X_scaled = self.scaler.transform(X)

            # 2. Apply same feature interaction
            if self.interaction_method == 'gaussian':
                X_interaction = self._gaussian_interaction(X_scaled)
            else:
                raise ValueError("Unknown interaction method: {}".format(self.interaction_method))

            # 3. Apply fitted post-interaction standardization
            if self.scale_after_interaction and self.interaction_scaler is not None:
                X_interaction = self.interaction_scaler.transform(X_interaction)

            # 4. Apply existing feature selection
            if self.selection_method is not None and self.selector is not None:
                X_final = self.selector.transform(X_interaction)
            else:
                X_final = X_interaction

            return X_final

        except Exception as e:
            logging.error("Error in transform: {}".format(str(e)))
            raise

    def _gaussian_interaction(self, X):
        """Enhanced Gaussian kernel feature interaction, adding distance features and local density features"""
        n_samples, n_features = X.shape

        # Determine gamma parameter
        if self.gaussian_gamma == 'auto':
            gamma = 1.0 / n_features
        elif isinstance(self.gaussian_gamma, (int, float)):
            gamma = self.gaussian_gamma
        else:
            gamma = 1.0 / n_features


        # Create base feature names
        base_names = ["feature_{}".format(i) for i in range(n_features)]

        # Calculate core features
        core_features = []
        core_names = []

        # 1. Direct features
        core_features.append(X)
        core_names.extend(base_names)

        # 2. Pairwise Gaussian kernel interaction
        for i, j in combinations(range(n_features), 2):
            # Calculate Gaussian kernel between two features
            diff = X[:, i:i + 1] - X[:, j:j + 1]
            kernel = np.exp(-gamma * (diff ** 2))
            core_features.append(kernel)
            core_names.append("gaussian_{}_{}".format(i, j))

            # Weighted combination
            weighted_combo = kernel * (X[:, i:i + 1] + X[:, j:j + 1])
            core_features.append(weighted_combo)
            core_names.append("weighted_gaussian_{}_{}".format(i, j))

        # 3. Distance features
        for i in range(n_features):
            # Calculate sum of RBF kernels between current feature and all other features
            rbf_sum = np.zeros((n_samples, 1))
            for j in range(n_features):
                if i != j:
                    diff = X[:, i:i + 1] - X[:, j:j + 1]
                    rbf_sum += np.exp(-gamma * (diff ** 2))
            core_features.append(rbf_sum)
            core_names.append("rbf_distance_{}".format(i))

        # 4. Local density features
        for i in range(n_features):
            # Calculate local density estimation
            density = np.zeros((n_samples, 1))
            for j in range(n_features):
                diff = X[:, i:i + 1] - X[:, j:j + 1]
                density += np.exp(-gamma * np.sum(diff ** 2, axis=1, keepdims=True))
            core_features.append(density)
            core_names.append("local_density_{}".format(i))

        # Combine all features
        X_gaussian = np.hstack(core_features)
        self.feature_names_ = core_names

        return X_gaussian

    def _select_features(self, X, y):
        """Enhanced feature selection, supporting multiple nonlinear methods"""
        # Import necessary libraries within the method
        from sklearn.feature_selection import SelectKBest, mutual_info_classif


        if self.selection_method == 'mutual_info':
            # Original mutual information selection
            if self.max_features is None:
                n_features = X.shape[1]
            else:
                n_features = min(self.max_features, X.shape[1])

            self.selector = SelectKBest(
                score_func=mutual_info_classif,
                k=n_features
            )
            X_selected = self.selector.fit_transform(X, y)
            self.feature_importance_ = self.selector.scores_

        elif self.selection_method == 'random_forest':
            # Random forest based feature selection
            from sklearn.ensemble import RandomForestClassifier
            if self.max_features is None:
                n_features = X.shape[1]
            else:
                n_features = min(self.max_features, X.shape[1])

            forest = RandomForestClassifier(n_estimators=100, random_state=42, n_jobs=-1)
            forest.fit(X, y)

            # Get feature importance
            self.feature_importance_ = forest.feature_importances_

            # Select top-k features
            indices = np.argsort(self.feature_importance_)[::-1][:n_features]

            # Create feature selection mask
            selected_mask = np.zeros(X.shape[1], dtype=bool)
            selected_mask[indices] = True

            # Implement selector to maintain API consistency
            from sklearn.feature_selection import SelectFromModel
            self.selector = SelectFromModel(forest, max_features=n_features, prefit=True)
            X_selected = self.selector.transform(X)

        elif self.selection_method == 'boruta':
            # Boruta feature selection
            try:
                from boruta import BorutaPy
                from sklearn.ensemble import RandomForestClassifier

                if self.max_features is None:
                    n_features = X.shape[1]
                else:
                    n_features = min(self.max_features, X.shape[1])


                # Create random forest classifier
                forest = RandomForestClassifier(
                    n_estimators=100,
                    max_depth=5,
                    random_state=42,
                    n_jobs=-1
                )

                # Create Boruta feature selector
                boruta_selector = BorutaPy(
                    estimator=forest,
                    n_estimators='auto',  # Automatically select number of estimators
                    verbose=0,
                    random_state=42
                )

                # Fit data
                # Boruta requires shape [samples, features], no need to transpose if X already has this shape
                boruta_selector.fit(X, y)

                # Select features considered important by Boruta
                X_selected = boruta_selector.transform(X)

                # If number of selected features is less than requested, use support ranking to select more
                if X_selected.shape[1] < n_features:
                    # Select features based on Boruta ranking
                    rank_indices = np.argsort(boruta_selector.ranking_)[:n_features]
                    selected_mask = np.zeros(X.shape[1], dtype=bool)
                    selected_mask[rank_indices] = True
                    X_selected = X[:, selected_mask]

                    # Update Boruta mask
                    boruta_mask = np.zeros(X.shape[1], dtype=bool)
                    boruta_mask[rank_indices] = True
                    boruta_selector.support_ = boruta_mask

                # Save feature importance (inverse ranking)
                self.feature_importance_ = 1.0 / boruta_selector.ranking_
                self.selector = boruta_selector

            except ImportError:
                logging.warning("Boruta not installed. Falling back to mutual_info selection.")
                # Fall back to mutual information selection
                return self._select_features_with_method(X, y, 'mutual_info')

        else:
            X_selected = X
            self.feature_importance_ = np.ones(X.shape[1])

            # Create a selector that does nothing
            class IdentitySelector:
                def transform(self, X):
                    return X

                def get_support(self):
                    return np.ones(X.shape[1], dtype=bool)

            self.selector = IdentitySelector()

        # Update feature names
        if len(self.feature_names_) != X.shape[1]:
            self.feature_names_ = ["feature_{}".format(i) for i in range(X.shape[1])]

        # Save selected feature names
        selected_mask = self.selector.get_support() if hasattr(self.selector, 'get_support') else np.ones(X.shape[1],
                                                                                                          dtype=bool)
        self.selected_features_ = np.array(self.feature_names_)[selected_mask]

        return X_selected

    def get_feature_names(self):
        """Get feature names"""
        if self.selected_features_ is not None:
            return self.selected_features_
        return self.feature_names_

    def get_feature_importance(self):
        """Get feature importance"""
        if self.feature_importance_ is None:
            return None

        importance_df = pd.DataFrame({
            'Feature': self.feature_names_,
            'Importance': self.feature_importance_
        })
        return importance_df.sort_values('Importance', ascending=False)


def weighted_specificity_score(y_true, y_pred):

    # Get all classes
    classes = np.unique(np.concatenate([y_true, y_pred]))

    # Calculate confusion matrix
    cm = confusion_matrix(y_true, y_pred, labels=classes)

    # Initialize variables
    specificity_sum = 0
    total_samples = 0

    for i, cls in enumerate(classes):

        true_neg_mask = np.ones(cm.shape, dtype=bool)
        true_neg_mask[i, :] = False
        true_neg_mask[:, i] = False

        tn = np.sum(cm[true_neg_mask])

        fp = np.sum(cm[:, i]) - cm[i, i]

        if tn + fp > 0:
            specificity = tn / (tn + fp)
        else:
            specificity = 0

        weight = np.sum(y_true == cls)
        specificity_sum += specificity * weight
        total_samples += weight

    return specificity_sum / total_samples if total_samples > 0 else 0


class ANNClassifier:
    def __init__(self, config):
        self.config = config
        self.classifiers = {}
        self.all_classes = config['all_classes']
        self.results_data = {}
        self.inner_cv_results = {}
        self.outer_cv_results = {}
        self.scaler = StandardScaler()

        # Get SMOTE parameters - New
        smote_sampling_strategy = config.get('smote_sampling_strategy', 'auto')
        smote_k_neighbors = config.get('smote_k_neighbors', 5)

        self.smote = SMOTE(
            sampling_strategy=smote_sampling_strategy,
            k_neighbors=smote_k_neighbors,
            random_state=42
        )

        # Feature interaction parameters
        feature_interaction_params = {
            'interaction_method': config.get('interaction_method', 'gaussian'),
            'selection_method': config.get('selection_method', 'mutual_info'),
            'max_features': config.get('max_features', None),
            'gaussian_gamma': config.get('gaussian_gamma', 'auto'),
            'scale_after_interaction': config.get('scale_after_interaction', True)
        }

        # Record feature processing parameters
        # Initialize feature interaction processor
        self.feature_interaction = AdvancedFeatureInteractions(**feature_interaction_params)

    def load_data(self):
        """Load and process data"""
        try:
            df = pd.read_excel(self.config['file_path'], sheet_name=self.config['sheet_name'])

            # Extract sample types and features
            labels = df['Sample'].values
            features = df.iloc[:, 1:14].values

            # Save original indices
            self.original_indices = np.arange(len(features))


            # Process features
            features_processed = self.feature_interaction.fit_transform(
                features,
                labels
            )

            self.feature_names = self.feature_interaction.get_feature_names()
            self.feature_importance = self.feature_interaction.get_feature_importance()

            return features_processed, np.array(labels), self.original_indices

        except Exception as e:
            logging.error("Error loading data: {}".format(str(e)))
            raise

    # Apply SMOTE method
    def apply_smote(self, X, y, positive_class):

        y_binary = (y == positive_class).astype(int)

        n_positive = sum(y_binary == 1)
        n_negative = sum(y_binary == 0)

        # Apply SMOTE oversampling
        try:
            X_resampled, y_binary_resampled = self.smote.fit_resample(X, y_binary)

            # Save synthetic data index information
            sample_indices = np.concatenate([
                self.original_indices,
                np.full(len(X_resampled) - len(X), -1)
            ])

            return X_resampled, y_binary_resampled, sample_indices

        except ValueError as e:
            logging.warning(f"SMOTE error: {str(e)}")
            logging.warning("Continuing without SMOTE resampling for this class")
            return X, y_binary, self.original_indices

    def process_fold(self, X, y_binary, sample_indices, train_index, test_index, positive_class, fold_idx):
        """Process single cross-validation fold - evaluate only with real data"""
        X_train, X_test = X[train_index], X[test_index]
        y_train, y_test = y_binary[train_index], y_binary[test_index]
        indices_train, indices_test = sample_indices[train_index], sample_indices[test_index]

        # Use only original data for test evaluation
        original_test_mask = indices_test >= 0
        X_test_original = X_test[original_test_mask]
        y_test_original = y_test[original_test_mask]

        # Check class distribution in test set
        test_class_dist = np.bincount(y_test_original)

        # If test set contains only one class, log warning
        if len(test_class_dist) == 1 or (len(test_class_dist) > 1 and test_class_dist[1] == 0):
            logging.warning("Test set for fold {} only contains negative samples".format(fold_idx + 1))
        if len(test_class_dist) > 1 and test_class_dist[0] == 0:
            logging.warning("Test set for fold {} only contains positive samples".format(fold_idx + 1))

        # Record class distribution
        neg_count = sum(y_train == 0)
        pos_count = sum(y_train == 1)
        class_ratio = neg_count / max(1, pos_count)

        # Optimize hyperparameters
        best_params = self._optimize_hyperparameters(X_train, y_train)

        # Train algorithms
        model = self._create_model(best_params)
        model.fit(X_train, y_train)

        # Get prediction probabilities
        y_pred_proba = model.predict_proba(X_test_original)[:, 1]

        # Use dynamic threshold selection - Find threshold with best F1
        thresholds = np.arange(0.1, 0.7, 0.025)
        best_f1 = 0
        best_threshold = 0.4

        # Only perform threshold optimization when test set contains two classes
        if len(test_class_dist) > 1 and test_class_dist[0] > 0 and test_class_dist[1] > 0:
            for threshold in thresholds:
                y_pred_thresh = (y_pred_proba >= threshold).astype(int)
                f1 = f1_score(y_test_original, y_pred_thresh, average='weighted')
                if f1 > best_f1:
                    best_f1 = f1
                    best_threshold = threshold


        # Make predictions using best threshold
        y_pred = (y_pred_proba >= best_threshold).astype(int)

        original_train_mask = indices_train >= 0
        X_train_original = X_train[original_train_mask]
        y_train_original = y_train[original_train_mask]
        y_train_pred_proba = model.predict_proba(X_train_original)[:, 1]
        y_train_pred = (y_train_pred_proba >= best_threshold).astype(int)

        # Calculate evaluation metrics
        train_f1 = f1_score(y_train_original, y_train_pred, average='weighted')
        test_f1 = f1_score(y_test_original, y_pred, average='weighted')
        test_precision = precision_score(y_test_original, y_pred, average='weighted')
        test_recall = recall_score(y_test_original, y_pred, average='weighted')

        sensitivity = recall_score(y_test_original, y_pred, average='weighted')

        specificity = weighted_specificity_score(y_test_original, y_pred)

        # ROC curve calculation
        try:
            fpr, tpr, _ = roc_curve(y_test_original, y_pred_proba)
            roc_auc = auc(fpr, tpr)
        except Exception as e:
            logging.warning("Error calculating ROC curve: {}".format(str(e)))
            fpr, tpr = np.array([0, 1]), np.array([0, 1])
            roc_auc = 0.5

        # PR curve calculation
        try:
            precision, recall, _ = precision_recall_curve(y_test_original, y_pred_proba)
            pr_auc = average_precision_score(y_test_original, y_pred_proba)
        except Exception as e:
            logging.warning("Error calculating PR curve: {}".format(str(e)))
            precision, recall = np.array([0, 1]), np.array([1, 0])
            pr_auc = 0.5

        # Confusion matrix
        conf_matrix = confusion_matrix(y_test_original, y_pred)

        # Calculate false negative rate and false positive rate
        tn, fp, fn, tp = 0, 0, 0, 0
        if conf_matrix.shape == (2, 2):
            tn, fp, fn, tp = conf_matrix.ravel()
        elif conf_matrix.shape == (1, 1):
            # Only one class
            if np.sum(y_test_original) == 0:
                tn = conf_matrix[0, 0]
            else:  # Only positive class
                tp = conf_matrix[0, 0]

        # Calculate false negative rate and false positive rate
        fnr = fn / (fn + tp) if (fn + tp) > 0 else 0
        fpr_val = fp / (fp + tn) if (fp + tn) > 0 else 0

        # Get feature importance (if available)
        feature_importance = None
        if hasattr(model, 'feature_importances_'):
            feature_importance = model.feature_importances_
        elif hasattr(model, 'coef_'):
            feature_importance = model.coef_[0] if model.coef_.ndim > 1 else model.coef_

        return {
            "fold_idx": fold_idx,
            "y_test": y_test_original,
            "y_pred": y_pred,
            "y_pred_proba": y_pred_proba,
            "optimal_threshold": best_threshold,
            "f1": test_f1,
            "precision": test_precision,
            "recall": test_recall,
            "sensitivity": sensitivity,
            "specificity": specificity,
            "train_f1": train_f1,
            "fpr": fpr,
            "tpr": tpr,
            "roc_auc": roc_auc,
            "pr_curve": {"precision": precision, "recall": recall},
            "pr_auc": pr_auc,
            "fnr": fnr,  # False negative rate
            "fpr_val": fpr_val,  # False positive rate
            "algorithms": model,
            "best_params": best_params,
            "confusion_matrix": conf_matrix,
            "feature_importance": feature_importance
        }

    def _get_hidden_layer_sizes(self, hidden_factor, neuron_params):
        """Generate hidden layer configuration based on hidden layer factor and neuron parameters"""

        # Get number of neurons in each layer (round to integer)
        first_layer = int(neuron_params['first_layer_neurons'])
        second_layer = int(neuron_params['second_layer_neurons'])
        third_layer = int(neuron_params['third_layer_neurons'])
        fourth_layer = int(neuron_params['fourth_layer_neurons'])

        # Determine number of layers to use based on hidden_factor
        if hidden_factor < 1:
            return (first_layer,)
        elif hidden_factor < 2:
            return (first_layer, second_layer)
        elif hidden_factor < 3:
            return (first_layer, second_layer, third_layer)
        else:
            return (first_layer, second_layer, third_layer, fourth_layer)

    def _optimize_hyperparameters(self, X_train, y_train):
        """Optimize hyperparameters using Bayesian optimization - For ANN"""
        # Get input feature dimension
        n_features = X_train.shape[1]

        def objective(**params):
            """Bayesian optimization objective function"""
            # Get and convert parameters
            hidden_layer_factor = float(params['hidden_layer_factor'])
            activation_idx = int(params['activation_idx'])
            solver_idx = int(params['solver_idx'])
            learning_rate_idx = int(params['learning_rate_idx'])
            batch_size = int(min(X_train.shape[0], params['batch_size_factor'] * 16))

            # Create neuron parameters dictionary
            neuron_params = {
                'first_layer_neurons': params['first_layer_neurons'],
                'second_layer_neurons': params['second_layer_neurons'],
                'third_layer_neurons': params['third_layer_neurons'],
                'fourth_layer_neurons': params['fourth_layer_neurons']
            }

            # Get hidden layer configuration
            hidden_layer_sizes = self._get_hidden_layer_sizes(hidden_layer_factor, neuron_params)

            # Activation functions, solvers and learning rate types
            activations = ['identity', 'logistic', 'tanh', 'relu']
            solvers = ['lbfgs', 'sgd', 'adam']
            learning_rates = ['constant', 'invscaling', 'adaptive']

            # Get corresponding string values
            activation = activations[activation_idx % len(activations)]
            solver = solvers[solver_idx % len(solvers)]
            learning_rate = learning_rates[learning_rate_idx % len(learning_rates)]

            # Create algorithms
            model = MLPClassifier(
                hidden_layer_sizes=hidden_layer_sizes,
                activation=activation,
                solver=solver,
                alpha=params['alpha'],
                batch_size=batch_size,
                learning_rate=learning_rate,
                learning_rate_init=params['learning_rate_init'],
                max_iter=500,
                early_stopping=True,
                validation_fraction=0.2,
                n_iter_no_change=10,
                random_state=42
            )

            # Use StratifiedKFold for cross-validation
            cv = StratifiedKFold(n_splits=self.config['inner_cv_splits'], shuffle=True, random_state=42)

            # Track multiple metrics
            f1_scores = []
            auc_scores = []
            recall_scores = []
            precision_scores = []
            sensitivity_scores = []
            specificity_scores = []

            # Wrap entire cross-validation process with try-except to avoid interruption due to algorithms failure
            try:
                for train_idx, val_idx in cv.split(X_train, y_train):
                    X_train_cv, X_val_cv = X_train[train_idx], X_train[val_idx]
                    y_train_cv, y_val_cv = y_train[train_idx], y_train[val_idx]

                    # Ensure label format is correct
                    y_train_cv = y_train_cv.astype(int)

                    # Train algorithms and capture warnings
                    with warnings.catch_warnings():
                        warnings.simplefilter("ignore")
                        try:
                            model.fit(X_train_cv, y_train_cv)
                        except Exception as e:
                            logging.debug(f"Model fitting failed: {str(e)}")
                            # Return a very low score
                            return 0.0

                    # Get prediction probabilities
                    try:
                        y_pred_proba = model.predict_proba(X_val_cv)[:, 1]
                    except Exception as e:
                        logging.debug(f"Probability prediction failed: {str(e)}")
                        return 0.0

                    # Try different thresholds to find best F1
                    best_f1 = 0
                    best_threshold = 0.5
                    best_pred = None

                    for threshold in [0.3, 0.35, 0.4, 0.45, 0.5]:
                        y_pred_thresh = (y_pred_proba >= threshold).astype(int)
                        f1 = f1_score(y_val_cv, y_pred_thresh, average='weighted')
                        if f1 > best_f1:
                            best_f1 = f1
                            best_threshold = threshold
                            best_pred = y_pred_thresh

                    # Record used threshold
                    logging.debug("Best threshold for this fold: {:.2f}".format(best_threshold))

                    # Calculate and save multiple metrics
                    f1 = f1_score(y_val_cv, best_pred, average='weighted')
                    f1_scores.append(f1)

                    try:
                        roc_auc = roc_auc_score(y_val_cv, y_pred_proba)
                        auc_scores.append(roc_auc)
                    except Exception as e:
                        logging.debug("Error calculating AUC: {}".format(str(e)))
                        auc_scores.append(0)

                    recall = recall_score(y_val_cv, best_pred, average='weighted')
                    recall_scores.append(recall)

                    sensitivity = recall_score(y_val_cv, best_pred, average='weighted')
                    sensitivity_scores.append(sensitivity)

                    precision = precision_score(y_val_cv, best_pred, average='weighted')
                    precision_scores.append(precision)

                    spec = weighted_specificity_score(y_val_cv, best_pred)
                    specificity_scores.append(spec)

            except Exception as e:
                logging.debug(f"Cross-validation error: {str(e)}")
                return 0.0

            # Comprehensive score - Weighted combination of multiple metrics
            mean_f1 = np.mean(f1_scores) if f1_scores else 0
            mean_auc = np.mean(auc_scores) if auc_scores else 0
            mean_recall = np.mean(recall_scores) if recall_scores else 0
            mean_specificity = np.mean(specificity_scores) if specificity_scores else 0
            mean_precision = np.mean(precision_scores) if precision_scores else 0

            combined_score = 0.55 * mean_f1 + 0.15 * mean_auc + 0.15 * mean_recall + 0.15 * mean_specificity

            # Record detailed evaluation
            param_str = f"hidden={hidden_layer_sizes}, act={activation}, solver={solver}, alpha={params['alpha']:.6f}, lr={learning_rate}, lr_init={params['learning_rate_init']:.6f}, batch={batch_size}"
            score_str = "F1={:.4f}, AUC={:.4f}, Recall={:.4f}, Spec={:.4f}, Score={:.4f}".format(
                mean_f1, mean_auc, mean_recall, mean_specificity, combined_score)
            logging.debug("Params: {} → {}".format(param_str, score_str))

            return combined_score

        # Use optimized parameter search space
        param_ranges = {
            'hidden_layer_factor': (0, self.config.get('max_hidden_layers', 3)),  # Hidden layer configuration factor
            'first_layer_neurons': self.config.get('neuron_ranges', {}).get('first_layer', (50, 300)),
            'second_layer_neurons': self.config.get('neuron_ranges', {}).get('second_layer', (30, 200)),
            'third_layer_neurons': self.config.get('neuron_ranges', {}).get('third_layer', (20, 150)),
            'fourth_layer_neurons': self.config.get('neuron_ranges', {}).get('fourth_layer', (10, 100)),
            'activation_idx': (1, 3),  # 0: 'identity', 1: 'logistic', 2: 'tanh', 3: 'relu'
            'solver_idx': (0, 2),  # 0: 'lbfgs', 1: 'sgd', 2: 'adam'
            'alpha': (1e-4, 1.0),  # L2 regularization parameter
            'learning_rate_idx': (0, 2),  # 0: 'constant', 1: 'invscaling', 2: 'adaptive'
            'learning_rate_init': (1e-5, 0.1),  # Initial learning rate
            'batch_size_factor': (1, 4)  # Batch size factor, actual batch size = factor * 16
        }

        # Use Bayesian optimization
        optimizer = BayesianOptimization(
            f=objective,
            pbounds=param_ranges,
            random_state=42,
            allow_duplicate_points=True
        )

        optimizer.maximize(
            init_points=self.config['bayes_opt']['init_points'],
            n_iter=self.config['bayes_opt']['n_iter']
        )

        # Get best parameters
        best_params = optimizer.max['params']

        # Convert parameters to correct types and values
        activations = ['identity', 'logistic', 'tanh', 'relu']
        solvers = ['lbfgs', 'sgd', 'adam']
        learning_rates = ['constant', 'invscaling', 'adaptive']

        # Calculate hidden layer sizes
        hidden_layer_factor = float(best_params['hidden_layer_factor'])

        # Neuron count parameters
        neuron_params = {
            'first_layer_neurons': best_params['first_layer_neurons'],
            'second_layer_neurons': best_params['second_layer_neurons'],
            'third_layer_neurons': best_params['third_layer_neurons'],
            'fourth_layer_neurons': best_params['fourth_layer_neurons']
        }

        # Get hidden layer configuration
        hidden_layer_sizes = self._get_hidden_layer_sizes(hidden_layer_factor, neuron_params)

        # Create dictionary containing all converted parameters
        params = {
            'hidden_layer_sizes': hidden_layer_sizes,
            'hidden_layer_factor': hidden_layer_factor,
            'first_layer_neurons': int(best_params['first_layer_neurons']),
            'second_layer_neurons': int(best_params['second_layer_neurons']),
            'third_layer_neurons': int(best_params['third_layer_neurons']),
            'fourth_layer_neurons': int(best_params['fourth_layer_neurons']),
            'activation': activations[int(best_params['activation_idx']) % len(activations)],
            'solver': solvers[int(best_params['solver_idx']) % len(solvers)],
            'alpha': best_params['alpha'],
            'learning_rate': learning_rates[int(best_params['learning_rate_idx']) % len(learning_rates)],
            'learning_rate_init': best_params['learning_rate_init'],
            'batch_size': int(min(X_train.shape[0], best_params['batch_size_factor'] * 16))
        }

        return params

    def _create_model(self, params):
        """Create ANN algorithms"""
        return MLPClassifier(
            hidden_layer_sizes=params['hidden_layer_sizes'],
            activation=params['activation'],
            solver=params['solver'],
            alpha=params['alpha'],
            batch_size=params['batch_size'],
            learning_rate=params['learning_rate'],
            learning_rate_init=params['learning_rate_init'],
            max_iter=1000,
            early_stopping=True,
            validation_fraction=0.2,
            n_iter_no_change=10,
            random_state=42
        )

    def plot_roc_curve(self, results, positive_class):
        """Plot ROC curve"""
        plt.figure(figsize=(10, 8))
        ax = plt.gca()

        # Set border thickness to 5
        for spine in ax.spines.values():
            spine.set_linewidth(5)

        # Set Arial font
        plt.rcParams['font.family'] = 'Arial'

        plt.title("ROC Curve for {}".format(positive_class), fontsize=20, pad=15, fontname='Arial')
        plt.xlabel("False Positive Rate", fontsize=24, labelpad=10, fontname='Arial')
        plt.ylabel("True Positive Rate", fontsize=24, labelpad=10, fontname='Arial')

        tprs = []
        aucs = []
        mean_fpr = np.linspace(0, 1, 500)

        for idx, result in enumerate(results):
            fpr = result['fpr']
            tpr = result['tpr']
            roc_auc = result['roc_auc']

            if fpr[0] != 0:
                fpr = np.concatenate([[0], fpr])
                tpr = np.concatenate([[0], tpr])

            if fpr[-1] != 1:
                fpr = np.concatenate([fpr, [1]])
                tpr = np.concatenate([tpr, [tpr[-1]]])

            interp_tpr = np.interp(mean_fpr, fpr, tpr)
            interp_tpr[0] = 0.0
            tprs.append(interp_tpr)
            aucs.append(roc_auc)

            plt.plot(fpr, tpr, alpha=0.15,
                     color=self.config['color_map'].get(positive_class, 'b'),
                     linestyle='-', linewidth=2)

        mean_tpr = np.mean(tprs, axis=0)
        mean_auc = np.mean(aucs)
        std_auc = np.std(aucs)

        plt.plot(mean_fpr, mean_tpr,
                 color=self.config['color_map'].get(positive_class, 'b'),
                 label='Mean ROC (AUC±std)\n{:.3f}±{:.3f}'.format(mean_auc, std_auc),
                 lw=5, alpha=0.8)

        std_tpr = np.std(tprs, axis=0)
        tprs_upper = np.minimum(mean_tpr + std_tpr, 1)
        tprs_lower = np.maximum(mean_tpr - std_tpr, 0)

        plt.fill_between(mean_fpr, tprs_lower, tprs_upper,
                         color=self.config['color_map'].get(positive_class, 'b'),
                         alpha=0.2, label='±1 std. dev.')

        plt.plot([0, 1], [0, 1], 'k--', lw=3, alpha=0.7)
        plt.grid(True, alpha=0.3, linewidth=0)
        plt.xlim([-0.02, 1.02])
        plt.ylim([-0.02, 1.02])

        ax.tick_params(width=5, length=10, labelsize=24)

        # Set tick label font
        for label in ax.get_xticklabels() + ax.get_yticklabels():
            label.set_fontname('Arial')

        legend = plt.legend(loc="lower right", fontsize=24, frameon=True)
        legend.get_frame().set_linewidth(3)

        # Set legend font
        for text in legend.get_texts():
            text.set_fontname('Arial')

        output_file = os.path.join(self.config['output_path'],
                                   "ROC_{}.png".format(positive_class))
        plt.savefig(output_file, dpi=600, bbox_inches='tight')
        plt.close()

    def plot_confusion_matrix(self, conf_matrix, class_name, fold_idx):
        """Plot confusion matrix"""
        plt.figure(figsize=(8, 6))

        # Set Arial font
        plt.rcParams['font.family'] = 'Arial'

        ax = plt.gca()
        sns.heatmap(conf_matrix, annot=True, fmt='d', cmap='Blues')

        plt.title('Confusion Matrix - {} (Fold {})'.format(class_name, fold_idx + 1),
                  fontsize=20, pad=15, fontname='Arial')
        plt.ylabel('True Label', fontsize=24, labelpad=10, fontname='Arial')
        plt.xlabel('Predicted Label', fontsize=24, labelpad=10, fontname='Arial')

        # Set tick label font
        ax.tick_params(labelsize=24)
        for label in ax.get_xticklabels() + ax.get_yticklabels():
            label.set_fontname('Arial')

        output_file = os.path.join(self.config['output_path'],
                                   "confusion_matrix_{}_fold_{}.png".format(class_name, fold_idx + 1))
        plt.savefig(output_file, dpi=300, bbox_inches='tight')
        plt.close()

    def format_excel_file(self, file_path):
        """Format Excel file: set font to Times New Roman and bold header row"""
        wb = load_workbook(file_path)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Apply Times New Roman to all cells and bold to header row
            for row_idx, row in enumerate(ws.iter_rows(), start=1):
                for cell in row:
                    if row_idx == 1:  # Header row
                        cell.font = Font(name='Times New Roman', size=11, bold=True)
                    else:
                        cell.font = Font(name='Times New Roman', size=11)

        wb.save(file_path)

    def save_results(self, all_results):
        """Save LOO results to Excel"""
        output_path = os.path.join(self.config['output_path'],
                                   self.config['output_filename'])

        with pd.ExcelWriter(output_path) as writer:
            # 1. Save overall summary
            summary_data = []
            for class_name, results in all_results.items():
                f1_scores = [r['f1'] for r in results]
                train_f1_scores = [r['train_f1'] for r in results]
                precision_scores = [r['precision'] for r in results]
                recall_scores = [r['recall'] for r in results]
                sensitivity_scores = [r['sensitivity'] for r in results]
                specificity_scores = [r['specificity'] for r in results]
                auc_scores = [r['roc_auc'] for r in results]
                pr_auc_scores = [r['pr_auc'] for r in results]
                fnr_scores = [r['fnr'] for r in results]  # False negative rate
                thresholds = [r['optimal_threshold'] for r in results]  # Optimal threshold

                summary_data.append({
                    'Class': class_name,
                    'Mean F1': np.mean(f1_scores),
                    'Std F1': np.std(f1_scores),
                    'Mean Precision': np.mean(precision_scores),
                    'Std Precision': np.std(precision_scores),
                    'Mean Recall': np.mean(recall_scores),
                    'Std Recall': np.std(recall_scores),
                    'Mean Sensitivity': np.mean(sensitivity_scores),
                    'Std Sensitivity': np.std(sensitivity_scores),
                    'Mean Specificity': np.mean(specificity_scores),
                    'Std Specificity': np.std(specificity_scores),
                    'Mean ROC AUC': np.mean(auc_scores),
                    'Std ROC AUC': np.std(auc_scores),
                    'Mean PR AUC': np.mean(pr_auc_scores),
                    'Std PR AUC': np.std(pr_auc_scores),
                    'Mean FNR': np.mean(fnr_scores),
                    'Mean Threshold': np.mean(thresholds),
                    'Best Params': str(results[-1]['best_params'])
                })

            pd.DataFrame(summary_data).to_excel(writer,
                                                sheet_name='Summary',
                                                index=False)

            # 2. Save detailed LOO results for each class
            for class_name, results in all_results.items():
                # Basic evaluation metrics
                fold_data = []
                for r in results:
                    params = r['best_params']
                    fold_data.append({
                        'Fold': r['fold_idx'] + 1,
                        'F1 Score': r['f1'],
                        'Precision': r['precision'],
                        'Recall': r['recall'],
                        'Sensitivity': r['sensitivity'],
                        'Specificity': r['specificity'],
                        'Train F1': r['train_f1'],
                        'ROC AUC': r['roc_auc'],
                        'PR AUC': r['pr_auc'],
                        'False Negative Rate': r['fnr'],
                        'Optimal Threshold': r['optimal_threshold'],
                        'Hidden Layers': str(params.get('hidden_layer_sizes', '')),
                        'First Layer Neurons': params.get('first_layer_neurons', 0),
                        'Second Layer Neurons': params.get('second_layer_neurons', 0) if len(
                            params.get('hidden_layer_sizes', ())) > 1 else 0,
                        'Third Layer Neurons': params.get('third_layer_neurons', 0) if len(
                            params.get('hidden_layer_sizes', ())) > 2 else 0,
                        'Fourth Layer Neurons': params.get('fourth_layer_neurons', 0) if len(
                            params.get('hidden_layer_sizes', ())) > 3 else 0,
                        'Activation': params.get('activation', ''),
                        'Solver': params.get('solver', ''),
                        'Alpha': params.get('alpha', 0),
                        'Learning Rate': params.get('learning_rate', ''),
                        'Learning Rate Init': params.get('learning_rate_init', 0),
                        'Batch Size': params.get('batch_size', 0)
                    })

                pd.DataFrame(fold_data).to_excel(writer,
                                                 sheet_name=f'{class_name}_details',
                                                 index=False)

                # Save ROC curve data
                # 1. ROC data for each fold
                for idx, result in enumerate(results):
                    fold_roc_data = pd.DataFrame({
                        'FPR': result['fpr'],
                        'TPR': result['tpr']
                    })
                    fold_roc_data.to_excel(writer,
                                           sheet_name=f'{class_name}_ROC_data',
                                           startcol=idx * 3,
                                           index=False)
                    # Add AUC value next to each fold's data
                    pd.DataFrame({'AUC': [result['roc_auc']]}).to_excel(
                        writer,
                        sheet_name=f'{class_name}_ROC_data',
                        startcol=idx * 3 + 2,
                        index=False
                    )

                # 2. Calculate and save mean ROC curve data
                mean_fpr = np.linspace(0, 1, 500)
                tprs = []
                for result in results:
                    # Ensure endpoints are included
                    fpr = result['fpr']
                    tpr = result['tpr']
                    if fpr[0] != 0:
                        fpr = np.concatenate([[0], fpr])
                        tpr = np.concatenate([[0], tpr])
                    if fpr[-1] != 1:
                        fpr = np.concatenate([fpr, [1]])
                        tpr = np.concatenate([tpr, [tpr[-1]]])

                    # Interpolation
                    interp_tpr = np.interp(mean_fpr, fpr, tpr)
                    interp_tpr[0] = 0.0
                    tprs.append(interp_tpr)

                mean_tpr = np.mean(tprs, axis=0)
                std_tpr = np.std(tprs, axis=0)

                # Save mean ROC curve data
                mean_roc_data = pd.DataFrame({
                    'FPR': mean_fpr,
                    'Mean_TPR': mean_tpr,
                    'Std_TPR': std_tpr,
                    'TPR_Upper': np.minimum(mean_tpr + std_tpr, 1),
                    'TPR_Lower': np.maximum(mean_tpr - std_tpr, 0)
                })
                mean_roc_data.to_excel(writer,
                                       sheet_name=f'{class_name}_mean_ROC',
                                       index=False)

        # Format Excel file
        self.format_excel_file(output_path)

    def run(self):
        """Run complete training and evaluation process"""
        # Load data
        X, y, original_indices = self.load_data()
        logging.info("Using ANN classifier")
        logging.info("Data shape: {} samples, {} features".format(X.shape[0], X.shape[1]))

        # Create cross-validation object
        outer_cv = StratifiedKFold(n_splits=self.config['outer_cv_splits'], shuffle=True, random_state=42)

        all_results = {}

        for class_name in self.all_classes:
            logging.info("Processing class: {}".format(class_name))

            X_balanced, y_binary_balanced, sample_indices = self.apply_smote(X, y, class_name)

            # Check class distribution
            class_counts = np.bincount(y_binary_balanced)
            pos_count = class_counts[1] if len(class_counts) > 1 else 0
            neg_count = class_counts[0]
            ratio = neg_count / max(1, pos_count)

            # If one class count is 0, skip
            if len(class_counts) <= 1 or pos_count == 0:
                logging.warning("No positive samples for class {}. Skipping.".format(class_name))
                continue

            # Use cross-validation - Parallel processing
            try:
                # Prepare cross-validation indices
                fold_indices = list(enumerate(outer_cv.split(X_balanced, y_binary_balanced)))

                # Process each fold in parallel
                results = Parallel(n_jobs=-1)(
                    delayed(self.process_fold)(
                        X_balanced, y_binary_balanced, sample_indices, train_idx, test_idx, class_name, fold_idx
                    )
                    for fold_idx, (train_idx, test_idx) in fold_indices
                )

                # If no valid LOO results, skip
                if not results:
                    logging.warning("No valid LOO results for class {}. Skipping.".format(class_name))
                    continue

                all_results[class_name] = results

                # Plot ROC curve
                self.plot_roc_curve(results, class_name)

                # Output performance metrics for current class
                f1_scores = [r['f1'] for r in results]
                auc_scores = [r['roc_auc'] for r in results]
                pr_auc_scores = [r['pr_auc'] for r in results]
                precision_scores = [r['precision'] for r in results]
                recall_scores = [r['recall'] for r in results]
                sensitivity_scores = [r['sensitivity'] for r in results]
                specificity_scores = [r['specificity'] for r in results]
                fnr_scores = [r['fnr'] for r in results]
                thresholds = [r['optimal_threshold'] for r in results]

                logging.info("Performance metrics for {}:".format(class_name))
                logging.info("F1: {:.3f}±{:.3f}, AUC: {:.3f}±{:.3f}, Precision: {:.3f}±{:.3f}, Recall: {:.3f}±{:.3f}, Sensitivity: {:.3f}±{:.3f}, Specificity: {:.3f}±{:.3f}".format(
                    np.mean(f1_scores), np.std(f1_scores),
                    np.mean(auc_scores), np.std(auc_scores),
                    np.mean(precision_scores), np.std(precision_scores),
                    np.mean(recall_scores), np.std(recall_scores),
                    np.mean(sensitivity_scores), np.std(sensitivity_scores),
                    np.mean(specificity_scores), np.std(specificity_scores)))

            except Exception as e:
                logging.error("Error processing class {}: {}".format(class_name, str(e)))
                import traceback
                logging.error(traceback.format_exc())
                # Continue processing other classes

        # Save LOO results
        if all_results:
            self.save_results(all_results)
            return all_results
        else:
            logging.error("No LOO results were generated. Check for errors above.")
            return {}


def main():
    """Main function"""
    # Configuration parameters
    config = {
        'file_path': os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '00_Data', 'Train.xlsx'),  # Path to the input Excel file containing training data
        'sheet_name': 'Sheet1',  # Name of the sheet in the Excel file (e.g., 'Sheet1')
        'output_path': os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '..', 'Results', 'Multicancer detection', '02_Algorithm optimization', 'ANN'),  # Directory path for saving output results
        'output_filename': 'ANN_results.xlsx',

        # Cross-validation configuration
        'outer_cv_splits': 5,
        'inner_cv_splits': 5,

        # Feature processing configuration
        'interaction_method': 'gaussian',
        'selection_method': 'boruta',
        'max_features': 35,
        'gaussian_gamma': 'auto',
        'scale_after_interaction': True,
        'debug_mode': True,

        # SMOTE configuration
        'smote_sampling_strategy': 0.5,
        'smote_k_neighbors': 5,

        # Bayesian optimization configuration
        'bayes_opt': {
            'init_points': 12,
            'n_iter': 80,
        },

        # ANN-specific configuration
        'max_hidden_layers': 2,

        # Define neuron search ranges
        'neuron_ranges': {
            'first_layer': (10, 100),
            'second_layer': (5, 50),
            'third_layer': (5, 20),
            'fourth_layer': (2, 10),
        },

        # Color configuration
        'color_map': {
            'LC': [254 / 255, 160 / 255, 64 / 255],
            'LuC': [242 / 255, 128 / 255, 128 / 255],
            'OC': [88 / 255, 97 / 255, 172 / 255],
            'N': [106 / 255, 180 / 255, 193 / 255]
        }
    }

    try:
        # Create output directory
        os.makedirs(config['output_path'], exist_ok=True)

        if config['debug_mode']:
            logging.info("Feature interaction method: {}".format(config['interaction_method']))
            logging.info("Feature selection method: {}".format(config['selection_method']))

        # Read Excel to get class information
        df = pd.read_excel(config['file_path'], sheet_name=config['sheet_name'])
        config['all_classes'] = sorted(df['Sample'].unique())

        # Validate data format
        required_columns = ['Sample']
        if not all(col in df.columns for col in required_columns):
            raise ValueError("Missing required columns. Required: {}".format(required_columns))

        # Run classifier
        classifier = ANNClassifier(config)
        results = classifier.run()

        # Output final LOO results summary
        logging.info("Training completed successfully!")

    except Exception as e:
        logging.error("An error occurred: {}".format(str(e)))
        import traceback
        logging.error(traceback.format_exc())
        raise


if __name__ == "__main__":
    main()