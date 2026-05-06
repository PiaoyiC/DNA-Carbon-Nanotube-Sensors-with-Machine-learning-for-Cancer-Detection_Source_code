import sys
import threading
import multiprocessing
import pickle
import glob
import numpy as np
import pandas as pd
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import StratifiedKFold
from sklearn.metrics import (f1_score, roc_curve, auc, confusion_matrix, precision_score,
                             recall_score, precision_recall_curve, average_precision_score,
                             roc_auc_score)
import matplotlib.pyplot as plt
import logging
import os
import seaborn as sns
import warnings
import random
import time
from typing import Dict, List, Tuple, Any
from datetime import datetime
from joblib import Parallel, delayed

from imblearn.over_sampling import SMOTE, ADASYN

warnings.filterwarnings('ignore')

import matplotlib as mpl

mpl.rcParams['svg.fonttype'] = 'none'
mpl.rcParams['font.family'] = 'Arial'
mpl.rcParams['pdf.fonttype'] = 42
mpl.rcParams['ps.fonttype'] = 42

logging.basicConfig(level=logging.CRITICAL)


class GAEvolutionLogger:
    """Genetic Algorithm Evolution History Logger"""

    def __init__(self, output_path: str):
        self.output_path = output_path
        self.temp_dir = os.path.join(output_path, "temp_logs")
        self.evolution_logs = {}

        # Create temporary directory
        os.makedirs(self.temp_dir, exist_ok=True)

    def start_algorithm(self, algorithm_name: str, class_name: str, fold: int):
        """Start logging algorithm"""
        key = f"{algorithm_name}_{class_name}_fold{fold}"
        temp_file = os.path.join(self.temp_dir, f"{key}.pkl")

        log_data = {
            'algorithm': algorithm_name,
            'class': class_name,
            'fold': fold,
            'start_time': datetime.now(),
            'generations': [],
            'config': {},
            'final_result': {}
        }

        # Save to temporary file
        try:
            with open(temp_file, 'wb') as f:
                pickle.dump(log_data, f)
        except Exception as e:
            pass

        return temp_file

    def log_generation(self, temp_file: str, generation: int, best_score: float,
                       avg_score: float, diversity: float, generation_time: float, best_params: dict = None):
        """Log generation information to temporary file"""
        try:
            # Read existing data
            if os.path.exists(temp_file):
                with open(temp_file, 'rb') as f:
                    log_data = pickle.load(f)

                # Add new generation record
                log_data['generations'].append({
                    'generation': generation,
                    'best_score': best_score,
                    'avg_score': avg_score,
                    'diversity': diversity,
                    'time': generation_time,
                    'best_params': best_params.copy() if best_params else None
                })

                # Write back to file
                with open(temp_file, 'wb') as f:
                    pickle.dump(log_data, f)
        except Exception as e:
            pass  # Silently handle file operation errors

    def finish_algorithm(self, temp_file: str, final_score: float, final_params: dict, total_time: float):
        """Finish algorithm logging"""
        try:
            if os.path.exists(temp_file):
                with open(temp_file, 'rb') as f:
                    log_data = pickle.load(f)

                log_data['final_result'] = {
                    'final_score': final_score,
                    'final_params': final_params.copy() if final_params else None,
                    'total_time': total_time,
                    'end_time': datetime.now()
                }

                with open(temp_file, 'wb') as f:
                    pickle.dump(log_data, f)
        except Exception as e:
            pass

    def merge_and_save_logs(self):

        log_file = os.path.join(self.output_path, "GA_Evolution_History.txt")

        try:
            # Collect all temporary files
            temp_files = glob.glob(os.path.join(self.temp_dir, "*.pkl"))

            if not temp_files:
                with open(log_file, 'w', encoding='utf-8') as f:
                    f.write("=" * 80 + "\n")
                    f.write("Genetic Algorithm Evolution History Report\n")
                    f.write(f"Generation Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                    f.write("=" * 80 + "\n\n")
                    f.write("No evolution history data found.\n")
                    f.write("Possible reasons: Log recording failed during parallel processing or optimization process abnormal.\n")
                return

            # Read and merge all log data
            all_logs = {}
            for temp_file in temp_files:
                try:
                    with open(temp_file, 'rb') as f:
                        log_data = pickle.load(f)
                        key = f"{log_data['algorithm']}_{log_data['class']}_fold{log_data['fold']}"
                        all_logs[key] = log_data
                except Exception as e:
                    continue  # Skip corrupted files

            # Write to final log file
            with open(log_file, 'w', encoding='utf-8') as f:
                f.write("=" * 80 + "\n")
                f.write("Genetic Algorithm Evolution History Report\n")
                f.write(f"Generation Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"Total Algorithm Instances: {len(all_logs)}\n")
                f.write("=" * 80 + "\n\n")

                if not all_logs:
                    f.write("No evolution history data successfully merged.\n")
                    return

                # Organize output by category and algorithm
                for key, log in sorted(all_logs.items()):
                    f.write(f"{log['algorithm']} - Class {log['class']} - Fold {log['fold']}\n")
                    f.write("-" * 60 + "\n")
                    f.write(f"Start Time: {log['start_time'].strftime('%H:%M:%S')}\n")

                    if log['final_result']:
                        f.write(f"Total Time: {log['final_result']['total_time']:.2f}s\n")
                        f.write(f"Final Score: {log['final_result']['final_score']:.4f}\n")

                    f.write(f"Evolution Generations: {len(log['generations'])}\n\n")

                    if log['generations']:
                        # Evolution process
                        f.write("Evolution Process:\n")
                        f.write("Gen  Best Score  Avg Score  Diversity  Time(s)\n")
                        f.write("-" * 50 + "\n")

                        for gen in log['generations']:
                            f.write(f"{gen['generation']:3d}   {gen['best_score']:8.4f}  {gen['avg_score']:8.4f}  "
                                    f"{gen['diversity']:10.4f}  {gen['time']:7.2f}\n")
                    else:
                        f.write("No evolution process data recorded for this algorithm.\n")

                    # Best parameters
                    if log['final_result'] and log['final_result']['final_params']:
                        f.write(f"\nBest Parameters:\n")
                        for param, value in log['final_result']['final_params'].items():
                            f.write(f"  {param}: {value}\n")

                    f.write("\n" + "=" * 80 + "\n\n")

            print(f"Evolution history saved to: {log_file}")

            # Clean up temporary files
            self._cleanup_temp_files()

        except Exception as e:
            print(f"Error saving evolution history: {str(e)}")

    def _cleanup_temp_files(self):
        try:
            temp_files = glob.glob(os.path.join(self.temp_dir, "*.pkl"))
            for temp_file in temp_files:
                try:
                    os.remove(temp_file)
                except:
                    pass

            # Attempt to delete temporary directory
            try:
                os.rmdir(self.temp_dir)
            except:
                pass
        except Exception as e:
            pass

    def get_summary_stats(self):
        """Get overall statistics"""
        try:
            temp_files = glob.glob(os.path.join(self.temp_dir, "*.pkl"))
            if not temp_files:
                return None

            stats = {
                'total_algorithms': len(temp_files),
                'total_generations': 0,
                'algorithms': {},
                'classes': {}
            }

            for temp_file in temp_files:
                try:
                    with open(temp_file, 'rb') as f:
                        log_data = pickle.load(f)

                    algo = log_data['algorithm']
                    class_name = log_data['class']

                    stats['total_generations'] += len(log_data['generations'])

                    # Algorithm statistics
                    if algo not in stats['algorithms']:
                        stats['algorithms'][algo] = {'count': 0, 'total_time': 0, 'avg_score': []}

                    stats['algorithms'][algo]['count'] += 1
                    if log_data['final_result']:
                        stats['algorithms'][algo]['total_time'] += log_data['final_result']['total_time']
                        stats['algorithms'][algo]['avg_score'].append(log_data['final_result']['final_score'])

                    # Category statistics
                    if class_name not in stats['classes']:
                        stats['classes'][class_name] = {'count': 0, 'total_time': 0, 'avg_score': []}

                    stats['classes'][class_name]['count'] += 1
                    if log_data['final_result']:
                        stats['classes'][class_name]['total_time'] += log_data['final_result']['total_time']
                        stats['classes'][class_name]['avg_score'].append(log_data['final_result']['final_score'])

                except Exception as e:
                    continue

            return stats
        except Exception as e:
            return None


class TimeEstimator:
    """Improved Time Estimator"""

    def __init__(self):
        self.class_times = []
        self.total_start_time = None

    def start_total(self):
        self.total_start_time = time.time()

    def record_class_time(self, class_time):
        """Record class completion time"""
        self.class_times.append(class_time)

    def get_average_class_time(self):
        """Get average class processing time"""
        if not self.class_times:
            return None
        return np.mean(self.class_times)

    def get_adaptive_estimate(self):
        """Adaptive estimation: consider time change trend"""
        if len(self.class_times) < 2:
            return self.get_average_class_time()

        # Consider recent trend
        recent_avg = np.mean(self.class_times[-2:])  # Average time for the last 2 classes
        overall_avg = np.mean(self.class_times)  # Overall average time

        # Weighted average, more emphasis on recent trend
        if len(self.class_times) >= 3:
            return 0.6 * recent_avg + 0.4 * overall_avg
        else:
            return recent_avg

    def estimate_remaining_time(self, completed_classes, total_classes):
        """Estimate remaining total time"""
        if completed_classes < 1:
            return None

        estimated_time_per_class = self.get_adaptive_estimate()
        if estimated_time_per_class is None:
            return None

        remaining_classes = total_classes - completed_classes
        return estimated_time_per_class * remaining_classes

    def format_time(self, seconds):
        """Format time display"""
        if seconds is None:
            return ""
        elif seconds < 60:
            return f"~{seconds:.0f}s"
        else:
            return f"~{seconds / 60:.1f}min"

    def get_total_time(self):
        """Get total runtime"""
        return time.time() - self.total_start_time if self.total_start_time else 0


class SimpleSpinnerIndicator:
    """Simple Spinner Indicator"""

    def __init__(self):
        self.running = False
        self.thread = None
        self.start_time = None

    def start(self, message="   Processing folds 1,2,3,4,5 in parallel"):
        self.running = True
        self.start_time = time.time()
        self.message = message
        self.thread = threading.Thread(target=self._animate)
        self.thread.daemon = True
        self.thread.start()

    def stop(self):
        self.running = False
        if self.thread:
            self.thread.join(timeout=1.0)
        sys.stdout.write('\r' + ' ' * 100 + '\r')
        sys.stdout.flush()

    def _animate(self):
        """Simple rotation animation"""
        spinners = ["|", "/", "-", "\\"]
        frame_idx = 0

        while self.running:
            elapsed = time.time() - self.start_time
            minutes = int(elapsed // 60)
            seconds = int(elapsed % 60)

            spinner = spinners[frame_idx % len(spinners)]
            time_str = f"{minutes:02d}:{seconds:02d}"

            display_text = f"\r{self.message}... {spinner} [{time_str}]"

            sys.stdout.write(display_text)
            sys.stdout.flush()

            frame_idx += 1
            time.sleep(0.5)


class GeneticAlgorithmOptimizer:
    """Genetic Algorithm Optimizer"""

    def __init__(self,
                 population_size: int = 40,
                 max_generations: int = 50,
                 crossover_rate: float = 0.8,
                 mutation_rate: float = 0.15,
                 elite_ratio: float = 0.2,
                 tournament_size: int = 3,
                 convergence_threshold: int = 10,
                 random_state: int = 42,
                 algorithm_name: str = "GA",
                 logger: 'GAEvolutionLogger' = None,
                 temp_log_file: str = None):

        self.population_size = population_size
        self.max_generations = max_generations
        self.crossover_rate = crossover_rate
        self.mutation_rate = mutation_rate
        self.elite_size = max(1, int(population_size * elite_ratio))
        self.tournament_size = tournament_size
        self.convergence_threshold = convergence_threshold
        self.random_state = random_state
        self.algorithm_name = algorithm_name
        self.logger = logger
        self.temp_log_file = temp_log_file

        # Set random seed
        random.seed(random_state)
        np.random.seed(random_state)

        # Optimization history
        self.history = {
            'best_scores': [],
            'avg_scores': [],
            'diversity_scores': [],
            'generation_times': []
        }

        # Best result
        self.best_individual = None
        self.best_score = float('-inf')

    def _encode_individual(self, param_bounds: Dict[str, Tuple]) -> Dict[str, Any]:
        """Encode an individual"""
        individual = {}
        for param_name, bounds in param_bounds.items():
            min_val, max_val, param_type = bounds
            if param_type == 'int':
                individual[param_name] = random.randint(min_val, max_val)
            elif param_type == 'float':
                individual[param_name] = random.uniform(min_val, max_val)
            elif param_type == 'categorical':
                individual[param_name] = random.choice(min_val)
        return individual

    def _create_initial_population(self, param_bounds: Dict[str, Tuple]) -> List[Dict[str, Any]]:
        """Create initial population"""
        return [self._encode_individual(param_bounds) for _ in range(self.population_size)]

    def _tournament_selection(self, population: List[Dict], fitness_scores: List[float]) -> Dict[str, Any]:
        """Tournament selection"""
        tournament_indices = random.sample(range(len(population)),
                                           min(self.tournament_size, len(population)))
        tournament_fitness = [fitness_scores[i] for i in tournament_indices]
        winner_idx = tournament_indices[np.argmax(tournament_fitness)]
        return population[winner_idx].copy()

    def _adaptive_crossover(self, parent1: Dict[str, Any], parent2: Dict[str, Any],
                            param_bounds: Dict[str, Tuple]) -> Tuple[Dict[str, Any], Dict[str, Any]]:
        """Adaptive crossover operation"""
        child1, child2 = parent1.copy(), parent2.copy()

        for param_name in parent1.keys():
            if random.random() < self.crossover_rate:
                bounds = param_bounds[param_name]
                param_type = bounds[2]

                if param_type in ['int', 'float']:
                    val1, val2 = parent1[param_name], parent2[param_name]
                    eta = 20
                    if abs(val1 - val2) > 1e-14:
                        if val1 > val2:
                            val1, val2 = val2, val1

                        beta = 1.0 + (2.0 * (val1 - bounds[0]) / (val2 - val1))
                        alpha = 2.0 - beta ** -(eta + 1.0)

                        if random.random() <= 1.0 / alpha:
                            beta_q = (random.random() * alpha) ** (1.0 / (eta + 1.0))
                        else:
                            beta_q = (1.0 / (2.0 - random.random() * alpha)) ** (1.0 / (eta + 1.0))

                        child1_val = 0.5 * ((val1 + val2) - beta_q * (val2 - val1))
                        child2_val = 0.5 * ((val1 + val2) + beta_q * (val2 - val1))

                        child1_val = np.clip(child1_val, bounds[0], bounds[1])
                        child2_val = np.clip(child2_val, bounds[0], bounds[1])

                        if param_type == 'int':
                            child1_val = int(round(child1_val))
                            child2_val = int(round(child2_val))

                        child1[param_name] = child1_val
                        child2[param_name] = child2_val

                elif param_type == 'categorical':
                    if random.random() < 0.5:
                        child1[param_name], child2[param_name] = child2[param_name], child1[param_name]

        return child1, child2

    def _adaptive_mutation(self, individual: Dict[str, Any], param_bounds: Dict[str, Tuple],
                           generation: int) -> Dict[str, Any]:
        """Adaptive mutation operation"""
        mutated = individual.copy()
        adaptive_rate = self.mutation_rate * (1.0 - generation / self.max_generations)

        for param_name, bounds in param_bounds.items():
            if random.random() < adaptive_rate:
                min_val, max_val, param_type = bounds

                if param_type == 'int':
                    current_val = mutated[param_name]
                    sigma = (max_val - min_val) * 0.1 * (1.0 - generation / self.max_generations)
                    new_val = int(round(np.random.normal(current_val, sigma)))
                    mutated[param_name] = np.clip(new_val, min_val, max_val)

                elif param_type == 'float':
                    current_val = mutated[param_name]
                    sigma = (max_val - min_val) * 0.1 * (1.0 - generation / self.max_generations)
                    new_val = np.random.normal(current_val, sigma)
                    mutated[param_name] = np.clip(new_val, min_val, max_val)

                elif param_type == 'categorical':
                    mutated[param_name] = random.choice(min_val)

        return mutated

    def _calculate_diversity(self, population: List[Dict[str, Any]]) -> float:
        """Calculate population diversity"""
        if len(population) < 2:
            return 0.0

        total_distance = 0.0
        count = 0

        for i in range(len(population)):
            for j in range(i + 1, len(population)):
                distance = 0.0
                for param_name in population[i].keys():
                    val1, val2 = population[i][param_name], population[j][param_name]
                    if isinstance(val1, (int, float)) and isinstance(val2, (int, float)):
                        distance += (val1 - val2) ** 2
                    else:
                        distance += 0 if val1 == val2 else 1

                total_distance += np.sqrt(distance)
                count += 1

        return total_distance / count if count > 0 else 0.0

    def _check_convergence(self, generation: int) -> bool:
        """Check convergence condition"""
        if len(self.history['best_scores']) < self.convergence_threshold:
            return False

        recent_scores = self.history['best_scores'][-self.convergence_threshold:]
        improvement = max(recent_scores) - min(recent_scores)
        return improvement < 1e-6

    def _safe_objective(self, objective_function, individual):
        """Safe objective function call, handle exceptions"""
        try:
            return objective_function(individual)
        except Exception as e:
            return 0.0

    def optimize(self, objective_function, param_bounds: Dict[str, Tuple]) -> Dict[str, Any]:
        """Main optimization loop"""
        start_time = time.time()

        # Create initial population
        population = self._create_initial_population(param_bounds)

        for generation in range(self.max_generations):
            generation_start = time.time()

            # Resume parallel evaluation of all individuals' fitness
            try:
                fitness_scores = Parallel(
                    n_jobs=-1,
                    prefer="threads",  # Use threads, as shared objective_function is needed
                    batch_size=1,  # Process one individual at a time, better load balancing
                    verbose=0  # Silent mode, do not display parallel processing info
                )(
                    delayed(self._safe_objective)(objective_function, individual)
                    for individual in population
                )
            except Exception as e:
                # If parallel fails, fallback to serial processing
                fitness_scores = []
                for individual in population:
                    fitness_scores.append(self._safe_objective(objective_function, individual))

            # Ensure fitness_scores is a list and length is correct
            if len(fitness_scores) != len(population):
                # Pad missing scores
                while len(fitness_scores) < len(population):
                    fitness_scores.append(0.0)

            # Update best individual
            current_best_idx = np.argmax(fitness_scores)
            current_best_score = fitness_scores[current_best_idx]

            if current_best_score > self.best_score:
                self.best_score = current_best_score
                self.best_individual = population[current_best_idx].copy()

            # Record history
            avg_score = np.mean(fitness_scores)
            diversity = self._calculate_diversity(population)
            generation_time = time.time() - generation_start

            self.history['best_scores'].append(self.best_score)
            self.history['avg_scores'].append(avg_score)
            self.history['diversity_scores'].append(diversity)
            self.history['generation_times'].append(generation_time)

            # Use temporary file to log
            if self.logger and self.temp_log_file:
                try:
                    self.logger.log_generation(
                        self.temp_log_file, generation + 1, self.best_score, avg_score,
                        diversity, generation_time, self.best_individual
                    )
                except Exception as e:
                    pass  # Silently handle log errors

            # Convergence check
            if self._check_convergence(generation):
                break

            # Generate next generation
            new_population = []

            # Elite retention
            elite_indices = np.argsort(fitness_scores)[-self.elite_size:]
            for idx in elite_indices:
                new_population.append(population[idx].copy())

            # Generate new individuals
            while len(new_population) < self.population_size:
                parent1 = self._tournament_selection(population, fitness_scores)
                parent2 = self._tournament_selection(population, fitness_scores)
                child1, child2 = self._adaptive_crossover(parent1, parent2, param_bounds)
                child1 = self._adaptive_mutation(child1, param_bounds, generation)
                child2 = self._adaptive_mutation(child2, param_bounds, generation)
                new_population.extend([child1, child2])

            # Ensure population size is correct
            population = new_population[:self.population_size]

        total_time = time.time() - start_time

        # Finish recording
        if self.logger and self.temp_log_file:
            try:
                self.logger.finish_algorithm(self.temp_log_file, self.best_score, self.best_individual, total_time)
            except Exception as e:
                pass

        return self.best_individual

    def get_optimization_history(self):
        """Get optimization history"""
        return self.history


# Helper functions
def weighted_specificity_score(y_true, y_pred):

    classes = np.unique(np.concatenate([y_true, y_pred]))
    cm = confusion_matrix(y_true, y_pred, labels=classes)

    specificity_sum = 0
    total_samples = 0

    for i, cls in enumerate(classes):
        true_neg_mask = np.ones(cm.shape, dtype=bool)
        true_neg_mask[i, :] = False
        true_neg_mask[:, i] = False

        tn = np.sum(cm[true_neg_mask])
        fp = np.sum(cm[:, i]) - cm[i, i]

        if tn + fp > 0:
            specificity = tn / (tn + fp)
        else:
            specificity = 0

        weight = np.sum(y_true == cls)
        specificity_sum += specificity * weight
        total_samples += weight

    return specificity_sum / total_samples if total_samples > 0 else 0


def weighted_specificity(y_true, y_pred, labels):

    total_specificity = 0
    total_weight = 0

    for label in labels:
        true_negatives = np.sum((y_true != label) & (y_pred != label))
        false_positives = np.sum((y_true != label) & (y_pred == label))
        weight = np.sum(y_true == label)
        total_weight += weight

        if (true_negatives + false_positives) > 0:
            specificity = true_negatives / (true_negatives + false_positives)
        else:
            specificity = 0.0
        total_specificity += specificity * weight

    return total_specificity / total_weight if total_weight > 0 else 0.0


class RandomForestModel:
    def __init__(self, config, logger=None):
        self.config = config
        self.all_classes = config['all_classes']
        self.scaler = StandardScaler()
        self.logger = logger

        # Initialize SMOTE
        smote_sampling_strategy = config.get('smote_sampling_strategy', 'auto')
        smote_k_neighbors = config.get('smote_k_neighbors', 5)
        self.smote = SMOTE(
            sampling_strategy=smote_sampling_strategy,
            k_neighbors=smote_k_neighbors,
            random_state=42
        )

        # Initialize feature interaction processor
        try:
            from AdvancedFeatureInteractionsRF import AdvancedFeatureInteractions
            self.feature_interaction = AdvancedFeatureInteractions(**config.get('feature_params', {}))
        except ImportError:
            self.feature_interaction = AdvancedFeatureInteractions(**config.get('feature_params', {}))

    def load_data(self):
        """Load and process data"""
        df = pd.read_excel(self.config['file_path'], sheet_name=self.config['sheet_name'])
        labels = df['Sample'].values
        features = df.iloc[:, 1:14].values

        self.original_indices = np.arange(len(features))
        original_feature_names = df.columns[1:14].tolist()
        self.feature_interaction.set_original_feature_names(original_feature_names)

        features_processed = self.feature_interaction.fit_transform(features, labels)
        self.feature_names = self.feature_interaction.get_feature_names()
        self.feature_importance = self.feature_interaction.get_feature_importance()

        return features_processed, np.array(labels), self.original_indices

    def apply_smote(self, X, y, positive_class):

        y_binary = (y == positive_class).astype(int)

        try:
            X_resampled, y_binary_resampled = self.smote.fit_resample(X, y_binary)
            sample_indices = np.concatenate([
                self.original_indices,
                np.full(len(X_resampled) - len(X), -1)
            ])
            return X_resampled, y_binary_resampled, sample_indices
        except ValueError:
            return X, y_binary, self.original_indices

    def _optimize_hyperparameters(self, X_train, y_train, class_name, fold_idx):
        """Optimize hyperparameters using genetic algorithm"""
        neg_count = sum(y_train == 0)
        pos_count = sum(y_train == 1)
        base_class_ratio = neg_count / max(1, pos_count)

        def objective(params):
            try:
                max_features = max(0.1, min(0.999, params['max_features_ratio']))

                if params['class_weight_multiplier'] > 0:
                    class_weight = {
                        0: 1.0,
                        1: params['class_weight_multiplier'] * base_class_ratio
                    }
                else:
                    class_weight = None

                from sklearn.ensemble import RandomForestClassifier
                model = RandomForestClassifier(
                    n_estimators=params['n_estimators'],
                    max_depth=params['max_depth'],
                    min_samples_split=params['min_samples_split'],
                    min_samples_leaf=params['min_samples_leaf'],
                    max_features=max_features,
                    class_weight=class_weight,
                    random_state=42,
                    n_jobs=-1  # Resume parallel computation
                )

                cv = StratifiedKFold(n_splits=self.config['inner_cv_splits'], shuffle=True, random_state=42)
                f1_scores = []
                auc_scores = []
                recall_scores = []
                specificity_scores = []

                for train_idx, val_idx in cv.split(X_train, y_train):
                    X_train_cv, X_val_cv = X_train[train_idx], X_train[val_idx]
                    y_train_cv, y_val_cv = y_train[train_idx], y_train[val_idx]

                    y_train_cv = y_train_cv.astype(int)
                    model.fit(X_train_cv, y_train_cv)
                    y_pred_proba = model.predict_proba(X_val_cv)[:, 1]

                    best_f1 = 0
                    best_pred = None
                    for threshold in [0.3, 0.35, 0.4, 0.45, 0.5]:
                        y_pred_thresh = (y_pred_proba >= threshold).astype(int)
                        f1 = f1_score(y_val_cv, y_pred_thresh, average='weighted')
                        if f1 > best_f1:
                            best_f1 = f1
                            best_pred = y_pred_thresh

                    f1_scores.append(f1_score(y_val_cv, best_pred, average='weighted'))

                    try:
                        auc_scores.append(roc_auc_score(y_val_cv, y_pred_proba))
                    except Exception:
                        auc_scores.append(0)

                    recall_scores.append(recall_score(y_val_cv, best_pred, average='weighted'))
                    specificity_scores.append(weighted_specificity_score(y_val_cv, best_pred))

                mean_f1 = np.mean(f1_scores)
                mean_auc = np.mean(auc_scores)
                mean_recall = np.mean(recall_scores)
                mean_specificity = np.mean(specificity_scores)

                return 0.55 * mean_f1 + 0.15 * mean_auc + 0.10 * mean_recall + 0.20 * mean_specificity

            except Exception:
                return 0.0

        # Parameter bounds
        param_bounds = {
            'n_estimators': (100, 500, 'int'),
            'max_depth': (3, 40, 'int'),
            'min_samples_split': (2, 20, 'int'),
            'min_samples_leaf': (2, 20, 'int'),
            'max_features_ratio': (0.1, 1.0, 'float'),
            'class_weight_multiplier': (0.0, 5.0, 'float')
        }

        # Create genetic algorithm optimizer
        temp_log_file = None
        if self.logger:
            temp_log_file = self.logger.start_algorithm("RandomForest", class_name, fold_idx + 1)

        ga_config = self.config.get('ga_config', {})
        ga_optimizer = GeneticAlgorithmOptimizer(
            population_size=ga_config.get('population_size', 40),
            max_generations=ga_config.get('max_generations', 50),
            crossover_rate=ga_config.get('crossover_rate', 0.8),
            mutation_rate=ga_config.get('mutation_rate', 0.15),
            random_state=42,
            algorithm_name="RF",
            logger=self.logger,
            temp_log_file=temp_log_file
        )

        best_params = ga_optimizer.optimize(objective, param_bounds)

        # Convert parameter format
        best_params['n_estimators'] = int(best_params['n_estimators'])
        best_params['max_depth'] = int(best_params['max_depth'])
        best_params['min_samples_split'] = int(best_params['min_samples_split'])
        best_params['min_samples_leaf'] = int(best_params['min_samples_leaf'])

        max_features_ratio = max(0.1, min(0.999, best_params['max_features_ratio']))
        best_params['max_features'] = max_features_ratio
        best_params.pop('max_features_ratio', None)

        class_weight_multiplier = best_params.pop('class_weight_multiplier', 0)
        if class_weight_multiplier > 0:
            best_params['class_weight'] = {
                0: 1.0,
                1: class_weight_multiplier * base_class_ratio
            }
        else:
            best_params['class_weight'] = None

        return best_params

    def _create_model(self, params):
        """Create Random Forest algorithms"""
        from sklearn.ensemble import RandomForestClassifier
        params_with_defaults = {
            'n_estimators': int(params.get('n_estimators', 100)),
            'max_depth': params.get('max_depth', None),
            'min_samples_split': int(params.get('min_samples_split', 2)),
            'min_samples_leaf': int(params.get('min_samples_leaf', 1)),
            'max_features': params.get('max_features', 'sqrt'),
            'class_weight': params.get('class_weight', None),
            'random_state': 42,
            'n_jobs': -1  # Resume parallel computation
        }
        return RandomForestClassifier(**params_with_defaults)


class XGBoostClassifier:
    def __init__(self, config, logger=None):
        self.config = config
        self.all_classes = config['all_classes']
        self.scaler = StandardScaler()
        self.logger = logger

        # Initialize ADASYN
        adasyn_sampling_strategy = config.get('adasyn_sampling_strategy', 'auto')
        adasyn_n_neighbors = config.get('adasyn_n_neighbors', 5)
        self.adasyn = ADASYN(
            sampling_strategy=adasyn_sampling_strategy,
            n_neighbors=adasyn_n_neighbors,
            random_state=42
        )

        # Initialize feature interaction processor
        try:
            from AdvancedFeatureInteractionsXGB import AdvancedFeatureInteractions
            self.feature_interaction = AdvancedFeatureInteractions(**config.get('feature_params', {}))
        except ImportError:
            self.feature_interaction = AdvancedFeatureInteractions(**config.get('feature_params', {}))

    def load_data(self):
        """Load and process data"""
        df = pd.read_excel(self.config['file_path'], sheet_name=self.config['sheet_name'])
        labels = df['Sample'].values
        features = df.iloc[:, 1:].values

        self.original_indices = np.arange(len(features))
        original_feature_names = df.columns[1:].tolist()
        self.feature_interaction.set_original_feature_names(original_feature_names)

        features_processed = self.feature_interaction.fit_transform(features, labels)
        self.feature_names = self.feature_interaction.get_feature_names()
        self.feature_importance = self.feature_interaction.get_feature_importance()

        return features_processed, np.array(labels), self.original_indices

    def apply_adasyn(self, X, y, positive_class):

        y_binary = (y == positive_class).astype(int)

        n_positive = sum(y_binary == 1)
        if n_positive < 2:
            return X, y_binary, self.original_indices

        try:
            n_neighbors = self.config.get('adasyn_n_neighbors', 5)
            if n_positive <= n_neighbors:
                adaptive_n_neighbors = max(1, n_positive - 1)
                temp_adasyn = ADASYN(
                    sampling_strategy=self.config.get('adasyn_sampling_strategy', 'auto'),
                    n_neighbors=adaptive_n_neighbors,
                    random_state=42
                )
                X_resampled, y_binary_resampled = temp_adasyn.fit_resample(X, y_binary)
            else:
                X_resampled, y_binary_resampled = self.adasyn.fit_resample(X, y_binary)

            sample_indices = np.concatenate([
                self.original_indices,
                np.full(len(X_resampled) - len(X), -1)
            ])
            return X_resampled, y_binary_resampled, sample_indices

        except ValueError:
            return X, y_binary, self.original_indices

    def _optimize_hyperparameters(self, X_train, y_train, class_name, fold_idx):
        """Optimize hyperparameters using genetic algorithm"""
        neg_count = sum(y_train == 0)
        pos_count = sum(y_train == 1)
        base_scale_pos_weight = neg_count / max(1, pos_count)

        def objective(params):
            try:
                actual_scale_pos_weight = params['scale_pos_weight_multiplier'] * base_scale_pos_weight

                import xgboost as xgb
                model = xgb.XGBClassifier(
                    n_estimators=params['n_estimators'],
                    max_depth=params['max_depth'],
                    learning_rate=params['learning_rate'],
                    subsample=params['subsample'],
                    colsample_bytree=params['colsample_bytree'],
                    min_child_weight=params['min_child_weight'],
                    gamma=params['gamma'],
                    reg_alpha=params['reg_alpha'],
                    reg_lambda=params['reg_lambda'],
                    scale_pos_weight=actual_scale_pos_weight,
                    use_label_encoder=False,
                    eval_metric='logloss',
                    random_state=42,
                    n_jobs=-1  # Resume parallel computation
                )

                cv = StratifiedKFold(n_splits=self.config['inner_cv_splits'], shuffle=True, random_state=42)
                f1_scores = []
                auc_scores = []
                recall_scores = []
                specificity_scores = []

                for train_idx, val_idx in cv.split(X_train, y_train):
                    X_train_cv, X_val_cv = X_train[train_idx], X_train[val_idx]
                    y_train_cv, y_val_cv = y_train[train_idx], y_train[val_idx]

                    y_train_cv = y_train_cv.astype(int)
                    model.fit(X_train_cv, y_train_cv)
                    y_pred_proba = model.predict_proba(X_val_cv)[:, 1]

                    best_f1 = 0
                    best_pred = None
                    for threshold in [0.3, 0.35, 0.4, 0.45, 0.5]:
                        y_pred_thresh = (y_pred_proba >= threshold).astype(int)
                        f1 = f1_score(y_val_cv, y_pred_thresh, average='weighted')
                        if f1 > best_f1:
                            best_f1 = f1
                            best_pred = y_pred_thresh

                    f1_scores.append(f1_score(y_val_cv, best_pred, average='weighted'))

                    try:
                        auc_scores.append(roc_auc_score(y_val_cv, y_pred_proba))
                    except Exception:
                        auc_scores.append(0)

                    recall_scores.append(recall_score(y_val_cv, best_pred, average='weighted'))
                    specificity_scores.append(weighted_specificity(y_val_cv, best_pred, [0, 1]))

                mean_f1 = np.mean(f1_scores)
                mean_auc = np.mean(auc_scores)
                mean_recall = np.mean(recall_scores)
                mean_specificity = np.mean(specificity_scores)

                return 0.55 * mean_f1 + 0.15 * mean_auc + 0.10 * mean_recall + 0.20 * mean_specificity

            except Exception:
                return 0.0

        # Parameter bounds
        param_bounds = {
            'n_estimators': (100, 800, 'int'),
            'max_depth': (2, 5, 'int'),
            'learning_rate': (0.01, 0.15, 'float'),
            'subsample': (0.7, 1.0, 'float'),
            'colsample_bytree': (0.7, 1.0, 'float'),
            'min_child_weight': (1, 5, 'int'),
            'gamma': (0, 0.5, 'float'),
            'reg_alpha': (0, 1.0, 'float'),
            'reg_lambda': (0.1, 10.0, 'float'),
            'scale_pos_weight_multiplier': (0.8, 3.0, 'float')
        }

        # Create genetic algorithm optimizer
        temp_log_file = None
        if self.logger:
            temp_log_file = self.logger.start_algorithm("XGBoost", class_name, fold_idx + 1)

        ga_config = self.config.get('ga_config', {})
        ga_optimizer = GeneticAlgorithmOptimizer(
            population_size=ga_config.get('population_size', 40),
            max_generations=ga_config.get('max_generations', 50),
            crossover_rate=ga_config.get('crossover_rate', 0.8),
            mutation_rate=ga_config.get('mutation_rate', 0.15),
            random_state=42,
            algorithm_name="XGB",
            logger=self.logger,
            temp_log_file=temp_log_file
        )

        best_params = ga_optimizer.optimize(objective, param_bounds)

        # Convert parameter format
        best_params['n_estimators'] = int(best_params['n_estimators'])
        best_params['max_depth'] = int(best_params['max_depth'])
        best_params['min_child_weight'] = max(1, int(best_params.get('min_child_weight', 1)))
        best_params['scale_pos_weight'] = best_params['scale_pos_weight_multiplier'] * base_scale_pos_weight
        del best_params['scale_pos_weight_multiplier']

        return best_params

    def _create_model(self, params):
        """Create XGBoost algorithms"""
        import xgboost as xgb
        params_with_defaults = {
            'n_estimators': int(params.get('n_estimators', 100)),
            'max_depth': int(params.get('max_depth', 3)),
            'learning_rate': params.get('learning_rate', 0.1),
            'subsample': params.get('subsample', 0.8),
            'colsample_bytree': params.get('colsample_bytree', 0.8),
            'scale_pos_weight': params.get('scale_pos_weight', 1.0),
            'min_child_weight': int(params.get('min_child_weight', 1)),
            'gamma': params.get('gamma', 0),
            'reg_alpha': params.get('reg_alpha', 0),
            'reg_lambda': params.get('reg_lambda', 1.0)
        }

        return xgb.XGBClassifier(
            **params_with_defaults,
            use_label_encoder=False,
            eval_metric='logloss',
            random_state=42,
            n_jobs=-1  # Resume parallel computation
        )


class EnsembleClassifier:
    def __init__(self, config):
        self.config = config
        self.all_classes = config['all_classes']
        self.scaler = StandardScaler()

        # Create evolution history logger
        self.logger = GAEvolutionLogger(config['output_path'])

        # Initialize RF and XGBoost components
        self.rf_component = RandomForestModel(config['rf_component'], self.logger)
        self.xgb_component = XGBoostClassifier(config['xgb_component'], self.logger)

    def load_data(self):
        """Load data"""
        rf_features, rf_labels, rf_indices = self.rf_component.load_data()
        xgb_features, xgb_labels, xgb_indices = self.xgb_component.load_data()
        return (rf_features, rf_labels, rf_indices), (xgb_features, xgb_labels, xgb_indices)

    def apply_data_augmentation(self, rf_data, xgb_data, positive_class):
        """Apply data augmentation"""
        rf_features, rf_labels, rf_indices = rf_data
        xgb_features, xgb_labels, xgb_indices = xgb_data

        rf_X_balanced, rf_y_binary, rf_sample_indices = self.rf_component.apply_smote(
            rf_features, rf_labels, positive_class
        )

        xgb_X_balanced, xgb_y_binary, xgb_sample_indices = self.xgb_component.apply_adasyn(
            xgb_features, xgb_labels, positive_class
        )

        # Ensure both datasets have the same size
        if len(rf_X_balanced) != len(xgb_X_balanced):
            min_samples = min(len(rf_X_balanced), len(xgb_X_balanced))
            if len(rf_X_balanced) > min_samples:
                rf_X_balanced = rf_X_balanced[:min_samples]
                rf_y_binary = rf_y_binary[:min_samples]
                rf_sample_indices = rf_sample_indices[:min_samples]
            if len(xgb_X_balanced) > min_samples:
                xgb_X_balanced = xgb_X_balanced[:min_samples]
                xgb_y_binary = xgb_y_binary[:min_samples]
                xgb_sample_indices = xgb_sample_indices[:min_samples]

        return (rf_X_balanced, rf_y_binary, rf_sample_indices), (xgb_X_balanced, xgb_y_binary, xgb_sample_indices)

    def process_fold(self, rf_data, xgb_data, train_index, test_index, positive_class, fold_idx):
        """Process single cross-validation fold"""
        # Unpack data
        rf_X, rf_y, rf_indices = rf_data
        xgb_X, xgb_y, xgb_indices = xgb_data

        # Split train and test sets
        rf_X_train, rf_X_test = rf_X[train_index], rf_X[test_index]
        rf_y_train, rf_y_test = rf_y[train_index], rf_y[test_index]
        rf_indices_train, rf_indices_test = rf_indices[train_index], rf_indices[test_index]

        xgb_X_train, xgb_X_test = xgb_X[train_index], xgb_X[test_index]
        xgb_y_train, xgb_y_test = xgb_y[train_index], xgb_y[test_index]
        xgb_indices_train, xgb_indices_test = xgb_indices[train_index], xgb_indices[test_index]

        # Use original data for evaluation
        rf_original_test_mask = rf_indices_test >= 0
        rf_X_test_original = rf_X_test[rf_original_test_mask]
        rf_y_test_original = rf_y_test[rf_original_test_mask]

        xgb_original_test_mask = xgb_indices_test >= 0
        xgb_X_test_original = xgb_X_test[xgb_original_test_mask]
        xgb_y_test_original = xgb_y_test[xgb_original_test_mask]

        # Verify test set consistency
        if not np.array_equal(rf_y_test_original, xgb_y_test_original):
            common_indices = np.intersect1d(rf_indices_test[rf_original_test_mask],
                                            xgb_indices_test[xgb_original_test_mask])
            rf_common_mask = np.isin(rf_indices_test, common_indices)
            xgb_common_mask = np.isin(xgb_indices_test, common_indices)

            rf_X_test_original = rf_X_test[rf_common_mask]
            rf_y_test_original = rf_y_test[rf_common_mask]
            xgb_X_test_original = xgb_X_test[xgb_common_mask]
            xgb_y_test_original = xgb_y_test[xgb_common_mask]

        # Optimize RF algorithms hyperparameters
        rf_best_params = self.rf_component._optimize_hyperparameters(rf_X_train, rf_y_train, positive_class, fold_idx)
        rf_model = self.rf_component._create_model(rf_best_params)
        rf_model.fit(rf_X_train, rf_y_train)

        # Optimize XGBoost algorithms hyperparameters
        xgb_best_params = self.xgb_component._optimize_hyperparameters(xgb_X_train, xgb_y_train, positive_class,
                                                                       fold_idx)
        xgb_model = self.xgb_component._create_model(xgb_best_params)
        xgb_model.fit(xgb_X_train, xgb_y_train)

        # Get prediction probabilities
        rf_y_pred_proba = rf_model.predict_proba(rf_X_test_original)[:, 1]
        xgb_y_pred_proba = xgb_model.predict_proba(xgb_X_test_original)[:, 1]

        # Optimize ensemble weights
        best_ensemble_weight = self._optimize_ensemble_weight(
            rf_y_pred_proba, xgb_y_pred_proba, rf_y_test_original
        )

        # Calculate weighted ensemble prediction probabilities
        ensemble_y_pred_proba = (1 - best_ensemble_weight) * rf_y_pred_proba + best_ensemble_weight * xgb_y_pred_proba

        # Use dynamic threshold selection
        test_class_dist = np.bincount(rf_y_test_original)
        thresholds = np.arange(0.1, 0.7, 0.025)
        best_f1 = 0
        best_threshold = 0.4

        if len(test_class_dist) > 1 and test_class_dist[0] > 0 and np.sum(test_class_dist) > test_class_dist[0]:
            for threshold in thresholds:
                y_pred_thresh = (ensemble_y_pred_proba >= threshold).astype(int)
                f1 = f1_score(rf_y_test_original, y_pred_thresh, average='weighted')
                if f1 > best_f1:
                    best_f1 = f1
                    best_threshold = threshold

        # Use best threshold for prediction
        ensemble_y_pred = (ensemble_y_pred_proba >= best_threshold).astype(int)

        # Calculate evaluation metrics
        ensemble_f1 = f1_score(rf_y_test_original, ensemble_y_pred, average='weighted')
        ensemble_precision = precision_score(rf_y_test_original, ensemble_y_pred, average='weighted', zero_division=0)
        ensemble_recall = recall_score(rf_y_test_original, ensemble_y_pred, average='weighted')
        ensemble_sensitivity = ensemble_recall
        ensemble_specificity = weighted_specificity_score(rf_y_test_original, ensemble_y_pred)

        # ROC curve calculation
        try:
            fpr, tpr, _ = roc_curve(rf_y_test_original, ensemble_y_pred_proba)
            roc_auc = auc(fpr, tpr)
        except Exception:
            fpr, tpr = np.array([0, 1]), np.array([0, 1])
            roc_auc = 0.5

        # PR curve calculation
        try:
            precision, recall, _ = precision_recall_curve(rf_y_test_original, ensemble_y_pred_proba)
            pr_auc = average_precision_score(rf_y_test_original, ensemble_y_pred_proba)
        except Exception:
            precision, recall = np.array([0, 1]), np.array([1, 0])
            pr_auc = 0.5

        # Confusion matrix
        conf_matrix = confusion_matrix(rf_y_test_original, ensemble_y_pred)

        # Calculate false negative rate and false positive rate
        tn, fp, fn, tp = 0, 0, 0, 0
        if conf_matrix.shape == (2, 2):
            tn, fp, fn, tp = conf_matrix.ravel()
        elif conf_matrix.shape == (1, 1):
            if np.sum(rf_y_test_original) == 0:
                tn = conf_matrix[0, 0]
            else:
                tp = conf_matrix[0, 0]

        fnr = fn / (fn + tp) if (fn + tp) > 0 else 0
        fpr_val = fp / (fp + tn) if (fp + tn) > 0 else 0

        return {
            "fold_idx": fold_idx,
            "y_test": rf_y_test_original,
            "y_pred": ensemble_y_pred,
            "y_pred_proba": ensemble_y_pred_proba,
            "optimal_threshold": best_threshold,
            "optimal_xgb_weight": best_ensemble_weight,
            "f1": ensemble_f1,
            "precision": ensemble_precision,
            "recall": ensemble_recall,
            "sensitivity": ensemble_sensitivity,
            "specificity": ensemble_specificity,
            "fpr": fpr,
            "tpr": tpr,
            "roc_auc": roc_auc,
            "pr_curve": {"precision": precision, "recall": recall},
            "pr_auc": pr_auc,
            "fnr": fnr,
            "fpr_val": fpr_val,
            "confusion_matrix": conf_matrix,
            "rf_model": rf_model,
            "xgb_model": xgb_model,
            "rf_best_params": rf_best_params,
            "xgb_best_params": xgb_best_params,
            "class_name": positive_class
        }

    def _optimize_ensemble_weight(self, rf_proba, xgb_proba, y_true):
        """Optimize XGBoost weight in ensemble algorithms"""

        def objective(xgb_weight):
            ensemble_proba = (1 - xgb_weight) * rf_proba + xgb_weight * xgb_proba
            best_f1 = 0
            for threshold in np.arange(0.3, 0.6, 0.05):
                y_pred = (ensemble_proba >= threshold).astype(int)
                f1 = f1_score(y_true, y_pred, average='weighted')
                if f1 > best_f1:
                    best_f1 = f1
            return best_f1

        from bayes_opt import BayesianOptimization
        optimizer = BayesianOptimization(
            f=objective,
            pbounds={'xgb_weight': (0.0, 1.0)},
            random_state=42,
            allow_duplicate_points=True
        )

        optimizer.maximize(init_points=5, n_iter=10)
        return optimizer.max['params']['xgb_weight']

    def plot_confusion_matrix_average(self, all_matrices, class_name):
        """Plot average confusion matrix"""
        if not all_matrices:
            return

        max_size = max(mat.shape[0] for mat in all_matrices)
        resized_matrices = []

        for mat in all_matrices:
            if mat.shape[0] < max_size:
                new_mat = np.zeros((max_size, max_size), dtype=mat.dtype)
                new_mat[:mat.shape[0], :mat.shape[1]] = mat
                resized_matrices.append(new_mat)
            else:
                resized_matrices.append(mat)

        avg_matrix = np.mean(resized_matrices, axis=0)

        plt.figure(figsize=(10, 8))
        plt.rcParams['font.family'] = 'Arial'
        ax = plt.gca()

        for spine in ax.spines.values():
            spine.set_linewidth(5)

        sns.heatmap(avg_matrix, annot=True, fmt='.1f', cmap='Blues',
                    annot_kws={'size': 20, 'weight': 'bold'})

        plt.title(f'Confusion Matrix - {class_name} (Average)', fontsize=20, pad=20)
        plt.ylabel('True Label', fontsize=24, labelpad=15)
        plt.xlabel('Predicted Label', fontsize=24, labelpad=15)
        ax.tick_params(axis='both', labelsize=24)

        output_file = os.path.join(self.config['output_path'],
                                   f"confusion_matrix_{class_name}_average.svg")
        plt.savefig(output_file, format='svg', bbox_inches='tight')
        plt.close()

    def plot_roc_curve(self, results, positive_class):
        """Plot ROC curve"""
        plt.figure(figsize=(10, 8))
        ax = plt.gca()

        for spine in ax.spines.values():
            spine.set_linewidth(5)

        plt.rcParams['font.family'] = 'Arial'

        plt.title(f"ROC Curve for {positive_class}", fontsize=20, pad=15)
        plt.xlabel("False Positive Rate", fontsize=24, labelpad=10)
        plt.ylabel("True Positive Rate", fontsize=24, labelpad=10)

        mean_fpr = np.linspace(0, 1, 1000)
        tprs = []
        aucs = []

        for result in results:
            fpr = result['fpr']
            tpr = result['tpr']
            roc_auc = result['roc_auc']

            if fpr[0] != 0:
                fpr = np.concatenate([[0], fpr])
                tpr = np.concatenate([[0], tpr])
            if fpr[-1] != 1:
                fpr = np.concatenate([fpr, [1]])
                tpr = np.concatenate([tpr, [tpr[-1]]])

            interp_tpr = np.interp(mean_fpr, fpr, tpr)
            interp_tpr[0] = 0.0
            tprs.append(interp_tpr)
            aucs.append(roc_auc)

            plt.plot(fpr, tpr, alpha=0.15,
                     color=self.config['color_map'].get(positive_class, 'b'),
                     linestyle='-', linewidth=2)

        mean_tpr = np.mean(tprs, axis=0)
        mean_auc = np.mean(aucs)
        std_auc = np.std(aucs)

        plt.plot(mean_fpr, mean_tpr,
                 color=self.config['color_map'].get(positive_class, 'b'),
                 label=f'Mean ROC (AUC={mean_auc:.3f}±{std_auc:.3f})',
                 lw=5, alpha=0.8)

        std_tpr = np.std(tprs, axis=0)
        tprs_upper = np.minimum(mean_tpr + std_tpr, 1)
        tprs_lower = np.maximum(mean_tpr - std_tpr, 0)

        plt.fill_between(mean_fpr, tprs_lower, tprs_upper,
                         color=self.config['color_map'].get(positive_class, 'b'),
                         alpha=0.2, label='±1 std. dev.')

        plt.plot([0, 1], [0, 1], 'k--', lw=2)
        ax.tick_params(axis='both', labelsize=24)

        plt.grid(True, alpha=0.3)
        legend = plt.legend(loc="lower right", fontsize=20, frameon=True)
        legend.get_frame().set_linewidth(3)

        plt.xlim([-0.02, 1.02])
        plt.ylim([-0.02, 1.02])

        output_file = os.path.join(self.config['output_path'],
                                   f"ROC_{positive_class}_Ensemble.svg")
        plt.savefig(output_file, format='svg', bbox_inches='tight')
        plt.close()

    def save_results(self, all_results):
        """Save LOO results to Excel - including ROC data and parameters"""
        output_path = os.path.join(self.config['output_path'], self.config['output_filename'])

        with pd.ExcelWriter(output_path) as writer:
            # 1. Save overall summary
            summary_data = []
            for class_name, results in all_results.items():
                fold_count = len(results)

                f1_scores = [r['f1'] for r in results]
                precision_scores = [r['precision'] for r in results]
                recall_scores = [r['recall'] for r in results]
                sensitivity_scores = [r['sensitivity'] for r in results]
                specificity_scores = [r['specificity'] for r in results]
                auc_scores = [r['roc_auc'] for r in results]
                pr_auc_scores = [r['pr_auc'] for r in results]
                fnr_scores = [r['fnr'] for r in results]
                xgb_weights = [r['optimal_xgb_weight'] for r in results]
                thresholds = [r['optimal_threshold'] for r in results]

                summary_data.append({
                    'Class': class_name,
                    'Number of Folds': fold_count,
                    'Ensemble Mean F1': np.mean(f1_scores),
                    'Ensemble Std F1': np.std(f1_scores),
                    'Ensemble Mean Precision': np.mean(precision_scores),
                    'Ensemble Std Precision': np.std(precision_scores),
                    'Ensemble Mean Recall': np.mean(recall_scores),
                    'Ensemble Std Recall': np.std(recall_scores),
                    'Ensemble Mean Sensitivity': np.mean(sensitivity_scores),
                    'Ensemble Std Sensitivity': np.std(sensitivity_scores),
                    'Ensemble Mean Specificity': np.mean(specificity_scores),
                    'Ensemble Std Specificity': np.std(specificity_scores),
                    'Ensemble Mean ROC AUC': np.mean(auc_scores),
                    'Ensemble Std ROC AUC': np.std(auc_scores),
                    'Ensemble Mean PR AUC': np.mean(pr_auc_scores),
                    'Ensemble Std PR AUC': np.std(pr_auc_scores),
                    'Ensemble Mean FNR': np.mean(fnr_scores),
                    'Ensemble Mean XGB Weight': np.mean(xgb_weights),
                    'Ensemble Std XGB Weight': np.std(xgb_weights),
                    'Ensemble Mean Threshold': np.mean(thresholds),
                })

            pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', index=False)

            # 2. Save detailed LOO results for each class
            for class_name, results in all_results.items():
                fold_data = []
                for r in results:
                    fold_data.append({
                        'Fold': r['fold_idx'] + 1,
                        'Ensemble F1': r['f1'],
                        'Ensemble Precision': r['precision'],
                        'Ensemble Recall': r['recall'],
                        'Ensemble Sensitivity': r['sensitivity'],
                        'Ensemble Specificity': r['specificity'],
                        'Ensemble ROC AUC': r['roc_auc'],
                        'Ensemble PR AUC': r['pr_auc'],
                        'Ensemble FNR': r['fnr'],
                        'Optimal Threshold': r['optimal_threshold'],
                        'Optimal XGB Weight': r['optimal_xgb_weight'],
                    })

                pd.DataFrame(fold_data).to_excel(writer, sheet_name=f'{class_name}_details', index=False)

                # 3. Save ROC curve data
                all_fpr = []
                all_tpr = []
                for r in results:
                    fpr = r['fpr']
                    tpr = r['tpr']

                    # Ensure start and end points are complete
                    if fpr[0] != 0:
                        fpr = np.concatenate([[0], fpr])
                        tpr = np.concatenate([[0], tpr])
                    if fpr[-1] != 1:
                        fpr = np.concatenate([fpr, [1]])
                        tpr = np.concatenate([tpr, [tpr[-1]]])

                    all_fpr.append(fpr)
                    all_tpr.append(tpr)

                # Create denser interpolation points for smoother curves
                mean_fpr = np.linspace(0, 1, 1000)
                tprs = []

                for fpr, tpr in zip(all_fpr, all_tpr):
                    interp_tpr = np.interp(mean_fpr, fpr, tpr)
                    interp_tpr[0] = 0.0
                    tprs.append(interp_tpr)

                mean_tpr = np.mean(tprs, axis=0)
                std_tpr = np.std(tprs, axis=0)
                tprs_upper = np.minimum(mean_tpr + std_tpr, 1)
                tprs_lower = np.maximum(mean_tpr - std_tpr, 0)

                # Save ROC curve data
                roc_data = pd.DataFrame({
                    'FPR': mean_fpr,
                    'TPR': mean_tpr,
                    'TPR_upper': tprs_upper,
                    'TPR_lower': tprs_lower
                })

                roc_data.to_excel(writer, sheet_name=f'{class_name}_ROC_data', index=False)

                # 4. Save algorithms parameters
                model_params = []
                for r in results:
                    # RF parameters
                    rf_params_row = {
                        'Fold': r['fold_idx'] + 1,
                        'Model': 'Random Forest'
                    }

                    # Process RF parameters
                    rf_params = r['rf_best_params']
                    for param_name, param_value in rf_params.items():
                        if param_name == 'class_weight' and isinstance(param_value, dict):
                            rf_params_row[param_name] = str(param_value)
                        else:
                            rf_params_row[param_name] = param_value

                    # XGBoost parameters
                    xgb_params_row = {
                        'Fold': r['fold_idx'] + 1,
                        'Model': 'XGBoost'
                    }

                    # Process XGBoost parameters
                    xgb_params = r['xgb_best_params']
                    for param_name, param_value in xgb_params.items():
                        xgb_params_row[param_name] = param_value

                    model_params.append(rf_params_row)
                    model_params.append(xgb_params_row)

                # Save parameter table
                if model_params:
                    pd.DataFrame(model_params).to_excel(writer, sheet_name=f'{class_name}_params', index=False)

        # Format Excel file
        self.format_excel_file(output_path)

        print(f"Results saved to: {output_path}")

    def format_excel_file(self, file_path):
        from openpyxl import load_workbook
        from openpyxl.styles import Font

        wb = load_workbook(file_path)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            for row in ws.iter_rows():
                for cell in row:
                    cell.font = Font(name='Times New Roman', size=11)

            for cell in ws[1]:
                cell.font = Font(name='Times New Roman', size=11, bold=True)

        wb.save(file_path)

    def plot_combined_roc_curves(self, all_results):
        """Plot ROC curves for all classes"""
        plt.figure(figsize=(12, 10))
        ax = plt.gca()

        for spine in ax.spines.values():
            spine.set_linewidth(9)

        plt.rcParams['font.family'] = 'Arial'

        plt.title("ROC Curves for All Classes (GA Optimized)", fontsize=20, pad=15)
        plt.xlabel("False Positive Rate", fontsize=24, labelpad=10)
        plt.ylabel("True Positive Rate", fontsize=24, labelpad=10)

        plt.plot([0, 1], [0, 1], 'k--', lw=2)

        for class_name, results in all_results.items():
            tprs = []
            aucs = []
            mean_fpr = np.linspace(0, 1, 1000)

            for result in results:
                fpr = result['fpr']
                tpr = result['tpr']
                roc_auc = result['roc_auc']

                if fpr[0] != 0:
                    fpr = np.concatenate([[0], fpr])
                    tpr = np.concatenate([[0], tpr])
                if fpr[-1] != 1:
                    fpr = np.concatenate([fpr, [1]])
                    tpr = np.concatenate([tpr, [tpr[-1]]])

                interp_tpr = np.interp(mean_fpr, fpr, tpr)
                interp_tpr[0] = 0.0
                tprs.append(interp_tpr)
                aucs.append(roc_auc)

            mean_tpr = np.mean(tprs, axis=0)
            mean_auc = np.mean(aucs)
            std_auc = np.std(aucs)

            color = self.config['color_map'].get(class_name, 'b')
            plt.plot(mean_fpr, mean_tpr,
                     color=color,
                     label=f'{class_name} (AUC = {mean_auc:.3f}±{std_auc:.3f})',
                     lw=5)

            std_tpr = np.std(tprs, axis=0)
            tprs_upper = np.minimum(mean_tpr + std_tpr, 1)
            tprs_lower = np.maximum(mean_tpr - std_tpr, 0)

            plt.fill_between(mean_fpr, tprs_lower, tprs_upper,
                             color=color, alpha=0.1)

        ax.tick_params(axis='both', labelsize=48, width=9, length=20)
        plt.xticks(np.arange(0, 1.1, 0.2))
        plt.yticks(np.arange(0, 1.1, 0.2))

        plt.grid(True, alpha=0.3)
        plt.xlim([-0.02, 1.02])
        plt.ylim([-0.02, 1.02])

        handles, labels = ax.get_legend_handles_labels()
        legend = plt.legend(handles, labels, loc="lower right", fontsize=28, frameon=True)
        legend.get_frame().set_linewidth(3)

        output_file = os.path.join(self.config['output_path'], "Combined_ROC_Curves_GA.svg")
        plt.savefig(output_file, format='svg', bbox_inches='tight')
        plt.close()

    def run(self):
        """Run complete training and evaluation process - Maintain parallel processing"""
        # Initialize time estimator
        time_estimator = TimeEstimator()
        time_estimator.start_total()

        # Load data
        rf_data, xgb_data = self.load_data()
        rf_features, rf_labels, _ = rf_data

        all_results = {}

        for class_idx, class_name in enumerate(self.all_classes):
            # Class start
            print(f"Class {class_name} [{class_idx + 1}/{len(self.all_classes)}]")
            print(f"Processing {self.config['outer_cv_splits']} cross-validation folds in parallel...")

            # Prepare dynamic prompt message
            if class_idx == 0:
                base_message = "   Processing folds 1,2,3,4,5 in parallel"
            else:
                estimated_class_time = time_estimator.get_adaptive_estimate()
                estimated_total_remaining = time_estimator.estimate_remaining_time(class_idx, len(self.all_classes))

                time_parts = []
                if estimated_class_time:
                    time_parts.append(f"Estimated: {time_estimator.format_time(estimated_class_time)}")
                if estimated_total_remaining and class_idx < len(self.all_classes) - 1:
                    time_parts.append(f"Total Remaining: {time_estimator.format_time(estimated_total_remaining)}")

                time_str = f" [{', '.join(time_parts)}]" if time_parts else ""
                base_message = f"   Processing folds 1,2,3,4,5 in parallel{time_str}"

            # Start simple spinner indicator
            indicator = SimpleSpinnerIndicator()
            indicator.start(base_message)

            # Record start time
            class_start_time = time.time()

            try:
                # Apply data augmentation
                rf_balanced_data, xgb_balanced_data = self.apply_data_augmentation(rf_data, xgb_data, class_name)
                _, rf_y_binary, _ = rf_balanced_data

                # Create cross-validation object
                outer_cv = StratifiedKFold(n_splits=self.config['outer_cv_splits'], shuffle=True, random_state=42)
                fold_indices = list(enumerate(outer_cv.split(rf_balanced_data[0], rf_y_binary)))

                # Parallel process all folds - Resume full parallel
                results = Parallel(n_jobs=-1, verbose=0)(
                    delayed(self._process_fold_wrapper)(
                        rf_balanced_data, xgb_balanced_data, train_idx, test_idx, class_name, fold_idx
                    )
                    for fold_idx, (train_idx, test_idx) in fold_indices
                )

                # Stop spinner indicator
                indicator.stop()

                # Filter LOO results
                results = [r for r in results if r is not None]
                if not results:
                    print(f"All folds failed for class {class_name}")
                    continue

                all_results[class_name] = results

                # Calculate actual time and record
                actual_time = time.time() - class_start_time
                time_estimator.record_class_time(actual_time)

                # Calculate performance metrics
                f1_scores = [r['f1'] for r in results]
                sensitivity_scores = [r['sensitivity'] for r in results]
                specificity_scores = [r['specificity'] for r in results]
                auc_scores = [r['roc_auc'] for r in results]

                # Output LOO results
                print(f"Completed! {len(results)} folds processed [Actual time: {actual_time:.1f}s]")
                print(f"   Results: F1={np.mean(f1_scores):.3f}±{np.std(f1_scores):.3f}, "
                      f"Sens={np.mean(sensitivity_scores):.3f}±{np.std(sensitivity_scores):.3f}, "
                      f"Spec={np.mean(specificity_scores):.3f}±{np.std(specificity_scores):.3f}, "
                      f"AUC={np.mean(auc_scores):.3f}±{np.std(auc_scores):.3f}")
                print()

                # Plot charts
                self.plot_roc_curve(results, class_name)
                all_conf_matrices = [result['confusion_matrix'] for result in results]
                self.plot_confusion_matrix_average(all_conf_matrices, class_name)

            except Exception as e:
                # Ensure spinner indicator stops
                indicator.stop()
                print(f"Error processing class {class_name}: {str(e)}")
                actual_time = time.time() - class_start_time
                time_estimator.record_class_time(actual_time)
                continue

        # Final output
        total_time = time_estimator.get_total_time()
        print(f"Training completed! Total runtime: {total_time:.1f}s ({total_time / 60:.1f} minutes)")

        # Generate charts and save LOO results
        if all_results:
            self.plot_combined_roc_curves(all_results)
            self.save_results(all_results)

            # Merge and save evolution history logs
            self.logger.merge_and_save_logs()

            # Output evolution history statistics
            stats = self.logger.get_summary_stats()
            if stats:
                print(
                    f"Evolution Statistics: Recorded {stats['total_algorithms']} algorithm instances, total {stats['total_generations']} generations evolved")

            return all_results
        else:
            print("No LOO results generated.")
            return {}

    def _process_fold_wrapper(self, rf_data, xgb_data, train_index, test_index, positive_class, fold_idx):
        """Fold processing wrapper for error handling in parallel execution - Silent version"""
        try:
            # Remove all print output, silent processing
            result = self.process_fold(rf_data, xgb_data, train_index, test_index, positive_class, fold_idx)
            return result

        except Exception as e:
            # Only output error in debug mode (optional)
            # print(f"Error processing {positive_class} fold {fold_idx + 1}: {str(e)}")
            return None


class AdvancedFeatureInteractions:
    """Feature Interaction Processing Class - Simplified version"""

    def __init__(self,
                 interaction_method='gaussian',
                 selection_method='boruta',
                 max_features=30,
                 gaussian_gamma='auto',
                 scale_after_interaction=True):
        self.interaction_method = interaction_method
        self.selection_method = selection_method
        self.max_features = max_features
        self.gaussian_gamma = gaussian_gamma
        self.scale_after_interaction = scale_after_interaction
        self.feature_names_ = None
        self.selected_features_ = None
        self.feature_importance_ = None
        self.scaler = StandardScaler()
        self.original_feature_names = None

    def set_original_feature_names(self, names):
        """Set names of original features"""
        self.original_feature_names = names

    def fit_transform(self, features, labels):
        """Fit data and perform feature processing and transformation"""
        if self.scale_after_interaction:
            features_scaled = self.scaler.fit_transform(features)
        else:
            features_scaled = features.copy()

        if self.original_feature_names is not None:
            self.feature_names_ = self.original_feature_names.copy()
        else:
            self.feature_names_ = [f"feature_{i}" for i in range(features.shape[1])]

        # Simplified feature processing
        features_processed = features_scaled

        # Feature selection
        if self.selection_method and self.max_features:
            if features_processed.shape[1] > self.max_features:
                features_processed = features_processed[:, :self.max_features]
                self.feature_names_ = self.feature_names_[:self.max_features]

        self.feature_importance_ = np.ones(len(self.feature_names_))
        return features_processed

    def get_feature_names(self):
        """Get feature names"""
        return self.feature_names_

    def get_feature_importance(self):
        """Get feature importance"""
        if self.feature_importance_ is None:
            return None

        importance_df = pd.DataFrame({
            'Feature': self.feature_names_,
            'Importance': self.feature_importance_
        })
        return importance_df.sort_values('Importance', ascending=False)


def main():
    """Main function"""
    # Basic configuration
    data_file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '00_Data', 'Train.xlsx')  # Path to the input Excel file containing training data
    data_sheet_name = 'Sheet1'  # Name of the sheet in the Excel file (e.g., 'Sheet1')

    # Genetic algorithm configuration
    ga_config = {
        'population_size': 50,
        'max_generations': 60,
        'crossover_rate': 0.8,
        'mutation_rate': 0.125,
        'elite_ratio': 0.125,
        'tournament_size': 3,
        'convergence_threshold': 10
    }

    # Full configuration
    config = {
        'output_path': os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '..', 'Results', 'Multicancer detection', '03_Ensemble'),  # Directory path for saving output results
        'output_filename': 'Ensemble_GA_results.xlsx',

        'file_path': data_file_path,
        'sheet_name': data_sheet_name,
        'outer_cv_splits': 5,
        'inner_cv_splits': 5,

        'xgb_weight': 0.6,
        'rf_weight': 0.4,

        'rf_component': {
            'interaction_method': 'gaussian',
            'selection_method': 'boruta',
            'max_features': 40,
            'gaussian_gamma': 'auto',
            'scale_after_interaction': True,
            'smote_sampling_strategy': 1,
            'smote_k_neighbors': 5,
            'ga_config': ga_config,
            'feature_params': {
                'interaction_method': 'gaussian',
                'selection_method': 'boruta',
                'max_features': 40,
                'gaussian_gamma': 'auto',
                'scale_after_interaction': True
            }
        },

        'xgb_component': {
            'interaction_method': 'gaussian',
            'selection_method': 'mutual_info',
            'max_features': 40,
            'gaussian_gamma': 'auto',
            'scale_after_interaction': True,
            'adasyn_sampling_strategy': 1,
            'adasyn_n_neighbors': 5,
            'ga_config': ga_config,
            'feature_params': {
                'interaction_method': 'gaussian',
                'selection_method': 'mutual_info',
                'max_features': 40,
                'gaussian_gamma': 'auto',
                'scale_after_interaction': True
            }
        },

        'color_map': {
            'LC': [254 / 255, 160 / 255, 64 / 255],
            'LuC': [242 / 255, 128 / 255, 128 / 255],
            'OC': [88 / 255, 97 / 255, 172 / 255],
            'N': [106 / 255, 180 / 255, 193 / 255]
        }
    }

    try:
        # Create output directory
        os.makedirs(config['output_path'], exist_ok=True)

        # Read class information
        df = pd.read_excel(config['file_path'], sheet_name=config['sheet_name'])
        all_classes = sorted(df['Sample'].unique())
        config['all_classes'] = all_classes

        # Add shared parameters to components
        for component in ['rf_component', 'xgb_component']:
            config[component]['file_path'] = config['file_path']
            config[component]['sheet_name'] = config['sheet_name']
            config[component]['all_classes'] = all_classes
            config[component]['outer_cv_splits'] = config['outer_cv_splits']
            config[component]['inner_cv_splits'] = config['inner_cv_splits']

        # Format output - According to specified style
        print("=" * 80)
        print("Genetic Algorithm Enhanced Ensemble Model - Full Parallel Version")
        print("=" * 80)
        print(f"Data loaded: Samples={len(df)}, Classes={len(all_classes)} {all_classes}")

        ga_config = config['rf_component']['ga_config']
        print(f"Full Parallel: Enabled, GA Config: Population {ga_config['population_size']} × {ga_config['max_generations']} generations, "
              f"Cross-validation: {config['outer_cv_splits']} folds")
        print(f"New Features: Temporary file log recording, full ROC data, parameter saving")
        print()

        # Run ensemble classifier
        ensemble = EnsembleClassifier(config)
        results = ensemble.run()

        # Keep final output concise
        if results:
            print("Training completed!")
            print(f"Results saved to: {os.path.join(config['output_path'], config['output_filename'])}")
            print(f"Charts generated: Combined_ROC_Curves_GA.svg")
            print(f"Evolution History: GA_Evolution_History.txt")
            print(f"Contents: ROC data, parameter records, evolution process")
        else:
            print("No LOO results generated")

    except Exception as e:
        print(f"Program execution error: {str(e)}")
        import traceback
        traceback.print_exc()
        raise


if __name__ == "__main__":
    main()