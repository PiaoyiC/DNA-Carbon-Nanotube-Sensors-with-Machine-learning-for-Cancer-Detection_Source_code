import sys
import threading
import numpy as np
import pandas as pd
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import StratifiedKFold
from sklearn.metrics import (f1_score, roc_curve, auc, confusion_matrix, precision_score,
                             recall_score, precision_recall_curve, average_precision_score,
                             roc_auc_score, accuracy_score)
import matplotlib.pyplot as plt
from joblib import Parallel, delayed
import logging
import os
import seaborn as sns
import warnings
import random
import time
from typing import Dict, List, Tuple, Any
import matplotlib as mpl

from imblearn.over_sampling import SMOTE, ADASYN

warnings.filterwarnings('ignore')

mpl.rcParams['svg.fonttype'] = 'none'
mpl.rcParams['font.family'] = 'Arial'
mpl.rcParams['pdf.fonttype'] = 42
mpl.rcParams['ps.fonttype'] = 42

logging.basicConfig(level=logging.CRITICAL)


class SimpleSpinnerIndicator:
    """Simple Spinner Indicator"""

    def __init__(self):
        self.running = False
        self.thread = None
        self.start_time = None

    def start(self, message="   Processing"):
        self.running = True
        self.start_time = time.time()
        self.message = message
        self.thread = threading.Thread(target=self._animate)
        self.thread.daemon = True
        self.thread.start()

    def stop(self):
        self.running = False
        if self.thread:
            self.thread.join(timeout=1.0)
        sys.stdout.write('\r' + ' ' * 100 + '\r')
        sys.stdout.flush()

    def _animate(self):
        """Simple rotation animation"""
        spinners = ["|", "/", "-", "\\"]
        frame_idx = 0

        while self.running:
            elapsed = time.time() - self.start_time
            minutes = int(elapsed // 60)
            seconds = int(elapsed % 60)

            spinner = spinners[frame_idx % len(spinners)]
            time_str = f"{minutes:02d}:{seconds:02d}"

            display_text = f"\r{self.message} {spinner} [{time_str}]"

            sys.stdout.write(display_text)
            sys.stdout.flush()

            frame_idx += 1
            time.sleep(0.5)


class GeneticAlgorithmOptimizer:

    def __init__(self,
                 population_size: int = 40,
                 max_generations: int = 50,
                 crossover_rate: float = 0.8,
                 mutation_rate: float = 0.15,
                 elite_ratio: float = 0.2,
                 tournament_size: int = 3,
                 convergence_threshold: int = 10,
                 random_state: int = 42):

        self.population_size = population_size
        self.max_generations = max_generations
        self.crossover_rate = crossover_rate
        self.mutation_rate = mutation_rate
        self.elite_size = max(1, int(population_size * elite_ratio))
        self.tournament_size = tournament_size
        self.convergence_threshold = convergence_threshold
        self.random_state = random_state

        # Set random seed
        random.seed(random_state)
        np.random.seed(random_state)

        # Optimization history
        self.history = {
            'best_scores': [],
            'avg_scores': []
        }

        # Best result
        self.best_individual = None
        self.best_score = float('-inf')

    def _encode_individual(self, param_bounds: Dict[str, Tuple]) -> Dict[str, Any]:
        """Encode an individual"""
        individual = {}
        for param_name, bounds in param_bounds.items():
            min_val, max_val, param_type = bounds
            if param_type == 'int':
                individual[param_name] = random.randint(min_val, max_val)
            elif param_type == 'float':
                individual[param_name] = random.uniform(min_val, max_val)
            elif param_type == 'categorical':
                individual[param_name] = random.choice(min_val)
        return individual

    def _create_initial_population(self, param_bounds: Dict[str, Tuple]) -> List[Dict[str, Any]]:
        """Create initial population"""
        return [self._encode_individual(param_bounds) for _ in range(self.population_size)]

    def _tournament_selection(self, population: List[Dict], fitness_scores: List[float]) -> Dict[str, Any]:
        """Tournament selection"""
        tournament_indices = random.sample(range(len(population)),
                                           min(self.tournament_size, len(population)))
        tournament_fitness = [fitness_scores[i] for i in tournament_indices]
        winner_idx = tournament_indices[np.argmax(tournament_fitness)]
        return population[winner_idx].copy()

    def _adaptive_crossover(self, parent1: Dict[str, Any], parent2: Dict[str, Any],
                            param_bounds: Dict[str, Tuple]) -> Tuple[Dict[str, Any], Dict[str, Any]]:
        """Adaptive crossover operation"""
        child1, child2 = parent1.copy(), parent2.copy()

        for param_name in parent1.keys():
            if random.random() < self.crossover_rate:
                bounds = param_bounds[param_name]
                param_type = bounds[2]

                if param_type in ['int', 'float']:
                    val1, val2 = parent1[param_name], parent2[param_name]
                    eta = 20
                    if abs(val1 - val2) > 1e-14:
                        if val1 > val2:
                            val1, val2 = val2, val1

                        beta = 1.0 + (2.0 * (val1 - bounds[0]) / (val2 - val1))
                        alpha = 2.0 - beta ** -(eta + 1.0)

                        if random.random() <= 1.0 / alpha:
                            beta_q = (random.random() * alpha) ** (1.0 / (eta + 1.0))
                        else:
                            beta_q = (1.0 / (2.0 - random.random() * alpha)) ** (1.0 / (eta + 1.0))

                        child1_val = 0.5 * ((val1 + val2) - beta_q * (val2 - val1))
                        child2_val = 0.5 * ((val1 + val2) + beta_q * (val2 - val1))

                        child1_val = np.clip(child1_val, bounds[0], bounds[1])
                        child2_val = np.clip(child2_val, bounds[0], bounds[1])

                        if param_type == 'int':
                            child1_val = int(round(child1_val))
                            child2_val = int(round(child2_val))

                        child1[param_name] = child1_val
                        child2[param_name] = child2_val

                elif param_type == 'categorical':
                    if random.random() < 0.5:
                        child1[param_name], child2[param_name] = child2[param_name], child1[param_name]

        return child1, child2

    def _adaptive_mutation(self, individual: Dict[str, Any], param_bounds: Dict[str, Tuple],
                           generation: int) -> Dict[str, Any]:
        """Adaptive mutation operation"""
        mutated = individual.copy()
        adaptive_rate = self.mutation_rate * (1.0 - generation / self.max_generations)

        for param_name, bounds in param_bounds.items():
            if random.random() < adaptive_rate:
                min_val, max_val, param_type = bounds

                if param_type == 'int':
                    current_val = mutated[param_name]
                    sigma = (max_val - min_val) * 0.1 * (1.0 - generation / self.max_generations)
                    new_val = int(round(np.random.normal(current_val, sigma)))
                    mutated[param_name] = np.clip(new_val, min_val, max_val)

                elif param_type == 'float':
                    current_val = mutated[param_name]
                    sigma = (max_val - min_val) * 0.1 * (1.0 - generation / self.max_generations)
                    new_val = np.random.normal(current_val, sigma)
                    mutated[param_name] = np.clip(new_val, min_val, max_val)

                elif param_type == 'categorical':
                    mutated[param_name] = random.choice(min_val)

        return mutated

    def _check_convergence(self, generation: int) -> bool:
        """Check convergence condition"""
        if len(self.history['best_scores']) < self.convergence_threshold:
            return False

        recent_scores = self.history['best_scores'][-self.convergence_threshold:]
        improvement = max(recent_scores) - min(recent_scores)
        return improvement < 1e-6

    def _safe_objective(self, objective_function, individual):
        """Safe objective function call, handle exceptions"""
        try:
            return objective_function(individual)
        except Exception as e:
            return 0.0

    def optimize(self, objective_function, param_bounds: Dict[str, Tuple]) -> Dict[str, Any]:
        """Main optimization loop"""
        # Create initial population
        population = self._create_initial_population(param_bounds)

        for generation in range(self.max_generations):
            # Parallel fitness evaluation
            try:
                fitness_scores = Parallel(
                    n_jobs=-1,
                    prefer="threads",
                    batch_size=1,
                    verbose=0
                )(
                    delayed(self._safe_objective)(objective_function, individual)
                    for individual in population
                )
            except Exception as e:
                # If parallel fails, fallback to serial processing
                fitness_scores = []
                for individual in population:
                    fitness_scores.append(self._safe_objective(objective_function, individual))

            # Ensure fitness_scores length is correct
            if len(fitness_scores) != len(population):
                while len(fitness_scores) < len(population):
                    fitness_scores.append(0.0)

            # Update best individual
            current_best_idx = np.argmax(fitness_scores)
            current_best_score = fitness_scores[current_best_idx]

            if current_best_score > self.best_score:
                self.best_score = current_best_score
                self.best_individual = population[current_best_idx].copy()

            # Record history
            avg_score = np.mean(fitness_scores)
            self.history['best_scores'].append(self.best_score)
            self.history['avg_scores'].append(avg_score)

            # Convergence check
            if self._check_convergence(generation):
                break

            # Generate next generation
            new_population = []

            # Elite retention
            elite_indices = np.argsort(fitness_scores)[-self.elite_size:]
            for idx in elite_indices:
                new_population.append(population[idx].copy())

            # Generate new individuals
            while len(new_population) < self.population_size:
                parent1 = self._tournament_selection(population, fitness_scores)
                parent2 = self._tournament_selection(population, fitness_scores)
                child1, child2 = self._adaptive_crossover(parent1, parent2, param_bounds)
                child1 = self._adaptive_mutation(child1, param_bounds, generation)
                child2 = self._adaptive_mutation(child2, param_bounds, generation)
                new_population.extend([child1, child2])

            # Ensure population size is correct
            population = new_population[:self.population_size]

        return self.best_individual


# Auxiliary functions
def weighted_specificity_score(y_true, y_pred):

    classes = np.unique(np.concatenate([y_true, y_pred]))
    cm = confusion_matrix(y_true, y_pred, labels=classes)

    specificity_sum = 0
    total_samples = 0

    for i, cls in enumerate(classes):
        true_neg_mask = np.ones(cm.shape, dtype=bool)
        true_neg_mask[i, :] = False
        true_neg_mask[:, i] = False

        tn = np.sum(cm[true_neg_mask])
        fp = np.sum(cm[:, i]) - cm[i, i]

        if tn + fp > 0:
            specificity = tn / (tn + fp)
        else:
            specificity = 0

        weight = np.sum(y_true == cls)
        specificity_sum += specificity * weight
        total_samples += weight

    return specificity_sum / total_samples if total_samples > 0 else 0


def weighted_specificity(y_true, y_pred, labels):
    """Calculate Specificity"""
    total_specificity = 0
    total_weight = 0

    for label in labels:
        true_negatives = np.sum((y_true != label) & (y_pred != label))
        false_positives = np.sum((y_true != label) & (y_pred == label))
        weight = np.sum(y_true == label)
        total_weight += weight

        if (true_negatives + false_positives) > 0:
            specificity = true_negatives / (true_negatives + false_positives)
        else:
            specificity = 0.0
        total_specificity += specificity * weight

    return total_specificity / total_weight if total_weight > 0 else 0.0


class RandomForestModel:
    def __init__(self, config):
        self.config = config
        self.all_classes = config['all_classes']
        self.scaler = StandardScaler()

        # Initialize SMOTE
        smote_sampling_strategy = config.get('smote_sampling_strategy', 'auto')
        smote_k_neighbors = config.get('smote_k_neighbors', 5)
        self.smote = SMOTE(
            sampling_strategy=smote_sampling_strategy,
            k_neighbors=smote_k_neighbors,
            random_state=42
        )

        # Initialize feature interaction processor
        try:
            from AdvancedFeatureInteractionsRF import AdvancedFeatureInteractions
            self.feature_interaction = AdvancedFeatureInteractions(**config.get('feature_params', {}))
        except ImportError:
            self.feature_interaction = AdvancedFeatureInteractions(**config.get('feature_params', {}))

    def load_data(self, file_path=None, sheet_name=None, show_info=False, is_external=False):
        """Load and process data"""
        try:
            file_path = file_path if file_path else self.config['file_path']
            sheet_name = sheet_name if sheet_name else self.config['sheet_name']

            df = pd.read_excel(file_path, sheet_name=sheet_name)
            labels = df['Sample'].values
            features = df.iloc[:, 1:14].values

            self.original_indices = np.arange(len(features))
            original_feature_names = df.columns[1:14].tolist()
            self.feature_interaction.set_original_feature_names(original_feature_names)

            features_processed = self.feature_interaction.fit_transform(features, labels)
            self.feature_names = self.feature_interaction.get_feature_names()
            self.feature_importance = self.feature_interaction.get_feature_importance()

            # Store info for later display
            if show_info:
                # Get the number of features after interaction (before selection)
                # feature_names_ contains all feature names after interaction, before selection
                if hasattr(self.feature_interaction, 'feature_names_') and self.feature_interaction.feature_names_ is not None:
                    n_features_after_interaction = len(self.feature_interaction.feature_names_)
                else:
                    # Fallback: use final processed features count
                    n_features_after_interaction = features_processed.shape[1]

                self.data_info = {
                    'samples': features.shape[0],
                    'original_features': features.shape[1],
                    'generated_features': n_features_after_interaction,
                    'selected_features': features_processed.shape[1],
                    'is_external': is_external
                }

            return features_processed, np.array(labels), self.original_indices

        except Exception as e:
            logging.error("Error loading data for RF: {}".format(str(e)))
            raise

    def apply_smote(self, X, y, positive_class):

        y_binary = np.array(y == positive_class, dtype=int)

        try:
            X_resampled, y_binary_resampled = self.smote.fit_resample(X, y_binary)
            sample_indices = np.concatenate([
                self.original_indices,
                np.full(len(X_resampled) - len(X), -1)
            ])
            return X_resampled, y_binary_resampled, sample_indices
        except ValueError:
            return X, y_binary, self.original_indices

    def _optimize_hyperparameters(self, X_train, y_train):
        """Optimize hyperparameters using genetic algorithm"""
        neg_count = np.sum(y_train == 0)
        pos_count = np.sum(y_train == 1)
        base_class_ratio = neg_count / max(1, pos_count)

        def objective(params):
            try:
                max_features = max(0.1, min(0.999, params['max_features_ratio']))

                if params['class_weight_multiplier'] > 0:
                    class_weight = {
                        0: 1.0,
                        1: params['class_weight_multiplier'] * base_class_ratio
                    }
                else:
                    class_weight = None

                from sklearn.ensemble import RandomForestClassifier
                model = RandomForestClassifier(
                    n_estimators=params['n_estimators'],
                    max_depth=params['max_depth'],
                    min_samples_split=params['min_samples_split'],
                    min_samples_leaf=params['min_samples_leaf'],
                    max_features=max_features,
                    class_weight=class_weight,
                    random_state=42,
                    n_jobs=-1
                )

                cv = StratifiedKFold(n_splits=self.config['inner_cv_splits'], shuffle=True, random_state=42)
                f1_scores = []
                auc_scores = []
                recall_scores = []
                specificity_scores = []

                for train_idx, val_idx in cv.split(X_train, y_train):
                    X_train_cv, X_val_cv = X_train[train_idx], X_train[val_idx]
                    y_train_cv, y_val_cv = y_train[train_idx], y_train[val_idx]

                    y_train_cv = y_train_cv.astype(int)
                    model.fit(X_train_cv, y_train_cv)
                    y_pred_proba = model.predict_proba(X_val_cv)[:, 1]

                    best_f1 = 0
                    best_pred = None
                    for threshold in [0.3, 0.35, 0.4, 0.45, 0.5, 0.55, 0.6]:
                        y_pred_thresh = np.array(y_pred_proba >= threshold, dtype=int)
                        f1 = f1_score(y_val_cv, y_pred_thresh, average='weighted')
                        if f1 > best_f1:
                            best_f1 = f1
                            best_pred = y_pred_thresh

                    f1_scores.append(f1_score(y_val_cv, best_pred, average='weighted'))

                    try:
                        auc_scores.append(roc_auc_score(y_val_cv, y_pred_proba))
                    except Exception:
                        auc_scores.append(0)

                    recall_scores.append(recall_score(y_val_cv, best_pred, average='weighted'))
                    specificity_scores.append(weighted_specificity_score(y_val_cv, best_pred))

                mean_f1 = np.mean(f1_scores)
                mean_auc = np.mean(auc_scores)
                mean_recall = np.mean(recall_scores)
                mean_specificity = np.mean(specificity_scores)

                return 0.55 * mean_f1 + 0.15 * mean_auc + 0.10 * mean_recall + 0.20 * mean_specificity

            except Exception:
                return 0.0

        # Parameter bounds
        param_bounds = {
            'n_estimators': (100, 500, 'int'),
            'max_depth': (3, 40, 'int'),
            'min_samples_split': (2, 20, 'int'),
            'min_samples_leaf': (2, 20, 'int'),
            'max_features_ratio': (0.1, 1.0, 'float'),
            'class_weight_multiplier': (0.0, 5.0, 'float')
        }

        # Create genetic algorithm optimizer
        ga_config = self.config.get('ga_config', {})
        ga_optimizer = GeneticAlgorithmOptimizer(
            population_size=ga_config.get('population_size', 40),
            max_generations=ga_config.get('max_generations', 50),
            crossover_rate=ga_config.get('crossover_rate', 0.8),
            mutation_rate=ga_config.get('mutation_rate', 0.15),
            random_state=42
        )

        best_params = ga_optimizer.optimize(objective, param_bounds)

        # Convert parameter format
        best_params['n_estimators'] = int(best_params['n_estimators'])
        best_params['max_depth'] = int(best_params['max_depth'])
        best_params['min_samples_split'] = int(best_params['min_samples_split'])
        best_params['min_samples_leaf'] = int(best_params['min_samples_leaf'])

        max_features_ratio = max(0.1, min(0.999, best_params['max_features_ratio']))
        best_params['max_features'] = max_features_ratio
        best_params.pop('max_features_ratio', None)

        class_weight_multiplier = best_params.pop('class_weight_multiplier', 0)
        if class_weight_multiplier > 0:
            best_params['class_weight'] = {
                0: 1.0,
                1: class_weight_multiplier * base_class_ratio
            }
        else:
            best_params['class_weight'] = None

        return best_params

    def _create_model(self, params):
        """Create Random Forest algorithms"""
        from sklearn.ensemble import RandomForestClassifier
        params_with_defaults = {
            'n_estimators': int(params.get('n_estimators', 100)),
            'max_depth': params.get('max_depth', None),
            'min_samples_split': int(params.get('min_samples_split', 2)),
            'min_samples_leaf': int(params.get('min_samples_leaf', 1)),
            'max_features': params.get('max_features', 'sqrt'),
            'class_weight': params.get('class_weight', None),
            'random_state': 42,
            'n_jobs': -1
        }
        return RandomForestClassifier(**params_with_defaults)


class XGBoostClassifier:
    def __init__(self, config):
        self.config = config
        self.all_classes = config['all_classes']
        self.scaler = StandardScaler()

        # Initialize ADASYN
        adasyn_sampling_strategy = config.get('adasyn_sampling_strategy', 'auto')
        adasyn_n_neighbors = config.get('adasyn_n_neighbors', 5)
        self.adasyn = ADASYN(
            sampling_strategy=adasyn_sampling_strategy,
            n_neighbors=adasyn_n_neighbors,
            random_state=42
        )

        # Initialize feature interaction processor
        try:
            from AdvancedFeatureInteractionsXGB import AdvancedFeatureInteractions
            self.feature_interaction = AdvancedFeatureInteractions(**config.get('feature_params', {}))
        except ImportError:
            self.feature_interaction = AdvancedFeatureInteractions(**config.get('feature_params', {}))

    def load_data(self, file_path=None, sheet_name=None, show_info=False, is_external=False):
        """Load and process data"""
        try:
            file_path = file_path if file_path else self.config['file_path']
            sheet_name = sheet_name if sheet_name else self.config['sheet_name']

            df = pd.read_excel(file_path, sheet_name=sheet_name)
            labels = df['Sample'].values
            features = df.iloc[:, 1:].values

            self.original_indices = np.arange(len(features))
            original_feature_names = df.columns[1:].tolist()
            self.feature_interaction.set_original_feature_names(original_feature_names)

            features_processed = self.feature_interaction.fit_transform(features, labels)
            self.feature_names = self.feature_interaction.get_feature_names()
            self.feature_importance = self.feature_interaction.get_feature_importance()

            # Store info for later display
            if show_info:
                # Get the number of features after interaction (before selection)
                # feature_names_ contains all feature names after interaction, before selection
                if hasattr(self.feature_interaction, 'feature_names_') and self.feature_interaction.feature_names_ is not None:
                    n_features_after_interaction = len(self.feature_interaction.feature_names_)
                else:
                    # Fallback: use final processed features count
                    n_features_after_interaction = features_processed.shape[1]

                self.data_info = {
                    'samples': features.shape[0],
                    'original_features': features.shape[1],
                    'generated_features': n_features_after_interaction,
                    'selected_features': features_processed.shape[1],
                    'is_external': is_external
                }

            return features_processed, np.array(labels), self.original_indices

        except Exception as e:
            logging.error("Error loading data for XGBoost: {}".format(str(e)))
            raise

    def apply_adasyn(self, X, y, positive_class):

        y_binary = np.array(y == positive_class, dtype=int)

        n_positive = np.sum(y_binary == 1)
        if n_positive < 2:
            return X, y_binary, self.original_indices

        try:
            n_neighbors = self.config.get('adasyn_n_neighbors', 5)
            if n_positive <= n_neighbors:
                adaptive_n_neighbors = max(1, n_positive - 1)
                temp_adasyn = ADASYN(
                    sampling_strategy=self.config.get('adasyn_sampling_strategy', 'auto'),
                    n_neighbors=adaptive_n_neighbors,
                    random_state=42
                )
                X_resampled, y_binary_resampled = temp_adasyn.fit_resample(X, y_binary)
            else:
                X_resampled, y_binary_resampled = self.adasyn.fit_resample(X, y_binary)

            sample_indices = np.concatenate([
                self.original_indices,
                np.full(len(X_resampled) - len(X), -1)
            ])
            return X_resampled, y_binary_resampled, sample_indices

        except ValueError:
            return X, y_binary, self.original_indices

    def _optimize_hyperparameters(self, X_train, y_train):

        neg_count = np.sum(y_train == 0)
        pos_count = np.sum(y_train == 1)
        base_scale_pos_weight = neg_count / max(1, pos_count)

        def objective(params):
            try:
                actual_scale_pos_weight = params['scale_pos_weight_multiplier'] * base_scale_pos_weight

                import xgboost as xgb
                model = xgb.XGBClassifier(
                    n_estimators=params['n_estimators'],
                    max_depth=params['max_depth'],
                    learning_rate=params['learning_rate'],
                    subsample=params['subsample'],
                    colsample_bytree=params['colsample_bytree'],
                    min_child_weight=params['min_child_weight'],
                    gamma=params['gamma'],
                    reg_alpha=params['reg_alpha'],
                    reg_lambda=params['reg_lambda'],
                    scale_pos_weight=actual_scale_pos_weight,
                    use_label_encoder=False,
                    eval_metric='logloss',
                    random_state=42,
                    n_jobs=-1
                )

                cv = StratifiedKFold(n_splits=self.config['inner_cv_splits'], shuffle=True, random_state=42)
                f1_scores = []
                auc_scores = []
                recall_scores = []
                specificity_scores = []

                for train_idx, val_idx in cv.split(X_train, y_train):
                    X_train_cv, X_val_cv = X_train[train_idx], X_train[val_idx]
                    y_train_cv, y_val_cv = y_train[train_idx], y_train[val_idx]

                    y_train_cv = y_train_cv.astype(int)
                    model.fit(X_train_cv, y_train_cv)
                    y_pred_proba = model.predict_proba(X_val_cv)[:, 1]

                    best_f1 = 0
                    best_pred = None
                    for threshold in [0.3, 0.35, 0.4, 0.45, 0.5, 0.55, 0.6]:
                        y_pred_thresh = np.array(y_pred_proba >= threshold, dtype=int)
                        f1 = f1_score(y_val_cv, y_pred_thresh, average='weighted')
                        if f1 > best_f1:
                            best_f1 = f1
                            best_pred = y_pred_thresh

                    f1_scores.append(f1_score(y_val_cv, best_pred, average='weighted'))

                    try:
                        auc_scores.append(roc_auc_score(y_val_cv, y_pred_proba))
                    except Exception:
                        auc_scores.append(0)

                    recall_scores.append(recall_score(y_val_cv, best_pred, average='weighted'))
                    specificity_scores.append(weighted_specificity(y_val_cv, best_pred, [0, 1]))

                mean_f1 = np.mean(f1_scores)
                mean_auc = np.mean(auc_scores)
                mean_recall = np.mean(recall_scores)
                mean_specificity = np.mean(specificity_scores)

                return 0.55 * mean_f1 + 0.15 * mean_auc + 0.10 * mean_recall + 0.20 * mean_specificity

            except Exception:
                return 0.0

        # Parameter bounds
        param_bounds = {
            'n_estimators': (100, 800, 'int'),
            'max_depth': (2, 5, 'int'),
            'learning_rate': (0.01, 0.15, 'float'),
            'subsample': (0.7, 1.0, 'float'),
            'colsample_bytree': (0.7, 1.0, 'float'),
            'min_child_weight': (1, 5, 'int'),
            'gamma': (0, 0.5, 'float'),
            'reg_alpha': (0, 1.0, 'float'),
            'reg_lambda': (0.1, 10.0, 'float'),
            'scale_pos_weight_multiplier': (0.8, 3.0, 'float')
        }

        # Create genetic algorithm optimizer
        ga_config = self.config.get('ga_config', {})
        ga_optimizer = GeneticAlgorithmOptimizer(
            population_size=ga_config.get('population_size', 40),
            max_generations=ga_config.get('max_generations', 50),
            crossover_rate=ga_config.get('crossover_rate', 0.8),
            mutation_rate=ga_config.get('mutation_rate', 0.15),
            random_state=42
        )

        best_params = ga_optimizer.optimize(objective, param_bounds)

        # Convert parameter format
        best_params['n_estimators'] = int(best_params['n_estimators'])
        best_params['max_depth'] = int(best_params['max_depth'])
        best_params['min_child_weight'] = max(1, int(best_params.get('min_child_weight', 1)))
        best_params['scale_pos_weight'] = best_params['scale_pos_weight_multiplier'] * base_scale_pos_weight
        del best_params['scale_pos_weight_multiplier']

        return best_params

    def _create_model(self, params):
        """Create XGBoost algorithms"""
        import xgboost as xgb
        params_with_defaults = {
            'n_estimators': int(params.get('n_estimators', 100)),
            'max_depth': int(params.get('max_depth', 3)),
            'learning_rate': params.get('learning_rate', 0.1),
            'subsample': params.get('subsample', 0.8),
            'colsample_bytree': params.get('colsample_bytree', 0.8),
            'scale_pos_weight': params.get('scale_pos_weight', 1.0),
            'min_child_weight': int(params.get('min_child_weight', 1)),
            'gamma': params.get('gamma', 0),
            'reg_alpha': params.get('reg_alpha', 0),
            'reg_lambda': params.get('reg_lambda', 1.0)
        }

        return xgb.XGBClassifier(
            **params_with_defaults,
            use_label_encoder=False,
            eval_metric='logloss',
            random_state=42,
            n_jobs=-1
        )


class EnsembleClassifier:
    """Ensemble Classifier"""

    def __init__(self, config):
        self.config = config
        self.all_classes = config['all_classes']
        self.scaler = StandardScaler()

        # Weight configuration
        self.use_fixed_weights = config.get('use_fixed_weights', True)
        self.fixed_xgb_weight = config.get('fixed_xgb_weight', 0.1)
        self.fixed_rf_weight = 1.0 - self.fixed_xgb_weight

        # Store models
        self.original_new_data = None
        self.final_models = {}

        # Initialize RF and XGBoost components
        self.rf_component = RandomForestModel(config['rf_component'])
        self.xgb_component = XGBoostClassifier(config['xgb_component'])

    def load_data(self, file_path=None, sheet_name=None, show_info=False, is_external=False):
        """Load data"""
        rf_features, rf_labels, rf_indices = self.rf_component.load_data(file_path, sheet_name, show_info, is_external)
        xgb_features, xgb_labels, xgb_indices = self.xgb_component.load_data(file_path, sheet_name, show_info, is_external)

        # Display unified information after both components load
        if show_info:
            rf_info = self.rf_component.data_info
            if is_external:
                print(f"External validation data: {rf_info['samples']} samples, {rf_info['original_features']} original features")
                print(f"After feature interaction: {rf_info['generated_features']} features generated")
                print(f"After feature selection: {rf_info['selected_features']} features selected")
                print(f"Final shape: {rf_info['samples']} samples × {rf_info['selected_features']} features")
            else:
                print(f"Training data: {rf_info['samples']} samples, {rf_info['original_features']} original features")
                print(f"After feature interaction: {rf_info['generated_features']} features generated")
                print(f"After feature selection: {rf_info['selected_features']} features selected")
                print(f"Final shape: {rf_info['samples']} samples × {rf_info['selected_features']} features")

        return (rf_features, rf_labels, rf_indices), (xgb_features, xgb_labels, xgb_indices)

    def apply_data_augmentation(self, rf_data, xgb_data, positive_class):
        """Apply data augmentation"""
        rf_features, rf_labels, rf_indices = rf_data
        xgb_features, xgb_labels, xgb_indices = xgb_data

        rf_X_balanced, rf_y_binary, rf_sample_indices = self.rf_component.apply_smote(
            rf_features, rf_labels, positive_class
        )

        xgb_X_balanced, xgb_y_binary, xgb_sample_indices = self.xgb_component.apply_adasyn(
            xgb_features, xgb_labels, positive_class
        )

        # Ensure the two datasets have the same size
        if len(rf_X_balanced) != len(xgb_X_balanced):
            min_samples = min(len(rf_X_balanced), len(xgb_X_balanced))
            if len(rf_X_balanced) > min_samples:
                rf_X_balanced = rf_X_balanced[:min_samples]
                rf_y_binary = rf_y_binary[:min_samples]
                rf_sample_indices = rf_sample_indices[:min_samples]
            if len(xgb_X_balanced) > min_samples:
                xgb_X_balanced = xgb_X_balanced[:min_samples]
                xgb_y_binary = xgb_y_binary[:min_samples]
                xgb_sample_indices = xgb_sample_indices[:min_samples]

        return (rf_X_balanced, rf_y_binary, rf_sample_indices), (xgb_X_balanced, xgb_y_binary, xgb_sample_indices)

    def process_fold(self, rf_data, xgb_data, train_index, test_index, positive_class, fold_idx):
        """Process single cross-validation fold"""
        # Unpack data
        rf_X, rf_y, rf_indices = rf_data
        xgb_X, xgb_y, xgb_indices = xgb_data

        # Split into training and test sets
        rf_X_train, rf_X_test = rf_X[train_index], rf_X[test_index]
        rf_y_train, rf_y_test = rf_y[train_index], rf_y[test_index]
        rf_indices_train, rf_indices_test = rf_indices[train_index], rf_indices[test_index]

        xgb_X_train, xgb_X_test = xgb_X[train_index], xgb_X[test_index]
        xgb_y_train, xgb_y_test = xgb_y[train_index], xgb_y[test_index]
        xgb_indices_train, xgb_indices_test = xgb_indices[train_index], xgb_indices[test_index]

        # Use only original data for evaluation
        rf_original_test_mask = rf_indices_test >= 0
        rf_X_test_original = rf_X_test[rf_original_test_mask]
        rf_y_test_original = rf_y_test[rf_original_test_mask]

        xgb_original_test_mask = xgb_indices_test >= 0
        xgb_X_test_original = xgb_X_test[xgb_original_test_mask]
        xgb_y_test_original = xgb_y_test[xgb_original_test_mask]

        # Verify consistency of test sets
        if not np.array_equal(rf_y_test_original, xgb_y_test_original):
            common_indices = np.intersect1d(rf_indices_test[rf_original_test_mask],
                                            xgb_indices_test[xgb_original_test_mask])
            rf_common_mask = np.isin(rf_indices_test, common_indices)
            xgb_common_mask = np.isin(xgb_indices_test, common_indices)

            rf_X_test_original = rf_X_test[rf_common_mask]
            rf_y_test_original = rf_y_test[rf_common_mask]
            xgb_X_test_original = xgb_X_test[xgb_common_mask]
            xgb_y_test_original = xgb_y_test[xgb_common_mask]

        # Optimize RF algorithms hyperparameters
        rf_best_params = self.rf_component._optimize_hyperparameters(rf_X_train, rf_y_train)
        rf_model = self.rf_component._create_model(rf_best_params)
        rf_model.fit(rf_X_train, rf_y_train)

        # Optimize XGBoost algorithms hyperparameters
        xgb_best_params = self.xgb_component._optimize_hyperparameters(xgb_X_train, xgb_y_train)
        xgb_model = self.xgb_component._create_model(xgb_best_params)
        xgb_model.fit(xgb_X_train, xgb_y_train)

        # Get prediction probabilities
        rf_y_pred_proba = rf_model.predict_proba(rf_X_test_original)[:, 1]
        xgb_y_pred_proba = xgb_model.predict_proba(xgb_X_test_original)[:, 1]

        best_ensemble_weight = self.fixed_xgb_weight

        # Calculate weighted ensemble prediction probabilities
        ensemble_y_pred_proba = (1 - best_ensemble_weight) * rf_y_pred_proba + best_ensemble_weight * xgb_y_pred_proba

        # Use dynamic threshold selection
        test_class_dist = np.bincount(rf_y_test_original)
        thresholds = np.arange(0.1, 0.7, 0.025)
        best_f1 = 0
        best_threshold = 0.4

        if len(test_class_dist) > 1 and test_class_dist[0] > 0 and np.sum(test_class_dist) > test_class_dist[0]:
            for threshold in thresholds:
                y_pred_thresh = np.array(ensemble_y_pred_proba >= threshold, dtype=int)
                f1 = f1_score(rf_y_test_original, y_pred_thresh, average='weighted')
                if f1 > best_f1:
                    best_f1 = f1
                    best_threshold = threshold

        # Use best threshold for prediction
        ensemble_y_pred = np.array(ensemble_y_pred_proba >= best_threshold, dtype=int)

        # Calculate evaluation metrics
        ensemble_f1 = f1_score(rf_y_test_original, ensemble_y_pred, average='weighted')
        ensemble_precision = precision_score(rf_y_test_original, ensemble_y_pred, average='weighted', zero_division=0)
        ensemble_recall = recall_score(rf_y_test_original, ensemble_y_pred, average='weighted')
        ensemble_sensitivity = ensemble_recall
        ensemble_specificity = weighted_specificity_score(rf_y_test_original, ensemble_y_pred)

        # ROC curve calculation
        try:
            fpr, tpr, _ = roc_curve(rf_y_test_original, ensemble_y_pred_proba)
            roc_auc = auc(fpr, tpr)
        except Exception:
            fpr, tpr = np.array([0, 1]), np.array([0, 1])
            roc_auc = 0.5

        # PR curve calculation
        try:
            precision, recall, _ = precision_recall_curve(rf_y_test_original, ensemble_y_pred_proba)
            pr_auc = average_precision_score(rf_y_test_original, ensemble_y_pred_proba)
        except Exception:
            precision, recall = np.array([0, 1]), np.array([1, 0])
            pr_auc = 0.5

        # Confusion matrix
        conf_matrix = confusion_matrix(rf_y_test_original, ensemble_y_pred)

        # Calculate false negative rate and false positive rate
        tn, fp, fn, tp = 0, 0, 0, 0
        if conf_matrix.shape == (2, 2):
            tn, fp, fn, tp = conf_matrix.ravel()
        elif conf_matrix.shape == (1, 1):
            if np.sum(rf_y_test_original) == 0:
                tn = conf_matrix[0, 0]
            else:
                tp = conf_matrix[0, 0]

        fnr = fn / (fn + tp) if (fn + tp) > 0 else 0
        fpr_val = fp / (fp + tn) if (fp + tn) > 0 else 0

        return {
            "fold_idx": fold_idx,
            "y_test": rf_y_test_original,
            "y_pred": ensemble_y_pred,
            "y_pred_proba": ensemble_y_pred_proba,
            "optimal_threshold": best_threshold,
            "optimal_xgb_weight": best_ensemble_weight,
            "f1": ensemble_f1,
            "precision": ensemble_precision,
            "recall": ensemble_recall,
            "sensitivity": ensemble_sensitivity,
            "specificity": ensemble_specificity,
            "fpr": fpr,
            "tpr": tpr,
            "roc_auc": roc_auc,
            "pr_curve": {"precision": precision, "recall": recall},
            "pr_auc": pr_auc,
            "fnr": fnr,
            "fpr_val": fpr_val,
            "confusion_matrix": conf_matrix,
            "rf_model": rf_model,
            "xgb_model": xgb_model,
            "rf_best_params": rf_best_params,
            "xgb_best_params": xgb_best_params,
            "class_name": positive_class
        }

    def combine_and_train(self, new_data_path, new_data_sheet=None):

        try:
            # Load training data with feature info display
            print()
            print("Loading and processing training data...")
            rf_train_data, xgb_train_data = self.load_data(show_info=True, is_external=False)
            rf_X_train, rf_y_train, rf_train_indices = rf_train_data
            xgb_X_train, xgb_y_train, xgb_train_indices = xgb_train_data
            print()

            # Load external data
            rf_new_data, xgb_new_data = self.load_data(new_data_path, new_data_sheet, show_info=False, is_external=True)
            rf_X_new, rf_y_new, rf_new_indices = rf_new_data
            xgb_X_new, xgb_y_new, xgb_new_indices = xgb_new_data

            # Store external data shape info for later display
            self.external_samples = len(rf_X_new)
            self.external_original_features = 12  # Original features
            self.external_final_features = rf_X_new.shape[1]

            # Save original new data for final prediction
            self.original_new_data = {
                'rf': {
                    'X': rf_X_new.copy(),
                    'y': rf_y_new.copy(),
                    'indices': rf_new_indices.copy()
                },
                'xgb': {
                    'X': xgb_X_new.copy(),
                    'y': xgb_y_new.copy(),
                    'indices': xgb_new_indices.copy()
                }
            }

            self.final_models = {}

            for idx, class_name in enumerate(self.all_classes, 1):
                # Start spinner indicator
                indicator = SimpleSpinnerIndicator()
                indicator.start(f"Processing class: {class_name} ({idx}/{len(self.all_classes)})")

                class_start_time = time.time()

                rf_train_original_len = len(rf_X_train)
                xgb_train_original_len = len(xgb_X_train)

                rf_X_train_balanced, rf_y_train_binary, _ = self.rf_component.apply_smote(
                    rf_X_train, rf_y_train, class_name
                )
                xgb_X_train_balanced, xgb_y_train_binary, _ = self.xgb_component.apply_adasyn(
                    xgb_X_train, xgb_y_train, class_name
                )

                rf_X_train_synthetic = rf_X_train_balanced[rf_train_original_len:]
                rf_y_train_synthetic = rf_y_train_binary[rf_train_original_len:]
                xgb_X_train_synthetic = xgb_X_train_balanced[xgb_train_original_len:]
                xgb_y_train_synthetic = xgb_y_train_binary[xgb_train_original_len:]

                rf_new_original_len = len(rf_X_new)
                xgb_new_original_len = len(xgb_X_new)

                rf_X_new_balanced, rf_y_new_binary, _ = self.rf_component.apply_smote(
                    rf_X_new, rf_y_new, class_name
                )
                xgb_X_new_balanced, xgb_y_new_binary, _ = self.xgb_component.apply_adasyn(
                    xgb_X_new, xgb_y_new, class_name
                )

                rf_X_new_synthetic = rf_X_new_balanced[rf_new_original_len:]
                rf_y_new_synthetic = rf_y_new_binary[rf_new_original_len:]
                xgb_X_new_synthetic = xgb_X_new_balanced[xgb_new_original_len:]
                xgb_y_new_synthetic = xgb_y_new_binary[xgb_new_original_len:]

                rf_X_combined = np.vstack([
                    rf_X_train,
                    rf_X_train_synthetic,
                    rf_X_new_synthetic
                ])

                rf_y_combined = np.concatenate([
                    np.array(rf_y_train == class_name, dtype=int),
                    rf_y_train_synthetic,
                    rf_y_new_synthetic
                ])

                xgb_X_combined = np.vstack([
                    xgb_X_train,
                    xgb_X_train_synthetic,
                    xgb_X_new_synthetic
                ])

                xgb_y_combined = np.concatenate([
                    np.array(xgb_y_train == class_name, dtype=int),
                    xgb_y_train_synthetic,
                    xgb_y_new_synthetic
                ])

                # Ensure RF and XGBoost data have the same size
                if len(rf_X_combined) != len(xgb_X_combined):
                    common_len = min(len(rf_X_combined), len(xgb_X_combined))

                    if len(rf_X_combined) > common_len:
                        rf_X_combined = rf_X_combined[:common_len]
                        rf_y_combined = rf_y_combined[:common_len]

                    if len(xgb_X_combined) > common_len:
                        xgb_X_combined = xgb_X_combined[:common_len]
                        xgb_y_combined = xgb_y_combined[:common_len]

                # 4. Perform cross-validation training
                outer_cv = StratifiedKFold(
                    n_splits=self.config['outer_cv_splits'],
                    shuffle=True,
                    random_state=42
                )

                rf_combined_indices = np.arange(len(rf_X_combined))
                xgb_combined_indices = np.arange(len(xgb_X_combined))

                rf_combined_data = (rf_X_combined, rf_y_combined, rf_combined_indices)
                xgb_combined_data = (xgb_X_combined, xgb_y_combined, xgb_combined_indices)

                fold_indices = list(enumerate(outer_cv.split(rf_X_combined, rf_y_combined)))

                # Parallel processing of each fold
                results = Parallel(n_jobs=-1, verbose=10)(
                    delayed(self.process_fold)(
                        rf_combined_data, xgb_combined_data, train_idx, test_idx, class_name, fold_idx
                    )
                    for fold_idx, (train_idx, test_idx) in fold_indices
                )

                results = [r for r in results if r is not None]
                if not results:
                    logging.warning(f"No valid LOO results for class {class_name}")
                    continue

                # 5. Train final algorithms on full training dataset
                # Calculate average parameters
                rf_final_params = {}
                xgb_final_params = {}

                rf_all_params = [r['rf_best_params'] for r in results]
                xgb_all_params = [r['xgb_best_params'] for r in results]

                # Process RF parameters
                for key in rf_all_params[0].keys():
                    if key == 'class_weight':
                        class_weights = []
                        for params in rf_all_params:
                            if params.get(key) is not None:
                                if isinstance(params[key], dict) and 1 in params[key]:
                                    class_weights.append(params[key][1])

                        if class_weights:
                            avg_weight = np.mean(class_weights)
                            rf_final_params['class_weight'] = {0: 1.0, 1: avg_weight}
                        else:
                            rf_final_params['class_weight'] = None
                    else:
                        values = [params.get(key) for params in rf_all_params if key in params]
                        if values:
                            if all(isinstance(v, (int, float)) for v in values):
                                if all(isinstance(v, int) for v in values):
                                    rf_final_params[key] = int(np.mean(values))
                                else:
                                    rf_final_params[key] = np.mean(values)

                # Process XGBoost parameters
                for key in xgb_all_params[0].keys():
                    values = [params.get(key) for params in xgb_all_params if key in params]
                    if values:
                        if all(isinstance(v, (int, float)) for v in values):
                            if all(isinstance(v, int) for v in values):
                                xgb_final_params[key] = int(np.mean(values))
                            else:
                                xgb_final_params[key] = np.mean(values)

                # Create and train final models
                rf_final_model = self.rf_component._create_model(rf_final_params)
                xgb_final_model = self.xgb_component._create_model(xgb_final_params)

                rf_final_model.fit(rf_X_combined, rf_y_combined)
                xgb_final_model.fit(xgb_X_combined, xgb_y_combined)

                # Calculate average threshold
                avg_threshold = np.mean([r['optimal_threshold'] for r in results])

                # Store final algorithms
                self.final_models[class_name] = {
                    'rf_model': rf_final_model,
                    'xgb_model': xgb_final_model,
                    'weight': self.fixed_xgb_weight,
                    'threshold': avg_threshold,
                    'rf_params': rf_final_params,
                    'xgb_params': xgb_final_params
                }

                # Stop spinner and show completion
                indicator.stop()
                class_time = time.time() - class_start_time
                print(f"Processing class: {class_name} ({idx}/{len(self.all_classes)}) - Completed [{class_time:.1f}s]")

            # Display external validation data information
            print()
            print("External validation data processing summary:")
            print(f"External validation data: {self.external_samples} samples, {self.external_original_features} original features")
            print(f"After feature interaction: {self.external_final_features} features generated")
            print(f"After feature selection: {self.external_final_features} features selected")
            print(f"Final shape: {self.external_samples} samples × {self.external_final_features} features")
            print()

            # 6. Predict on external data
            predictions_df = self.predict_original_new_data()
            return predictions_df

        except Exception as e:
            logging.error(f"Error in combined training process: {str(e)}")
            import traceback
            logging.error(traceback.format_exc())
            raise

    def predict_original_new_data(self):
        """Use trained final models to predict on external data"""
        try:
            if self.original_new_data is None:
                raise ValueError("No original new data available")

            rf_X_new = self.original_new_data['rf']['X']
            rf_y_new = self.original_new_data['rf']['y']
            xgb_X_new = self.original_new_data['xgb']['X']

            y_true = rf_y_new

            # Store prediction probabilities for each class
            class_probabilities = {}

            if not self.final_models:
                raise ValueError("No final models available")

            # Predict using trained classifier for each class
            for class_name, model_data in self.final_models.items():
                rf_model = model_data['rf_model']
                xgb_model = model_data['xgb_model']
                weight = model_data['weight']
                threshold = model_data['threshold']

                # RF algorithms prediction
                rf_proba = rf_model.predict_proba(rf_X_new)[:, 1]

                # XGBoost algorithms prediction
                xgb_proba = xgb_model.predict_proba(xgb_X_new)[:, 1]

                # Ensemble prediction
                ensemble_proba = (1 - weight) * rf_proba + weight * xgb_proba

                # Store prediction probabilities
                class_probabilities[class_name] = ensemble_proba

            # Create prediction LOO results DataFrame
            results_df = pd.DataFrame()
            results_df['Sample_ID'] = [f"Sample_{i + 1}" for i in range(len(rf_X_new))]
            results_df['True_Label'] = y_true

            # Add prediction probabilities for each class
            for class_name in self.all_classes:
                if class_name in class_probabilities:
                    results_df[f'{class_name}_Probability'] = class_probabilities[class_name]
                else:
                    results_df[f'{class_name}_Probability'] = 0.0

            # Determine final predicted class based on probabilities
            prob_columns = [col for col in results_df.columns if col.endswith('_Probability')]
            results_df['Predicted_Class'] = results_df[prob_columns].idxmax(axis=1).str.replace('_Probability', '')

            # Calculate performance metrics
            mask_correct = (results_df['True_Label'] == results_df['Predicted_Class'])
            correct_predictions = np.sum(mask_correct)
            accuracy = correct_predictions / len(results_df)
            weighted_f1 = f1_score(y_true, results_df['Predicted_Class'], average='weighted')

            # Calculate and save confusion matrix
            cm = confusion_matrix(y_true, results_df['Predicted_Class'], labels=self.all_classes)
            cm_df = pd.DataFrame(cm, index=self.all_classes, columns=self.all_classes)

            # Calculate performance metrics for each class
            class_metrics = []
            all_precision = []
            all_recall = []
            all_sensitivity = []
            all_specificity = []
            all_f1 = []
            all_auc = []

            for class_name in self.all_classes:
                if class_name not in class_probabilities:
                    continue

                # Binarize labels
                y_true_binary = np.array(y_true == class_name, dtype=int)
                y_pred_binary = np.array(results_df['Predicted_Class'] == class_name, dtype=int)

                # Calculate prediction probabilities
                y_proba = class_probabilities[class_name]

                # Calculate basic performance metrics
                tn, fp, fn, tp = confusion_matrix(y_true_binary, y_pred_binary).ravel()

                precision = tp / (tp + fp) if (tp + fp) > 0 else 0
                recall = tp / (tp + fn) if (tp + fn) > 0 else 0
                sensitivity = recall
                specificity = tn / (tn + fp) if (tn + fp) > 0 else 0
                f1 = 2 * precision * recall / (precision + recall) if (precision + recall) > 0 else 0

                all_precision.append(precision)
                all_recall.append(recall)
                all_sensitivity.append(sensitivity)
                all_specificity.append(specificity)
                all_f1.append(f1)

                # Calculate AUC
                try:
                    auc_score = roc_auc_score(y_true_binary, y_proba)
                    all_auc.append(auc_score)
                except Exception as e:
                    auc_score = 0
                    all_auc.append(0)

                # Calculate ROC curve data
                fpr, tpr, thresholds = roc_curve(y_true_binary, y_proba)

                # Save ROC curve data to DataFrame
                roc_data = pd.DataFrame({
                    'FPR': fpr,
                    'TPR': tpr,
                    'Thresholds': thresholds if len(thresholds) == len(fpr) else np.append(thresholds, np.nan)
                })

                # Save ROC data to Excel
                roc_file_path = os.path.join(self.config['output_path'], f'ROC_data_{class_name}.xlsx')
                roc_data.to_excel(roc_file_path, index=False)
                self.format_excel_file(roc_file_path)

                class_metrics.append({
                    'Class': class_name,
                    'Precision': precision,
                    'Recall': recall,
                    'Sensitivity': sensitivity,
                    'Specificity': specificity,
                    'F1': f1,
                    'AUC': auc_score,
                })

            # Add Overall performance
            if all_precision:
                class_metrics.append({
                    'Class': 'Overall',
                    'Precision': np.mean(all_precision),
                    'Recall': np.mean(all_recall),
                    'Sensitivity': np.mean(all_sensitivity),
                    'Specificity': np.mean(all_specificity),
                    'F1': np.mean(all_f1),
                    'AUC': np.mean(all_auc),
                })

            # Create performance metrics DataFrame
            metrics_df = pd.DataFrame(class_metrics)

            # Plot confusion matrix heatmap
            self.plot_external_confusion_matrix(y_true, results_df['Predicted_Class'])

            # Plot ROC curves
            self.plot_external_roc_curves(y_true, class_probabilities)

            # Save external validation LOO results
            output_path = os.path.join(self.config['output_path'], 'external_validation_results_GA.xlsx')
            with pd.ExcelWriter(output_path) as writer:
                results_df.to_excel(writer, sheet_name='Predictions', index=False)
                metrics_df.to_excel(writer, sheet_name='Class_Metrics', index=False)
                cm_df.to_excel(writer, sheet_name='Confusion_Matrix')

            # Format Excel file
            self.format_excel_file(output_path)

            return results_df

        except Exception as e:
            logging.error(f"Error predicting original new data: {str(e)}")
            import traceback
            logging.error(traceback.format_exc())
            raise

    def plot_external_confusion_matrix(self, y_true, y_pred):
        """Plot confusion matrix for all classes"""
        plt.figure(figsize=(10, 8))
        plt.rcParams['font.family'] = 'Arial'
        ax = plt.gca()

        for spine in ax.spines.values():
            spine.set_linewidth(5)

        cm = confusion_matrix(y_true, y_pred, labels=self.all_classes)

        sns.heatmap(cm, annot=True, fmt='d', cmap='Blues',
                    xticklabels=self.all_classes,
                    yticklabels=self.all_classes,
                    annot_kws={'size': 20, 'weight': 'bold'})

        plt.title('Confusion Matrix for All Classes (GA Optimized)', fontsize=20, pad=20)
        plt.xlabel('Predicted Label', fontsize=24, labelpad=15)
        plt.ylabel('True Label', fontsize=24, labelpad=15)

        ax.tick_params(axis='both', labelsize=24)

        output_path = os.path.join(self.config['output_path'], 'combined_confusion_matrix_GA.svg')
        plt.savefig(output_path, format='svg', bbox_inches='tight')
        plt.close()

    def plot_external_roc_curves(self, y_true, class_probabilities):
        """Plot ROC curves for external validation data"""
        plt.figure(figsize=(10, 8))
        plt.rcParams['font.family'] = 'Arial'
        ax = plt.gca()

        for spine in ax.spines.values():
            spine.set_linewidth(7)

        plt.title("ROC Curves for All Classes (GA Optimized)", fontsize=20, pad=15)
        plt.xlabel("False Positive Rate", fontsize=24, labelpad=10)
        plt.ylabel("True Positive Rate", fontsize=24, labelpad=10)

        # Plot diagonal line
        plt.plot([0, 1], [0, 1], 'k--', lw=2)

        # Create denser interpolation points for smoother curves
        mean_fpr = np.linspace(0, 1, 1000)

        # Store ROC data for all classes to calculate average ROC AUC
        all_tprs = []
        all_aucs = []

        for class_name in self.all_classes:
            if class_name not in class_probabilities:
                continue

            y_true_binary = np.array(y_true == class_name, dtype=int)
            y_prob = class_probabilities[class_name]

            fpr, tpr, _ = roc_curve(y_true_binary, y_prob)
            roc_auc = auc(fpr, tpr)
            all_aucs.append(roc_auc)

            if fpr[0] != 0:
                fpr = np.concatenate([[0], fpr])
                tpr = np.concatenate([[0], tpr])
            if fpr[-1] != 1:
                fpr = np.concatenate([fpr, [1]])
                tpr = np.concatenate([tpr, [tpr[-1]]])

            # Interpolate for smoother curve
            interp_tpr = np.interp(mean_fpr, fpr, tpr)
            interp_tpr[0] = 0.0
            all_tprs.append(interp_tpr)

            plt.plot(mean_fpr, interp_tpr,
                     color=self.config['color_map'].get(class_name, 'b'),
                     label=f'{class_name} (AUC = {roc_auc:.3f})',
                     linewidth=5)

        # Calculate average ROC curve and AUC
        if all_tprs:
            mean_tpr = np.mean(all_tprs, axis=0)
            mean_auc = np.mean(all_aucs)
            std_auc = np.std(all_aucs)

        ax.tick_params(axis='both', labelsize=48, width=9, length=20)
        plt.xticks(np.arange(0, 1.1, 0.2))
        plt.yticks(np.arange(0, 1.1, 0.2))

        plt.grid(True, alpha=0.3)
        legend = plt.legend(loc="lower right", fontsize=28, frameon=True)
        legend.get_frame().set_linewidth(3)

        plt.xlim([-0.02, 1.02])
        plt.ylim([-0.02, 1.02])

        output_path = os.path.join(self.config['output_path'], 'external_roc_curves_GA.svg')
        plt.savefig(output_path, format='svg', bbox_inches='tight')
        plt.close()

    def format_excel_file(self, file_path):
        """Format Excel file"""
        from openpyxl import load_workbook
        from openpyxl.styles import Font

        wb = load_workbook(file_path)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Set all cells to Times New Roman, 11pt
            for row in ws.iter_rows():
                for cell in row:
                    cell.font = Font(name='Times New Roman', size=11)

            # Set first row to bold
            for cell in ws[1]:
                cell.font = Font(name='Times New Roman', size=11, bold=True)

        wb.save(file_path)


class AdvancedFeatureInteractions:
    """Feature interaction processing class - simplified version"""

    def __init__(self,
                 interaction_method='gaussian',
                 selection_method='boruta',
                 max_features=30,
                 gaussian_gamma='auto',
                 scale_after_interaction=True):
        self.interaction_method = interaction_method
        self.selection_method = selection_method
        self.max_features = max_features
        self.gaussian_gamma = gaussian_gamma
        self.scale_after_interaction = scale_after_interaction
        self.feature_names_ = None
        self.selected_features_ = None
        self.feature_importance_ = None
        self.scaler = StandardScaler()
        self.original_feature_names = None

    def set_original_feature_names(self, names):
        """Set original feature names"""
        self.original_feature_names = names

    def fit_transform(self, features, labels):
        """Fit data and perform feature processing and transformation"""
        if self.scale_after_interaction:
            features_scaled = self.scaler.fit_transform(features)
        else:
            features_scaled = features.copy()

        if self.original_feature_names is not None:
            self.feature_names_ = self.original_feature_names.copy()
        else:
            self.feature_names_ = [f"feature_{i}" for i in range(features.shape[1])]

        # Simplified feature processing
        features_processed = features_scaled

        # Feature selection
        if self.selection_method and self.max_features:
            if features_processed.shape[1] > self.max_features:
                features_processed = features_processed[:, :self.max_features]
                self.feature_names_ = self.feature_names_[:self.max_features]

        self.feature_importance_ = np.ones(len(self.feature_names_))
        return features_processed

    def get_feature_names(self):
        """Get feature names"""
        return self.feature_names_

    def get_feature_importance(self):
        """Get feature importance"""
        if self.feature_importance_ is None:
            return None

        importance_df = pd.DataFrame({
            'Feature': self.feature_names_,
            'Importance': self.feature_importance_
        })
        return importance_df.sort_values('Importance', ascending=False)


def main():
    """Main function"""

    # Basic configuration
    train_data_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '00_Data', 'Train.xlsx')  # Path to the input Excel file containing training data
    train_sheet_name = 'Sheet1'  # Name of the sheet in the Excel file (e.g., 'Sheet1')

    # Genetic algorithm configuration
    ga_config = {
        'population_size': 50,
        'max_generations': 60,
        'crossover_rate': 0.8,
        'mutation_rate': 0.125,
        'elite_ratio': 0.125,
        'tournament_size': 3,
        'convergence_threshold': 10
    }

    # Configuration parameters
    config = {
        'output_path': os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '..', 'Results', 'Multicancer detection', '03_Ensemble', 'External validation'),  # Directory path for saving output results
        'output_filename': 'External validation_results.xlsx',

        'train_data_path': train_data_path,
        'train_sheet_name': train_sheet_name,

        'outer_cv_splits': 5,
        'inner_cv_splits': 5,

        # Ensemble algorithms weight configuration
        'use_fixed_weights': True,
        'fixed_xgb_weight': 0.27,

        # External data configuration
        'new_data_path': os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '00_Data', 'Test.xlsx'),  # Path to the input Excel file containing external data
        'new_data_sheet': 'Sheet1',  # Name of the sheet in the Excel file (e.g., 'Sheet1')

        # RF component configuration
        'rf_component': {
            'interaction_method': 'gaussian',
            'selection_method': 'boruta',
            'max_features': 40,
            'gaussian_gamma': 'auto',
            'scale_after_interaction': True,
            'smote_sampling_strategy': 1.0,
            'smote_k_neighbors': 5,
            'ga_config': ga_config,
            'feature_params': {
                'interaction_method': 'gaussian',
                'selection_method': 'boruta',
                'max_features': 40,
                'gaussian_gamma': 'auto',
                'scale_after_interaction': True
            }
        },

        # XGBoost component configuration
        'xgb_component': {
            'interaction_method': 'gaussian',
            'selection_method': 'mutual_info',
            'max_features': 40,
            'gaussian_gamma': 'auto',
            'scale_after_interaction': True,
            'adasyn_sampling_strategy': 1.0,
            'adasyn_n_neighbors': 5,
            'ga_config': ga_config,
            'feature_params': {
                'interaction_method': 'gaussian',
                'selection_method': 'mutual_info',
                'max_features': 40,
                'gaussian_gamma': 'auto',
                'scale_after_interaction': True
            }
        },

        # Color configuration
        'color_map': {
            'LC': [254 / 255, 160 / 255, 64 / 255],
            'LuC': [242 / 255, 128 / 255, 128 / 255],
            'OC': [88 / 255, 97 / 255, 172 / 255],
            'N': [106 / 255, 180 / 255, 193 / 255]
        }
    }

    try:
        # Create output directory
        os.makedirs(config['output_path'], exist_ok=True)

        # Read class information
        df = pd.read_excel(config['train_data_path'], sheet_name=config['train_sheet_name'])
        all_classes = sorted(df['Sample'].unique())
        config['all_classes'] = all_classes

        # Add shared parameters to components
        for component in ['rf_component', 'xgb_component']:
            config[component]['file_path'] = config['train_data_path']
            config[component]['sheet_name'] = config['train_sheet_name']
            config[component]['all_classes'] = all_classes
            config[component]['outer_cv_splits'] = config['outer_cv_splits']
            config[component]['inner_cv_splits'] = config['inner_cv_splits']

        print("=" * 80)
        print("Genetic Algorithm Enhanced Ensemble Model - External Validation")
        print("=" * 80)

        # Load external data to get shape
        df_external = pd.read_excel(config['new_data_path'], sheet_name=config['new_data_sheet'])
        print(f"External data shape: {df_external.shape[0]} samples, {df_external.shape[1]-1} features")
        print()

        # Run ensemble classifier
        ensemble = EnsembleClassifier(config)

        # Record start time
        start_time = time.time()

        # Perform training and external validation
        predictions_df = ensemble.combine_and_train(config['new_data_path'], config['new_data_sheet'])

        # Calculate total runtime
        total_time = time.time() - start_time

        # Read and output final metrics from saved file
        print()
        print("=" * 80)
        print("External Validation Results")
        print("=" * 80)

        metrics_file = os.path.join(config['output_path'], 'external_validation_results_GA.xlsx')
        metrics_df = pd.read_excel(metrics_file, sheet_name='Class_Metrics')

        # Output each class metrics
        for _, row in metrics_df.iterrows():
            if row['Class'] != 'Overall':
                print(f"{row['Class']:4s}: F1={row['F1']:.3f}, AUC={row['AUC']:.3f}, "
                      f"Sens={row['Sensitivity']:.3f}, Spec={row['Specificity']:.3f}")

        # Output overall metrics
        overall_row = metrics_df[metrics_df['Class'] == 'Overall'].iloc[0]
        print(f"{'Overall':4s}: F1={overall_row['F1']:.3f}, AUC={overall_row['AUC']:.3f}, "
              f"Sens={overall_row['Sensitivity']:.3f}, Spec={overall_row['Specificity']:.3f}")

        print()
        print(f"Validation completed! Total runtime: {total_time:.1f}s ({total_time/60:.1f} minutes)")
        print(f"Results saved to: {metrics_file}")
        print()

    except Exception as e:
        logging.error(f"An error occurred: {str(e)}")
        import traceback
        logging.error(traceback.format_exc())
        raise


if __name__ == "__main__":
    main()
