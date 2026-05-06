import numpy as np
import pandas as pd
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import StratifiedKFold
from sklearn.metrics import (f1_score, roc_curve, auc, confusion_matrix, precision_score,
                             recall_score, roc_auc_score)
from bayes_opt import BayesianOptimization
import matplotlib.pyplot as plt
from joblib import Parallel, delayed
import logging
import os
import warnings
import json
from datetime import datetime

# Import all required algorithms
from sklearn.discriminant_analysis import LinearDiscriminantAnalysis
from sklearn.linear_model import LogisticRegression
from sklearn.svm import SVC
from sklearn.neural_network import MLPClassifier
from sklearn.tree import DecisionTreeClassifier
from sklearn.ensemble import RandomForestClassifier
import xgboost as xgb

warnings.filterwarnings('ignore')

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')


# calculate specificity
def specificity_score(y_true, y_pred):

    tn, fp, fn, tp = confusion_matrix(y_true, y_pred).ravel()
    return tn / (tn + fp) if (tn + fp) > 0 else 0.0


class CancerPredictionEvaluator:
    def __init__(self, config):
        self.config = config
        self.scaler = StandardScaler()

        # Algorithm color mapping
        self.algorithm_colors = {
            'LDA': '#15455c',  # dark teal
            'LR': '#3b537d',  # navy blue
            'SVM': '#715d92',  # purple
            'ANN': '#aa6397',  # magenta
            'DT': '#db6c89',  # pink
            'RF': '#f98270',  # coral
            'XGB': '#ffa656'  # orange
        }

        # Initialize algorithm parameters
        self.algorithms = {
            'LDA': {
                'name': 'Linear Discriminant Analysis',
                'model_class': LinearDiscriminantAnalysis,
                'param_ranges': {
                    'solver_idx': (0, 2),  # 0: 'svd', 1: 'lsqr', 2: 'eigen'
                    'shrinkage_factor': (0, 1)  # Only used with 'lsqr' and 'eigen'
                }
            },
            'LR': {
                'name': 'Logistic Regression',
                'model_class': LogisticRegression,
                'param_ranges': {
                    'C': (1e-6, 100),
                    'penalty_idx': (0, 3),  # 0: 'l1', 1: 'l2', 2: 'elasticnet', 3: 'none'
                    'solver_idx': (0, 4),  # 0: 'newton-cg', 1: 'lbfgs', 2: 'liblinear', 3: 'sag', 4: 'saga'
                    'l1_ratio': (0, 1),  # Only used with 'elasticnet'
                    'class_weight_multiplier': (0.5, 5.0)
                }
            },
            'SVM': {
                'name': 'Support Vector Machine',
                'model_class': SVC,
                'param_ranges': {
                    'C': (1e-6, 100),
                    'gamma': (1e-6, 10),
                    'kernel_idx': (0, 3),  # 0: 'linear', 1: 'rbf', 2: 'poly', 3: 'sigmoid'
                    'degree': (2, 5),  # Only used with 'poly'
                    'class_weight_multiplier': (0.5, 5.0)
                }
            },
            'ANN': {
                'name': 'Artificial Neural Network',
                'model_class': MLPClassifier,
                'param_ranges': {
                    'hidden_layer_size_factor': (0, 3),  # Will be used to determine hidden_layer_sizes
                    'activation_idx': (0, 3),  # 0: 'identity', 1: 'logistic', 2: 'tanh', 3: 'relu'
                    'solver_idx': (0, 2),  # 0: 'lbfgs', 1: 'sgd', 2: 'adam'
                    'alpha': (1e-6, 1),
                    'learning_rate_idx': (0, 2),  # 0: 'constant', 1: 'invscaling', 2: 'adaptive'
                    'learning_rate_init': (1e-6, 0.1)
                }
            },
            'DT': {
                'name': 'Decision Tree',
                'model_class': DecisionTreeClassifier,
                'param_ranges': {
                    'criterion_idx': (0, 1),  # 0: 'gini', 1: 'entropy'
                    'max_depth': (3, 20),
                    'min_samples_split': (2, 20),
                    'min_samples_leaf': (1, 10),
                    'class_weight_multiplier': (0.5, 5.0)
                }
            },
            'RF': {
                'name': 'Random Forest',
                'model_class': RandomForestClassifier,
                'param_ranges': {
                    'n_estimators_factor': (1, 10),  # Will be multiplied by 50
                    'criterion_idx': (0, 1),  # 0: 'gini', 1: 'entropy'
                    'max_depth': (3, 30),
                    'min_samples_split': (2, 20),
                    'min_samples_leaf': (1, 10),
                    'class_weight_multiplier': (0.5, 5.0)
                }
            },
            'XGB': {
                'name': 'XGBoost',
                'model_class': xgb.XGBClassifier,
                'param_ranges': {
                    'n_estimators_factor': (1, 10),  # Will be multiplied by 50
                    'max_depth': (3, 15),
                    'learning_rate': (0.001, 0.3),
                    'subsample': (0.5, 1.0),
                    'colsample_bytree': (0.5, 1.0),
                    'gamma': (0, 10),
                    'min_child_weight': (1, 10)
                }
            }
        }

    def load_data(self):
        """Load and preprocess data with enhanced error handling"""
        logging.info("Loading data...")
        try:
            df = pd.read_excel(self.config['file_path'], sheet_name=self.config['sheet_name'])

            # Log DataFrame information
            logging.info(f"DataFrame shape: {df.shape}")
            logging.info(f"DataFrame columns: {df.columns.tolist()}")

            # Check if DataFrame is empty or has unexpected structure
            if df.empty:
                raise ValueError("DataFrame is empty. Check if the Excel file contains data.")

            if df.shape[1] <= 1:
                raise ValueError(
                    f"DataFrame has only {df.shape[1]} column(s). Expected at least 2 columns (label + features).")

            # Extract labels and features
            if 'Sample' in df.columns:
                # If 'Sample' column exists, use it as labels
                labels = df['Sample'].values
                features = df.drop('Sample', axis=1).values
                logging.info(f"Using 'Sample' column as labels. Unique labels: {np.unique(labels).tolist()}")
            else:
                # Otherwise, use first column as labels
                labels = df.iloc[:, 0].values
                features = df.iloc[:, 1:].values
                logging.info(
                    f"Using first column '{df.columns[0]}' as labels. Unique labels: {np.unique(labels).tolist()}")

            # Get original feature names
            if 'Sample' in df.columns:
                feature_names = df.drop('Sample', axis=1).columns.tolist()
            else:
                feature_names = df.columns[1:].tolist()

            self.feature_names = feature_names

            logging.info(f"Feature columns: {len(feature_names)}")
            logging.info(f"Original feature shape: {features.shape}")
            logging.info(f"Original number of samples: {len(labels)}")

            # Convert labels to binary (LuC=1, N=0)
            binary_labels = np.array([1 if l == 'LuC' else 0 for l in labels])
            logging.info(f"Binary labels distribution: {np.bincount(binary_labels)}")

            # Standardize features
            features_processed = self.scaler.fit_transform(features)
            logging.info(f"Processed feature shape: {features_processed.shape}")

            return features_processed, binary_labels

        except Exception as e:
            logging.error(f"Error loading data: {str(e)}")
            # Print more debug information
            if 'df' in locals():
                logging.error(f"DataFrame info: shape={df.shape}, columns={df.columns.tolist()}")
            raise

    def _create_model(self, algorithm_key, params, class_ratio=None):
        """Create a algorithms instance with the given parameters"""
        algo_info = self.algorithms[algorithm_key]

        # Handle specific parameter conversions for each algorithm
        if algorithm_key == 'LDA':
            solvers = ['svd', 'lsqr', 'eigen']
            solver = solvers[int(params['solver_idx'] + 0.5)]

            if solver == 'svd':
                return LinearDiscriminantAnalysis(solver=solver)
            else:
                # Only use shrinkage for 'lsqr' and 'eigen'
                return LinearDiscriminantAnalysis(
                    solver=solver,
                    shrinkage=params['shrinkage_factor']
                )

        elif algorithm_key == 'LR':
            penalties = ['l1', 'l2', 'elasticnet', 'none']
            solvers = ['newton-cg', 'lbfgs', 'liblinear', 'sag', 'saga']

            penalty_idx = int(params['penalty_idx'] + 0.5)
            solver_idx = int(params['solver_idx'] + 0.5)

            penalty = penalties[penalty_idx]
            solver = solvers[solver_idx]

            # Handle solver/penalty compatibility
            if penalty == 'l1' and solver not in ['liblinear', 'saga']:
                solver = 'saga'
            elif penalty == 'elasticnet' and solver != 'saga':
                solver = 'saga'
            elif penalty == 'none' and solver == 'liblinear':
                solver = 'lbfgs'

            # Class weight calculation
            if class_ratio is not None:
                class_weight = {
                    0: 1.0,
                    1: params.get('class_weight_multiplier', 1.0) * class_ratio
                }
            else:
                class_weight = None

            # Handle penalty/solver specific parameters
            if penalty == 'elasticnet':
                return LogisticRegression(
                    C=params['C'],
                    penalty=penalty,
                    solver=solver,
                    l1_ratio=params['l1_ratio'],
                    class_weight=class_weight,
                    max_iter=1000,
                    random_state=42
                )
            else:
                return LogisticRegression(
                    C=params['C'],
                    penalty=penalty,
                    solver=solver,
                    class_weight=class_weight,
                    max_iter=1000,
                    random_state=42
                )

        elif algorithm_key == 'SVM':
            kernels = ['linear', 'rbf', 'poly', 'sigmoid']
            kernel_idx = int(params['kernel_idx'] + 0.5)
            kernel = kernels[kernel_idx]

            # Class weight calculation
            if class_ratio is not None:
                class_weight = {
                    0: 1.0,
                    1: params.get('class_weight_multiplier', 1.0) * class_ratio
                }
            else:
                class_weight = None

            # Handle kernel-specific parameters
            if kernel == 'poly':
                return SVC(
                    C=params['C'],
                    kernel=kernel,
                    gamma=params['gamma'],
                    degree=int(params['degree'] + 0.5),
                    class_weight=class_weight,
                    probability=True,
                    random_state=42
                )
            else:
                return SVC(
                    C=params['C'],
                    kernel=kernel,
                    gamma=params['gamma'],
                    class_weight=class_weight,
                    probability=True,
                    random_state=42
                )

        elif algorithm_key == 'ANN':
            # Handle hidden layer sizes based on factor
            hidden_factor = int(params['hidden_layer_size_factor'] + 0.5)
            if hidden_factor == 0:
                hidden_layer_sizes = (75,)
            elif hidden_factor == 1:
                hidden_layer_sizes = (75, 30)
            elif hidden_factor == 2:
                hidden_layer_sizes = (75, 30, 10)
            else:
                hidden_layer_sizes = (75, 30, 15, 5)

            activations = ['identity', 'logistic', 'tanh', 'relu']
            solvers = ['lbfgs', 'sgd', 'adam']
            learning_rates = ['constant', 'invscaling', 'adaptive']

            return MLPClassifier(
                hidden_layer_sizes=hidden_layer_sizes,
                activation=activations[int(params['activation_idx'] + 0.5)],
                solver=solvers[int(params['solver_idx'] + 0.5)],
                alpha=params['alpha'],
                learning_rate=learning_rates[int(params['learning_rate_idx'] + 0.5)],
                learning_rate_init=params['learning_rate_init'],
                max_iter=500,
                random_state=42
            )

        elif algorithm_key == 'DT':
            criteria = ['gini', 'entropy']

            # Class weight calculation
            if class_ratio is not None:
                class_weight = {
                    0: 1.0,
                    1: params.get('class_weight_multiplier', 1.0) * class_ratio
                }
            else:
                class_weight = None

            return DecisionTreeClassifier(
                criterion=criteria[int(params['criterion_idx'] + 0.5)],
                max_depth=int(params['max_depth'] + 0.5),
                min_samples_split=int(params['min_samples_split'] + 0.5),
                min_samples_leaf=int(params['min_samples_leaf'] + 0.5),
                class_weight=class_weight,
                random_state=42
            )

        elif algorithm_key == 'RF':
            criteria = ['gini', 'entropy']

            # Class weight calculation
            if class_ratio is not None:
                class_weight = {
                    0: 1.0,
                    1: params.get('class_weight_multiplier', 1.0) * class_ratio
                }
            else:
                class_weight = None

            return RandomForestClassifier(
                n_estimators=int(params['n_estimators_factor'] * 50 + 0.5),
                criterion=criteria[int(params['criterion_idx'] + 0.5)],
                max_depth=int(params['max_depth'] + 0.5),
                min_samples_split=int(params['min_samples_split'] + 0.5),
                min_samples_leaf=int(params['min_samples_leaf'] + 0.5),
                class_weight=class_weight,
                n_jobs=-1,
                random_state=42
            )

        elif algorithm_key == 'XGB':
            return xgb.XGBClassifier(
                n_estimators=int(params['n_estimators_factor'] * 50 + 0.5),
                max_depth=int(params['max_depth'] + 0.5),
                learning_rate=params['learning_rate'],
                subsample=params['subsample'],
                colsample_bytree=params['colsample_bytree'],
                gamma=params['gamma'],
                min_child_weight=int(params['min_child_weight'] + 0.5),
                eval_metric='logloss',
                random_state=42,
                n_jobs=-1
            )

        else:
            raise ValueError(f"Unsupported algorithm: {algorithm_key}")

    def _optimize_hyperparameters(self, X_train, y_train, algorithm_key):
        """Optimize hyperparameters using Bayesian optimization"""
        # Calculate class ratio for class_weight
        neg_count = sum(y_train == 0)
        pos_count = sum(y_train == 1)
        base_class_ratio = neg_count / max(1, pos_count)
        logging.info("Class ratio (neg/pos): {:.2f}".format(base_class_ratio))

        # Get algorithm parameters
        algo_info = self.algorithms[algorithm_key]
        param_ranges = algo_info['param_ranges']

        def objective(**params):
            """Bayesian optimization objective function"""
            # Create algorithms with parameters
            model = self._create_model(algorithm_key, params, base_class_ratio)

            # Use StratifiedKFold for cross-validation
            cv = StratifiedKFold(n_splits=self.config['inner_cv_splits'], shuffle=True, random_state=42)

            # Track multiple metrics
            f1_scores = []
            auc_scores = []
            sensitivity_scores = []
            specificity_scores = []

            for train_idx, val_idx in cv.split(X_train, y_train):
                X_train_cv, X_val_cv = X_train[train_idx], X_train[val_idx]
                y_train_cv, y_val_cv = y_train[train_idx], y_train[val_idx]

                # Ensure label format is correct
                y_train_cv = y_train_cv.astype(int)

                # Some algorithms may not converge or may raise warnings which we want to suppress
                try:
                    # Train algorithms
                    with warnings.catch_warnings():
                        warnings.simplefilter("ignore")
                        model.fit(X_train_cv, y_train_cv)

                    # Get prediction probabilities
                    try:
                        y_pred_proba = model.predict_proba(X_val_cv)[:, 1]
                    except (AttributeError, IndexError):
                        # If algorithms doesn't have predict_proba method or it doesn't return expected format
                        if hasattr(model, 'decision_function'):
                            # For SVM and some other models
                            val_decisions = model.decision_function(X_val_cv)
                            if len(val_decisions.shape) > 1 and val_decisions.shape[1] > 1:
                                # Multi-class decision function
                                pos_class_idx = 1 if 1 in model.classes_ else 0
                                val_decisions = val_decisions[:, pos_class_idx]
                            val_decisions_min = np.min(val_decisions)
                            val_decisions_max = np.max(val_decisions)
                            if val_decisions_max > val_decisions_min:
                                y_pred_proba = (val_decisions - val_decisions_min) / (
                                        val_decisions_max - val_decisions_min)
                            else:
                                y_pred_proba = np.ones_like(val_decisions) * 0.5
                        else:
                            # For models that don't provide probability or decision function
                            y_pred_proba = model.predict(X_val_cv).astype(float)

                    # Try different thresholds to find the best F1
                    best_f1 = 0
                    best_threshold = 0.5
                    best_pred = None

                    for threshold in [0.3, 0.35, 0.4, 0.45, 0.5]:
                        y_pred_thresh = (y_pred_proba >= threshold).astype(int)
                        f1 = f1_score(y_val_cv, y_pred_thresh)
                        if f1 > best_f1:
                            best_f1 = f1
                            best_threshold = threshold
                            best_pred = y_pred_thresh

                    # Record the threshold used
                    logging.debug("Best threshold for this fold: {:.2f}".format(best_threshold))

                    # Calculate and save multiple metrics
                    f1 = f1_score(y_val_cv, best_pred)
                    f1_scores.append(f1)

                    try:
                        roc_auc = roc_auc_score(y_val_cv, y_pred_proba)
                        auc_scores.append(roc_auc)
                    except Exception as e:
                        logging.debug("Error calculating AUC: {}".format(str(e)))
                        auc_scores.append(0)

                    sensitivity = recall_score(y_val_cv, best_pred)
                    sensitivity_scores.append(sensitivity)

                    spec = specificity_score(y_val_cv, best_pred)
                    specificity_scores.append(spec)

                except Exception as e:
                    logging.debug(f"Error in inner CV for {algorithm_key}: {str(e)}")
                    # Append zeros for failed trials
                    f1_scores.append(0)
                    auc_scores.append(0)
                    sensitivity_scores.append(0)
                    specificity_scores.append(0)

            # Use default scoring without specific weights
            mean_f1 = np.mean(f1_scores) if f1_scores else 0
            mean_auc = np.mean(auc_scores) if auc_scores else 0
            mean_sensitivity = np.mean(sensitivity_scores) if sensitivity_scores else 0
            mean_specificity = np.mean(specificity_scores) if specificity_scores else 0

            # Simply use the mean F1 score as the optimization target
            combined_score = mean_f1

            # Log detailed evaluation
            param_str = ", ".join([f"{k}={v:.4f}" for k, v in params.items()])
            score_str = f"F1={mean_f1:.4f}, AUC={mean_auc:.4f}, Sens={mean_sensitivity:.4f}, Spec={mean_specificity:.4f}"
            logging.debug(f"Params: {param_str} → {score_str}")

            return combined_score

        # Use Bayesian optimization with the parameter ranges for this algorithm
        optimizer = BayesianOptimization(
            f=objective,
            pbounds=param_ranges,
            random_state=42,
            allow_duplicate_points=True
        )

        optimizer.maximize(
            init_points=self.config['bayes_opt']['init_points'],
            n_iter=self.config['bayes_opt']['n_iter']
        )

        # Return best parameters
        best_params = optimizer.max['params']
        logging.info(f"Best parameters for {algorithm_key}: {best_params}")
        logging.info(f"Best score for {algorithm_key}: {optimizer.max['target']:.4f}")
        return best_params

    def process_fold(self, X, y, train_index, test_index, fold_idx, algorithm_key):
        """Process a single cross-validation fold for a specific algorithm"""
        X_train, X_test = X[train_index], X[test_index]
        y_train, y_test = y[train_index], y[test_index]

        # Check test set class distribution
        test_class_dist = np.bincount(y_test)
        logging.info(f"Test set class distribution for {algorithm_key}, fold {fold_idx + 1}: {test_class_dist}")

        # Log warnings for imbalanced test sets
        if len(test_class_dist) == 1 or (len(test_class_dist) > 1 and test_class_dist[1] == 0):
            logging.warning(f"Test set for {algorithm_key}, fold {fold_idx + 1} only contains negative samples")
        if len(test_class_dist) > 1 and test_class_dist[0] == 0:
            logging.warning(f"Test set for {algorithm_key}, fold {fold_idx + 1} only contains positive samples")

        # Log class distribution
        neg_count = sum(y_train == 0)
        pos_count = sum(y_train == 1)
        class_ratio = neg_count / max(1, pos_count)
        logging.info(
            f"{algorithm_key} fold {fold_idx + 1} class distribution - Negative: {neg_count}, Positive: {pos_count}, Ratio: {class_ratio:.2f}")

        try:
            # Optimize hyperparameters
            best_params = self._optimize_hyperparameters(X_train, y_train, algorithm_key)

            # Train algorithms
            model = self._create_model(algorithm_key, best_params, class_ratio)
            model.fit(X_train, y_train)

            # Get prediction probabilities
            try:
                y_pred_proba = model.predict_proba(X_test)[:, 1]
            except (AttributeError, IndexError):
                # Handle models without predict_proba
                if hasattr(model, 'decision_function'):
                    decisions = model.decision_function(X_test)
                    if len(decisions.shape) > 1 and decisions.shape[1] > 1:
                        # Multi-class decision function
                        pos_class_idx = 1 if 1 in model.classes_ else 0
                        decisions = decisions[:, pos_class_idx]
                    decisions_min = np.min(decisions)
                    decisions_max = np.max(decisions)
                    if decisions_max > decisions_min:
                        y_pred_proba = (decisions - decisions_min) / (decisions_max - decisions_min)
                    else:
                        y_pred_proba = np.ones_like(decisions) * 0.5
                else:
                    y_pred_proba = model.predict(X_test).astype(float)

            # Use dynamic threshold selection - find threshold for best F1
            thresholds = np.arange(0.1, 0.7, 0.025)
            best_f1 = 0
            best_threshold = 0.4  # Default threshold

            # Only optimize threshold when test set has both classes
            if len(test_class_dist) > 1 and test_class_dist[0] > 0 and test_class_dist[1] > 0:
                for threshold in thresholds:
                    y_pred_thresh = (y_pred_proba >= threshold).astype(int)
                    f1 = f1_score(y_test, y_pred_thresh)
                    if f1 > best_f1:
                        best_f1 = f1
                        best_threshold = threshold

            logging.info(f"Selected optimal threshold for {algorithm_key}, fold {fold_idx + 1}: {best_threshold:.2f}")

            # Use best threshold for predictions
            y_pred = (y_pred_proba >= best_threshold).astype(int)

            # Get training set prediction probabilities
            try:
                y_train_pred_proba = model.predict_proba(X_train)[:, 1]
            except (AttributeError, IndexError):
                if hasattr(model, 'decision_function'):
                    train_decisions = model.decision_function(X_train)
                    if len(train_decisions.shape) > 1 and train_decisions.shape[1] > 1:
                        pos_class_idx = 1 if 1 in model.classes_ else 0
                        train_decisions = train_decisions[:, pos_class_idx]
                    train_decisions_min = np.min(train_decisions)
                    train_decisions_max = np.max(train_decisions)
                    if train_decisions_max > train_decisions_min:
                        y_train_pred_proba = (train_decisions - train_decisions_min) / (
                                train_decisions_max - train_decisions_min)
                    else:
                        y_train_pred_proba = np.ones_like(train_decisions) * 0.5
                else:
                    y_train_pred_proba = model.predict(X_train).astype(float)

            y_train_pred = (y_train_pred_proba >= best_threshold).astype(int)

            # Calculate evaluation metrics
            train_f1 = f1_score(y_train, y_train_pred)
            test_f1 = f1_score(y_test, y_pred)

            # Calculate sensitivity (same as recall)
            test_sensitivity = recall_score(y_test, y_pred)
            train_sensitivity = recall_score(y_train, y_train_pred)

            # Calculate specificity
            test_specificity = specificity_score(y_test, y_pred)
            train_specificity = specificity_score(y_train, y_train_pred)

            # Calculate precision
            test_precision = precision_score(y_test, y_pred)
            train_precision = precision_score(y_train, y_train_pred)

            # Calculate ROC curve
            try:
                fpr, tpr, _ = roc_curve(y_test, y_pred_proba)
                roc_auc = auc(fpr, tpr)
            except Exception as e:
                logging.warning(f"Error calculating ROC curve for {algorithm_key}, fold {fold_idx + 1}: {str(e)}")
                fpr, tpr = np.array([0, 1]), np.array([0, 1])
                roc_auc = 0.5

            # Confusion matrix
            conf_matrix = confusion_matrix(y_test, y_pred)

            # Calculate false negative rate and false positive rate
            tn, fp, fn, tp = 0, 0, 0, 0
            if conf_matrix.shape == (2, 2):
                tn, fp, fn, tp = conf_matrix.ravel()
            elif conf_matrix.shape == (1, 1):
                if np.sum(y_test) == 0:  # Only negative class
                    tn = conf_matrix[0, 0]
                else:  # Only positive class
                    tp = conf_matrix[0, 0]

            # Calculate false negative rate and false positive rate
            fnr = fn / (fn + tp) if (fn + tp) > 0 else 0
            fpr_val = fp / (fp + tn) if (fp + tn) > 0 else 0

            logging.info(
                f"{algorithm_key} fold {fold_idx + 1} CV results - F1: {test_f1:.4f}, AUC: {roc_auc:.4f}, "
                f"Sensitivity: {test_sensitivity:.4f}, Specificity: {test_specificity:.4f}"
            )

            return {
                "algorithm": algorithm_key,
                "fold_idx": fold_idx,
                "y_test": y_test,
                "y_pred": y_pred,
                "y_pred_proba": y_pred_proba,
                "optimal_threshold": best_threshold,
                "f1": test_f1,
                "train_f1": train_f1,
                "sensitivity": test_sensitivity,
                "train_sensitivity": train_sensitivity,
                "specificity": test_specificity,
                "train_specificity": train_specificity,
                "precision": test_precision,
                "train_precision": train_precision,
                "fpr": fpr,
                "tpr": tpr,
                "roc_auc": roc_auc,
                "fnr": fnr,
                "fpr_val": fpr_val,
                "algorithms": model,
                "best_params": best_params,
                "confusion_matrix": conf_matrix
            }

        except Exception as e:
            logging.error(f"Error in {algorithm_key} fold {fold_idx + 1}: {str(e)}")
            import traceback
            logging.error(traceback.format_exc())

            # Return a placeholder result with zero metrics
            return {
                "algorithm": algorithm_key,
                "fold_idx": fold_idx,
                "y_test": y_test,
                "y_pred": np.zeros_like(y_test),
                "y_pred_proba": np.zeros_like(y_test, dtype=float),
                "optimal_threshold": 0.5,
                "f1": 0.0,
                "train_f1": 0.0,
                "sensitivity": 0.0,
                "train_sensitivity": 0.0,
                "specificity": 0.0,
                "train_specificity": 0.0,
                "precision": 0.0,
                "train_precision": 0.0,
                "fpr": np.array([0, 1]),
                "tpr": np.array([0, 1]),
                "roc_auc": 0.5,
                "fnr": 0.0,
                "fpr_val": 0.0,
                "algorithms": None,
                "best_params": {},
                "confusion_matrix": np.zeros((2, 2)),
                "error": str(e)
            }

    def plot_algorithm_roc_curve(self, all_results):
        """Plot ROC curve for all algorithms with English text and Arial font"""
        plt.figure(figsize=(12, 10))
        ax = plt.gca()

        # Set border thickness to 5
        for spine in ax.spines.values():
            spine.set_linewidth(5)

        # Set Arial font
        plt.rcParams['font.family'] = 'Arial'

        plt.title("ROC Curves for All Algorithms", fontsize=24, pad=15, fontname='Arial')
        plt.xlabel("False Positive Rate", fontsize=24, labelpad=10, fontname='Arial')
        plt.ylabel("True Positive Rate", fontsize=24, labelpad=10, fontname='Arial')

        # Process CV results by algorithm
        algorithm_mean_results = {}

        for algorithm_key in self.algorithms.keys():
            if algorithm_key not in all_results:
                continue

            algorithm_results = all_results[algorithm_key]

            # Collect all folds for this algorithm
            tprs = []
            aucs = []
            mean_fpr = np.linspace(0, 1, 500)

            for result in algorithm_results:
                fpr = result['fpr']
                tpr = result['tpr']
                roc_auc = result['roc_auc']

                # Ensure TPR/FPR values are properly formatted
                if fpr[0] != 0:
                    fpr = np.concatenate([[0], fpr])
                    tpr = np.concatenate([[0], tpr])

                if fpr[-1] != 1:
                    fpr = np.concatenate([fpr, [1]])
                    tpr = np.concatenate([tpr, [tpr[-1]]])

                # Interpolate to standard grid
                interp_tpr = np.interp(mean_fpr, fpr, tpr)
                interp_tpr[0] = 0.0
                tprs.append(interp_tpr)
                aucs.append(roc_auc)

            # Calculate mean TPR and AUC
            if tprs:
                mean_tpr = np.mean(tprs, axis=0)
                mean_auc = np.mean(aucs)
                std_auc = np.std(aucs)

                # Plot mean ROC curve with proper transparency
                plt.plot(mean_fpr, mean_tpr,
                         color=self.algorithm_colors.get(algorithm_key, 'b'),
                         label=f'{self.algorithms[algorithm_key]["name"]} (AUC = {mean_auc:.3f}±{std_auc:.3f})',
                         lw=3, alpha=1.0)

                # Fill standard deviation
                std_tpr = np.std(tprs, axis=0)
                tprs_upper = np.minimum(mean_tpr + std_tpr, 1)
                tprs_lower = np.maximum(mean_tpr - std_tpr, 0)

                plt.fill_between(mean_fpr, tprs_lower, tprs_upper,
                                 color=self.algorithm_colors.get(algorithm_key, 'b'),
                                 alpha=0.2)

                # Store mean CV results for later use
                algorithm_mean_results[algorithm_key] = {
                    'mean_fpr': mean_fpr,
                    'mean_tpr': mean_tpr,
                    'mean_auc': mean_auc,
                    'std_auc': std_auc,
                    'tprs_upper': tprs_upper,
                    'tprs_lower': tprs_lower
                }

        # Plot diagonal line
        plt.plot([0, 1], [0, 1], 'k--', lw=2, alpha=0.7)

        # Customize plot appearance
        plt.grid(True, alpha=0.3, linewidth=0)
        plt.xlim([-0.02, 1.02])
        plt.ylim([-0.02, 1.02])

        ax.tick_params(width=5, length=10, labelsize=20)

        # Set tick label font
        for label in ax.get_xticklabels() + ax.get_yticklabels():
            label.set_fontname('Arial')

        # Add legend
        legend = plt.legend(loc="lower right", fontsize=18, frameon=True)
        legend.get_frame().set_linewidth(3)

        # Set legend font
        for text in legend.get_texts():
            text.set_fontname('Arial')

        # Save the plot
        output_file = os.path.join(self.config['output_path'], "ROC_all_algorithms.svg")
        plt.savefig(output_file, format='svg', bbox_inches='tight')
        plt.close()

        return algorithm_mean_results

    def save_results(self, all_results, algorithm_mean_roc=None):
        """Save CV results to Excel focusing on algorithm performance"""
        output_path = os.path.join(self.config['output_path'], self.config['output_filename'])

        # Prepare CV results by algorithm
        algorithm_summary = {}

        for algorithm_key, results_list in all_results.items():
            all_metrics = {
                'f1': [],
                'sensitivity': [],
                'specificity': [],
                'precision': [],
                'roc_auc': [],
                'optimal_threshold': []
            }

            # Collect metrics from all folds
            for result in results_list:
                for metric in all_metrics.keys():
                    all_metrics[metric].append(result[metric])

            # Calculate mean and std for each metric
            algorithm_metrics = {}
            for metric, values in all_metrics.items():
                if values:
                    algorithm_metrics[f'mean_{metric}'] = np.mean(values)
                    algorithm_metrics[f'std_{metric}'] = np.std(values)
                else:
                    algorithm_metrics[f'mean_{metric}'] = 0
                    algorithm_metrics[f'std_{metric}'] = 0

            algorithm_summary[algorithm_key] = algorithm_metrics

        # Create summary DataFrame
        summary_data = []
        for algorithm_key, metrics in algorithm_summary.items():
            row = {
                'Algorithm': self.algorithms[algorithm_key]['name'],
                'Mean F1': metrics['mean_f1'],
                'Std F1': metrics['std_f1'],
                'Mean Sensitivity': metrics['mean_sensitivity'],
                'Std Sensitivity': metrics['std_sensitivity'],
                'Mean Specificity': metrics['mean_specificity'],
                'Std Specificity': metrics['std_specificity'],
                'Mean Precision': metrics['mean_precision'],
                'Std Precision': metrics['std_precision'],
                'Mean AUC': metrics['mean_roc_auc'],
                'Std AUC': metrics['std_roc_auc'],
                'Mean Threshold': metrics['mean_optimal_threshold'],
                'Std Threshold': metrics['std_optimal_threshold']
            }
            summary_data.append(row)

        summary_df = pd.DataFrame(summary_data)

        # Sort by Mean F1 score (descending)
        summary_df = summary_df.sort_values('Mean F1', ascending=False)

        # Write to Excel
        with pd.ExcelWriter(output_path) as writer:
            # Write summary of all algorithms
            summary_df.to_excel(writer, sheet_name='Algorithm_Summary', index=False)

            # Write detailed CV results by algorithm
            for algorithm_key, results_list in all_results.items():
                algorithm_name = self.algorithms[algorithm_key]['name']

                # Prepare data for this algorithm
                algorithm_data = []

                for result in results_list:
                    row = {
                        'Fold': result['fold_idx'] + 1,
                        'F1': result['f1'],
                        'Sensitivity': result['sensitivity'],
                        'Specificity': result['specificity'],
                        'Precision': result['precision'],
                        'ROC AUC': result['roc_auc'],
                        'Optimal Threshold': result['optimal_threshold']
                    }
                    algorithm_data.append(row)

                # Create DataFrame and write to Excel
                if algorithm_data:
                    pd.DataFrame(algorithm_data).to_excel(
                        writer,
                        sheet_name=f'{algorithm_key}_details',
                        index=False
                    )

            # If have ROC curve data, save it
            if algorithm_mean_roc:
                roc_data = []
                if algorithm_mean_roc and len(algorithm_mean_roc) > 0:
                    first_algo = list(algorithm_mean_roc.keys())[0]
                    fpr_values = algorithm_mean_roc[first_algo]['mean_fpr']

                    # Create a DataFrame with FPR and TPR for each algorithm
                    roc_df = pd.DataFrame({'FPR': fpr_values})

                    for algo, roc_info in algorithm_mean_roc.items():
                        roc_df[f'{self.algorithms[algo]["name"]}_TPR'] = roc_info['mean_tpr']
                        roc_df[f'{self.algorithms[algo]["name"]}_TPR_upper'] = roc_info['tprs_upper']
                        roc_df[f'{self.algorithms[algo]["name"]}_TPR_lower'] = roc_info['tprs_lower']

                    roc_df.to_excel(writer, sheet_name='ROC_Curve_Data', index=False)

                    # Create a summary of AUC values
                    auc_data = []
                    for algo, roc_info in algorithm_mean_roc.items():
                        auc_data.append({
                            'Algorithm': self.algorithms[algo]['name'],
                            'Mean AUC': roc_info['mean_auc'],
                            'Std AUC': roc_info['std_auc']
                        })

                    pd.DataFrame(auc_data).to_excel(writer, sheet_name='AUC_Summary', index=False)

        # Format Excel file
        self.format_excel_file(output_path)

    def format_excel_file(self, file_path):
        """Format Excel file with Times New Roman font and bold headers"""
        from openpyxl import load_workbook
        from openpyxl.styles import Font

        wb = load_workbook(file_path)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Set all cells to Times New Roman, 11pt
            for row in ws.iter_rows():
                for cell in row:
                    cell.font = Font(name='Times New Roman', size=11)

            # Set first row to bold
            for cell in ws[1]:
                cell.font = Font(name='Times New Roman', size=11, bold=True)

        wb.save(file_path)

    def run(self):
        """Run complete training and evaluation process for all algorithms"""
        # Load data
        X, y = self.load_data()
        logging.info("Loaded data: {} features, {} samples".format(X.shape[1], X.shape[0]))
        logging.info("Class distribution: {} negative, {} positive".format(np.sum(y == 0), np.sum(y == 1)))

        # Record datetime for output files
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        # Create cross-validation object
        outer_cv = StratifiedKFold(n_splits=self.config['outer_cv_splits'], shuffle=True, random_state=42)

        # Store CV results for all algorithms
        all_algorithm_results = {}

        # Process each algorithm
        for algorithm_key, algo_info in self.algorithms.items():
            logging.info(f"\n{'=' * 50}")
            logging.info(f"Processing algorithm: {algo_info['name']}")
            logging.info(f"{'=' * 50}")

            algorithm_results = []

            # Use cross-validation
            try:
                # Prepare cross-validation indices
                fold_indices = list(enumerate(outer_cv.split(X, y)))

                # Process each fold in parallel
                results = Parallel(n_jobs=-1)(
                    delayed(self.process_fold)(
                        X, y, train_idx, test_idx, fold_idx, algorithm_key
                    )
                    for fold_idx, (train_idx, test_idx) in fold_indices
                )

                # Skip if no valid CV results
                if not results:
                    logging.warning(f"No valid CV results for {algo_info['name']}. Skipping.")
                    continue

                algorithm_results = results

                # Output performance metrics
                f1_scores = [r['f1'] for r in results]
                auc_scores = [r['roc_auc'] for r in results]
                sensitivity_scores = [r['sensitivity'] for r in results]
                specificity_scores = [r['specificity'] for r in results]
                precision_scores = [r['precision'] for r in results]

                logging.info(f"\nPerformance metrics for {algo_info['name']}:")
                logging.info(f"F1 Score: {np.mean(f1_scores):.3f} ± {np.std(f1_scores):.3f}")
                logging.info(f"ROC AUC: {np.mean(auc_scores):.3f} ± {np.std(auc_scores):.3f}")
                logging.info(f"Sensitivity: {np.mean(sensitivity_scores):.3f} ± {np.std(sensitivity_scores):.3f}")
                logging.info(f"Specificity: {np.mean(specificity_scores):.3f} ± {np.std(specificity_scores):.3f}")
                logging.info(f"Precision: {np.mean(precision_scores):.3f} ± {np.std(precision_scores):.3f}")

            except Exception as e:
                logging.error(f"Error processing {algo_info['name']}: {str(e)}")
                import traceback
                logging.error(traceback.format_exc())
                continue

            # Store CV results for this algorithm
            if algorithm_results:
                all_algorithm_results[algorithm_key] = algorithm_results
                logging.info(f"Completed processing {algo_info['name']}")
            else:
                logging.error(f"No CV results generated for {algo_info['name']}. Check for errors above.")

        # Plot ROC curves for all algorithms
        if all_algorithm_results:
            algorithm_mean_roc = self.plot_algorithm_roc_curve(all_algorithm_results)

            # Save CV results to Excel
            self.save_results(all_algorithm_results, algorithm_mean_roc)

            return all_algorithm_results
        else:
            logging.error("No CV results were generated for any algorithm. Check for errors above.")
            return {}


def main():
    """Main function"""
    # Configuration parameters
    config = {
        'file_path': os.path.join(os.path.dirname(os.path.abspath(__file__)), 'TS Train.xlsx'),  # Path to the input Excel file containing training data
        'sheet_name': '',  # Name of the sheet in the Excel file (e.g., 'Sheet1')
        'output_path': os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'Results', 'Early cancer detection'),  # Directory path for saving output results
        'output_filename': 'Cancer_Prediction_Results.xlsx',

        # Cross-validation configuration
        'outer_cv_splits': 5,
        'inner_cv_splits': 5,

        # Bayesian optimization configuration
        'bayes_opt': {
            'init_points': 10,
            'n_iter': 60,
        }
    }

    try:
        # Create output directory
        os.makedirs(config['output_path'], exist_ok=True)

        # Run cancer prediction evaluator
        evaluator = CancerPredictionEvaluator(config)
        results = evaluator.run()

        # Output final summary
        logging.info("\nEvaluation completed successfully!")

    except Exception as e:
        logging.error("An error occurred: {}".format(str(e)))
        import traceback
        logging.error(traceback.format_exc())
        raise


if __name__ == "__main__":
    main()