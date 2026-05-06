# -*- coding: utf-8 -*-
import numpy as np
import pandas as pd
import os
import logging
import warnings
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import (
    accuracy_score, f1_score, confusion_matrix,
    precision_score, recall_score, roc_curve, auc,
    precision_recall_curve, average_precision_score, roc_auc_score
)
from sklearn.model_selection import StratifiedKFold, train_test_split
import importlib.util
import sys
from scipy.spatial.distance import pdist, squareform
from scipy.stats import multivariate_normal
from datetime import datetime
import json
from tqdm import tqdm
import time

# Optuna imports
import optuna
from optuna.samplers import TPESampler
from optuna.pruners import HyperbandPruner

# Ignore warnings
warnings.filterwarnings('ignore')

logging.basicConfig(
    level=logging.INFO,
    format='%(message)s',
    handlers=[
        logging.FileHandler("external_validation.log"),
        logging.StreamHandler()
    ]
)

optuna.logging.set_verbosity(optuna.logging.WARNING)


def specificity_score(y_true, y_pred):
    """Calculate specificity"""
    tn, fp, fn, tp = confusion_matrix(y_true, y_pred).ravel()
    return tn / (tn + fp) if (tn + fp) > 0 else 0.0


class SyntheticDataGenerator:

    @staticmethod
    def generate_synthetic_samples(X, y, multiplier=5, noise_range=(0.05, 0.15), preserve_correlation=True):
        """Generate synthetic samples"""

        n_samples, n_features = X.shape
        X_synthetic = np.zeros((n_samples * multiplier, n_features))
        y_synthetic = np.zeros(n_samples * multiplier, dtype=y.dtype)

        # Copy original data
        X_synthetic[:n_samples] = X.copy()
        y_synthetic[:n_samples] = y.copy()

        unique_classes = np.unique(y)

        if preserve_correlation:
            for class_label in unique_classes:
                class_mask = (y == class_label)
                X_class = X[class_mask]

                if len(X_class) < 3:
                    for i in range(1, multiplier):
                        for j, x_orig in enumerate(X_class):
                            idx = n_samples * i + np.where(class_mask)[0][j]
                            noise_factor = np.random.uniform(noise_range[0], noise_range[1])
                            X_synthetic[idx] = x_orig + np.random.normal(0, noise_factor * np.std(X_class, axis=0),
                                                                         n_features)
                            y_synthetic[idx] = class_label
                    continue

                # Calculate covariance matrix
                mean_vector = np.mean(X_class, axis=0)
                cov_matrix = np.cov(X_class, rowvar=False)
                cov_matrix = (cov_matrix + cov_matrix.T) / 2

                # Ensure positive definiteness
                eigvals, eigvecs = np.linalg.eigh(cov_matrix)
                min_eig = 1e-6
                eigvals = np.maximum(eigvals, min_eig)
                cov_matrix = eigvecs.dot(np.diag(eigvals)).dot(eigvecs.T)
                cov_matrix += np.eye(cov_matrix.shape[0]) * 1e-6

                for i in range(1, multiplier):
                    for j, x_orig in enumerate(X_class):
                        idx = n_samples * i + np.where(class_mask)[0][j]

                        try:
                            random_sample = multivariate_normal.rvs(mean_vector, cov_matrix)
                            noise_factor = np.random.uniform(noise_range[0], noise_range[1])
                            X_synthetic[idx] = x_orig * (1 - noise_factor) + random_sample * noise_factor

                        except (np.linalg.LinAlgError, ValueError):
                            noise_factor = np.random.uniform(noise_range[0], noise_range[1])
                            noise = np.random.normal(0, noise_factor * np.std(X_class, axis=0), n_features)
                            X_synthetic[idx] = x_orig + noise

                        y_synthetic[idx] = class_label
        else:
            for i in range(1, multiplier):
                for j, (x_orig, y_orig) in enumerate(zip(X, y)):
                    idx = n_samples * i + j
                    noise_factor = np.random.uniform(noise_range[0], noise_range[1])
                    X_synthetic[idx] = x_orig + np.random.normal(0, noise_factor * np.std(X, axis=0), n_features)
                    y_synthetic[idx] = y_orig

        return X_synthetic, y_synthetic


class ExtendedEnsembleValidator:
    """Extended ensemble algorithms validator"""

    def __init__(self, config):
        """Initialize validator"""
        self.config = config
        self.rf_lda_ensemble_module_path = config['rf_lda_ensemble_module_path']
        self.output_path = config['output_path']
        self.training_data_path = config['training_data_path']
        self.training_sheet_name = config['training_sheet_name']
        self.external_data_path = config['external_data_path']
        self.external_sheet_name = config['external_sheet_name']
        self.n_select = config.get('n_select', 5)
        self.fine_tune_fraction = config.get('fine_tune_fraction', 0.15)
        self.perturb_range = config.get('perturb_range', (0.1, 0.3))

        # Synthetic data parameters
        self.synthetic_multiplier = config.get('synthetic_multiplier', 5)
        self.synthetic_noise_range = config.get('synthetic_noise_range', (0.05, 0.15))
        self.cv_folds = config.get('cv_folds', 5)
        self.preserve_correlation = config.get('preserve_correlation', True)

        self.n_repeats = config.get('n_repeats', 5)

        os.makedirs(self.output_path, exist_ok=True)

        # Load original algorithms module
        self.load_original_model_module()
        self.save_config()

    def save_config(self):
        """Save configuration to JSON file"""
        config_path = os.path.join(self.output_path, "validation_config.json")
        try:
            serializable_config = {}
            for k, v in self.config.items():
                if isinstance(v, (dict, list, str, int, float, bool)) or v is None:
                    serializable_config[k] = v
                else:
                    serializable_config[k] = str(v)

            with open(config_path, 'w') as f:
                json.dump(serializable_config, f, indent=4)

        except Exception as e:
            logging.error(f"Error saving configuration: {str(e)}")

    def load_original_model_module(self):
        """Dynamically load original algorithms module"""
        try:
            module_dir = os.path.dirname(self.rf_lda_ensemble_module_path)
            if module_dir and module_dir not in sys.path:
                sys.path.append(module_dir)

            module_name = os.path.basename(self.rf_lda_ensemble_module_path)
            if module_name.endswith('.py'):
                module_name = module_name[:-3]

            spec = importlib.util.spec_from_file_location(module_name, self.rf_lda_ensemble_module_path)
            self.original_module = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(self.original_module)

            self.RFLDAEnsembleClassifier = getattr(self.original_module, 'RFLDAEnsembleClassifier')

        except Exception as e:
            logging.error(f"Error loading original algorithms module: {str(e)}")
            raise

    def train_ensemble_classifier(self):
        """Train and obtain cross-validation LOO results"""
        model_config = {
            'file_path': self.training_data_path,
            'sheet_name': self.training_sheet_name,
            'output_path': os.path.join(self.output_path, 'cv_results'),
            'outer_cv_splits': 5,
            'inner_cv_splits': 5,
            'optuna_config': {
                'n_trials': 80,
                'timeout': 300,
                'sampler': 'TPE',
                'pruner': 'Hyperband'
            }
        }

        os.makedirs(model_config['output_path'], exist_ok=True)

        print("Model training...")
        ensemble = self.RFLDAEnsembleClassifier(model_config)
        cv_results = ensemble.run()

        if not cv_results:
            raise ValueError("Cross-validation training failed, no LOO results returned")

        return cv_results

    def collect_all_params(self, cv_results):
        """Collect all parameter combinations from cross-validation"""
        params_list = []
        for result in cv_results:
            if 'rf_best_params' in result and 'lda_best_params' in result:
                params = {
                    'rf_params': result['rf_best_params'],
                    'lda_params': result['lda_best_params'],
                    'threshold': result.get('optimal_threshold', 0.5),
                    'lda_weight': result.get('optimal_lda_weight', 0.5)
                }
                params_list.append(params)

        if not params_list:
            logging.warning("Insufficient parameter information")
            return []

        return params_list

    def train_final_models(self, all_params):
        """Train multiple models on complete training set using all parameter combinations"""
        # Prepare training data
        df = pd.read_excel(self.training_data_path, sheet_name=self.training_sheet_name)

        if 'Sample' in df.columns:
            labels = df['Sample'].values
            features = df.drop('Sample', axis=1).values
        else:
            labels = df.iloc[:, 0].values
            features = df.iloc[:, 1:].values

        # Standardize features
        scaler = StandardScaler()
        X = scaler.fit_transform(features)

        # Create binary classification labels (LuC=1, N=0)
        y_binary = (labels == 'LuC').astype(int)

        # Train multiple models
        models = []
        for i, params in enumerate(all_params):
            rf_params = params['rf_params']
            lda_params = params['lda_params']
            threshold = params['threshold']
            lda_weight = params['lda_weight']

            # Create and train RF algorithms
            rf_model = self.create_rf_model(rf_params)
            rf_model.fit(X, y_binary)

            # Create and train LDA algorithms
            lda_model = self.create_lda_model(lda_params)
            lda_model.fit(X, y_binary)

            # Save algorithms and parameters
            models.append({
                'rf_model': rf_model,
                'lda_model': lda_model,
                'threshold': threshold,
                'lda_weight': lda_weight
            })

        final_models = {
            'models': models,
            'scaler': scaler
        }

        return final_models

    def create_rf_model(self, params):
        """Create RF algorithms"""
        from sklearn.ensemble import RandomForestClassifier

        model_params = {
            'n_estimators': int(params.get('n_estimators', 100)),
            'max_depth': params.get('max_depth'),
            'min_samples_split': max(2, int(params.get('min_samples_split', 2))),
            'min_samples_leaf': int(params.get('min_samples_leaf', 1)),
            'max_features': params.get('max_features', 'sqrt'),
            'class_weight': params.get('class_weight'),
            'random_state': 42,
            'n_jobs': -1
        }

        return RandomForestClassifier(**model_params)

    def create_lda_model(self, params):
        """Create LDA algorithms"""
        from sklearn.discriminant_analysis import LinearDiscriminantAnalysis

        solvers = ['svd', 'lsqr', 'eigen']
        solver_idx = int(params.get('solver_idx', 0))
        solver = solvers[solver_idx % len(solvers)]

        if solver == 'svd':
            return LinearDiscriminantAnalysis(solver=solver)
        else:
            return LinearDiscriminantAnalysis(
                solver=solver,
                shrinkage=params.get('shrinkage_factor', 0.5)
            )

    def model_fine_tune(self, all_params, random_seed=42):

        try:
            df_external = pd.read_excel(self.external_data_path, sheet_name=self.external_sheet_name)

            if 'Sample' in df_external.columns:
                external_labels = df_external['Sample'].values
                external_features = df_external.drop('Sample', axis=1).values
            else:
                external_labels = df_external.iloc[:, 0].values
                external_features = df_external.iloc[:, 1:].values

            X_tune, X_test, y_tune, y_test = train_test_split(
                external_features, external_labels,
                test_size=1 - self.fine_tune_fraction,
                stratify=external_labels,
                random_state=random_seed
            )

            complete_external_df = pd.DataFrame()
            complete_external_df['Sample'] = external_labels
            for i in range(external_features.shape[1]):
                complete_external_df[f'Feature_{i + 1}'] = external_features[:, i]

            # Create binary classification labels for fine-tuning
            y_tune_binary = (y_tune == 'LuC').astype(int)

            # Standardize features
            scaler = StandardScaler()
            X_tune_scaled = scaler.fit_transform(X_tune)

            # Generate synthetic data from fine-tuning subset
            # Set random seed for synthetic data generation
            np.random.seed(random_seed)
            X_synthetic, y_synthetic = SyntheticDataGenerator.generate_synthetic_samples(
                X_tune_scaled, y_tune_binary,
                multiplier=self.synthetic_multiplier,
                noise_range=self.synthetic_noise_range,
                preserve_correlation=self.preserve_correlation
            )

            # Fine-tune parameters using synthetic data
            fine_tuned_params_list = []
            fine_tuned_models_list = []

            for i, params in enumerate(all_params):
                rf_params = params['rf_params']
                lda_params = params['lda_params']
                threshold = params['threshold']
                lda_weight = params['lda_weight']

                # Use Optuna for parameter fine-tuning
                fine_tuned_params = self._optuna_fine_tune_params(
                    X_synthetic, y_synthetic, rf_params, lda_params, threshold, lda_weight, i, random_seed
                )

                # Train final algorithms
                final_rf_model = self.create_rf_model(fine_tuned_params['rf_params'])
                final_rf_model.fit(X_synthetic, y_synthetic)

                final_lda_model = self.create_lda_model(fine_tuned_params['lda_params'])
                final_lda_model.fit(X_synthetic, y_synthetic)

                fine_tuned_model = {
                    'rf_model': final_rf_model,
                    'lda_model': final_lda_model,
                    'threshold': fine_tuned_params['threshold'],
                    'lda_weight': fine_tuned_params['lda_weight']
                }

                fine_tuned_params_list.append(fine_tuned_params)
                fine_tuned_models_list.append(fine_tuned_model)

            fine_tuned_models = {
                'models': fine_tuned_models_list,
                'scaler': scaler
            }

            return fine_tuned_params_list, fine_tuned_models, complete_external_df

        except Exception as e:
            logging.error(f"Error during enhanced fine-tuning process: {str(e)}")
            import traceback
            logging.error(traceback.format_exc())
            raise

    def _optuna_fine_tune_params(self, X_synthetic, y_synthetic, rf_params, lda_params, threshold, lda_weight,
                                 param_idx, random_seed):
        """Use Optuna for parameter fine-tuning"""

        def objective(trial):
            # Suggest perturbations based on original parameters
            perturb_factor = 0.2

            # RF parameter perturbation
            n_estimators = rf_params.get('n_estimators', 100)
            n_est_min = max(50, int(n_estimators * (1 - perturb_factor)))
            n_est_max = int(n_estimators * (1 + perturb_factor))
            rf_n_estimators = trial.suggest_int('rf_n_estimators', n_est_min, n_est_max)

            max_depth = rf_params.get('max_depth')
            if max_depth is not None:
                depth_min = max(3, int(max_depth * (1 - perturb_factor)))
                depth_max = int(max_depth * (1 + perturb_factor))
                rf_max_depth = trial.suggest_int('rf_max_depth', depth_min, depth_max)
            else:
                rf_max_depth = None

            min_samples_split = rf_params.get('min_samples_split', 2)
            split_min = max(2, int(min_samples_split * (1 - perturb_factor)))
            split_max = int(min_samples_split * (1 + perturb_factor))
            rf_min_samples_split = trial.suggest_int('rf_min_samples_split', split_min, split_max)

            min_samples_leaf = rf_params.get('min_samples_leaf', 1)
            leaf_min = max(1, int(min_samples_leaf * (1 - perturb_factor)))
            leaf_max = int(min_samples_leaf * (1 + perturb_factor))
            rf_min_samples_leaf = trial.suggest_int('rf_min_samples_leaf', leaf_min, leaf_max)

            max_features = rf_params.get('max_features', 0.5)
            if isinstance(max_features, str):
                max_features = 0.5 if max_features == 'sqrt' else 0.3 if max_features == 'log2' else 0.7
            rf_max_features = trial.suggest_float('rf_max_features', max(0.1, max_features - 0.2),
                                                  min(1.0, max_features + 0.2))

            # LDA parameter perturbation
            solver_idx = lda_params.get('solver_idx', 0)
            lda_solver_idx = trial.suggest_int('lda_solver_idx', max(0, solver_idx - 1), min(2, solver_idx + 1))

            shrinkage_factor = lda_params.get('shrinkage_factor', 0.5)
            lda_shrinkage_factor = trial.suggest_float('lda_shrinkage_factor', max(0.0, shrinkage_factor - 0.3),
                                                       min(1.0, shrinkage_factor + 0.3))

            # Ensemble parameter perturbation
            ensemble_threshold = trial.suggest_float('threshold', max(0.1, threshold - 0.2), min(0.9, threshold + 0.2))
            ensemble_lda_weight = trial.suggest_float('lda_weight', max(0.0, lda_weight - 0.3),
                                                      min(1.0, lda_weight + 0.3))

            # Build parameter dictionary
            perturbed_rf_params = rf_params.copy()
            perturbed_rf_params.update({
                'n_estimators': rf_n_estimators,
                'max_depth': rf_max_depth,
                'min_samples_split': rf_min_samples_split,
                'min_samples_leaf': rf_min_samples_leaf,
                'max_features': rf_max_features
            })

            perturbed_lda_params = lda_params.copy()
            perturbed_lda_params.update({
                'solver_idx': lda_solver_idx,
                'shrinkage_factor': lda_shrinkage_factor
            })

            # Cross-validation evaluation
            cv = StratifiedKFold(n_splits=min(self.cv_folds, len(y_synthetic)), shuffle=True, random_state=random_seed)
            f1_scores = []

            for train_idx, test_idx in cv.split(X_synthetic, y_synthetic):
                X_fold_train, X_fold_test = X_synthetic[train_idx], X_synthetic[test_idx]
                y_fold_train, y_fold_test = y_synthetic[train_idx], y_synthetic[test_idx]

                # Train models
                rf_model = self.create_rf_model(perturbed_rf_params)
                rf_model.fit(X_fold_train, y_fold_train)

                lda_model = self.create_lda_model(perturbed_lda_params)
                lda_model.fit(X_fold_train, y_fold_train)

                # Predict
                rf_proba = rf_model.predict_proba(X_fold_test)[:, 1]
                lda_proba = lda_model.predict_proba(X_fold_test)[:, 1]

                # Ensemble prediction
                ensemble_proba = (1 - ensemble_lda_weight) * rf_proba + ensemble_lda_weight * lda_proba
                ensemble_pred = (ensemble_proba >= ensemble_threshold).astype(int)

                # Calculate F1 score
                f1 = f1_score(y_fold_test, ensemble_pred)
                f1_scores.append(f1)

            return np.mean(f1_scores)

        # Create Optuna study
        study = optuna.create_study(
            direction='maximize',
            sampler=TPESampler(seed=random_seed),
            pruner=HyperbandPruner()
        )

        # Optimize
        study.optimize(objective, n_trials=30, timeout=60, n_jobs=1)

        # Extract best parameters
        best_params = study.best_trial.params

        # Build returned parameter dictionary
        final_rf_params = rf_params.copy()
        final_rf_params.update({
            'n_estimators': best_params['rf_n_estimators'],
            'max_depth': best_params.get('rf_max_depth'),
            'min_samples_split': best_params['rf_min_samples_split'],
            'min_samples_leaf': best_params['rf_min_samples_leaf'],
            'max_features': best_params['rf_max_features']
        })

        final_lda_params = lda_params.copy()
        final_lda_params.update({
            'solver_idx': best_params['lda_solver_idx'],
            'shrinkage_factor': best_params['lda_shrinkage_factor']
        })

        return {
            'rf_params': final_rf_params,
            'lda_params': final_lda_params,
            'threshold': best_params['threshold'],
            'lda_weight': best_params['lda_weight']
        }

    def validate_external_data(self, fine_tuned_models, validation_df):
        """Validate external data using soft voting strategy"""
        try:
            # Extract features and labels
            if 'Sample' in validation_df.columns:
                external_labels = validation_df['Sample'].values
                external_features = validation_df.drop('Sample', axis=1).values
            else:
                external_labels = validation_df.iloc[:, 0].values
                external_features = validation_df.iloc[:, 1:].values

            # Get models and scaler
            model_list = fine_tuned_models['models']
            scaler = fine_tuned_models['scaler']

            # Standardize features
            X_external = scaler.transform(external_features)

            # Initialize probability vectors
            ensemble_probas = np.zeros((len(external_labels), len(model_list)))

            # Use each algorithms for prediction
            for i, model in enumerate(model_list):
                rf_model = model['rf_model']
                lda_model = model['lda_model']
                lda_weight = model['lda_weight']

                # RF algorithms prediction
                rf_proba = rf_model.predict_proba(X_external)[:, 1]

                # LDA algorithms prediction
                lda_proba = lda_model.predict_proba(X_external)[:, 1]

                # Calculate weighted ensemble prediction
                ensemble_proba = (1 - lda_weight) * rf_proba + lda_weight * lda_proba

                # Save prediction probability
                ensemble_probas[:, i] = ensemble_proba

            # Calculate average prediction probability across all models (soft voting)
            final_probabilities = np.mean(ensemble_probas, axis=1)

            # Create binary classification labels (LuC=1, N=0)
            y_true_binary = (external_labels == 'LuC').astype(int)

            # Use 0.5 threshold for prediction
            y_pred_binary = (final_probabilities >= 0.5).astype(int)

            # Calculate performance metrics for LuC class
            precision = precision_score(y_true_binary, y_pred_binary, zero_division=0)
            recall = recall_score(y_true_binary, y_pred_binary, zero_division=0)
            sensitivity = recall
            specificity = specificity_score(y_true_binary, y_pred_binary)
            f1 = f1_score(y_true_binary, y_pred_binary, zero_division=0)
            accuracy = accuracy_score(y_true_binary, y_pred_binary)

            # Calculate AUC
            if len(np.unique(y_true_binary)) > 1:
                auc_score = roc_auc_score(y_true_binary, final_probabilities)
                fpr, tpr, _ = roc_curve(y_true_binary, final_probabilities)
                roc_auc = auc(fpr, tpr)
            else:
                auc_score = 0.5
                fpr, tpr = np.array([0, 1]), np.array([0, 1])
                roc_auc = 0.5

            # Calculate confusion matrix
            cm = confusion_matrix(y_true_binary, y_pred_binary)

            # Build single run result
            single_result = {
                'accuracy': accuracy,
                'precision': precision,
                'recall': recall,
                'sensitivity': sensitivity,
                'specificity': specificity,
                'f1': f1,
                'auc': auc_score,
                'confusion_matrix': cm,
                'fpr': fpr,
                'tpr': tpr,
                'probabilities': final_probabilities,
                'true_labels': y_true_binary,
                'predicted_labels': y_pred_binary
            }

            return single_result

        except Exception as e:
            logging.error(f"Error validating external data: {str(e)}")
            import traceback
            logging.error(traceback.format_exc())
            raise

    def run_multiple_validation(self):

        print("=" * 50)
        print("External validation for early cancer detection")
        print("=" * 50)
        print()

        # 1. Execute cross-validation training
        cv_results = self.train_ensemble_classifier()

        # 2. Collect all parameter combinations
        all_params = self.collect_all_params(cv_results)

        if not all_params:
            raise ValueError("Unable to collect valid parameter combinations")

        # Display external data information
        df_external = pd.read_excel(self.external_data_path, sheet_name=self.external_sheet_name)
        if 'Sample' in df_external.columns:
            external_labels = df_external['Sample'].values
            external_features = df_external.drop('Sample', axis=1).values
        else:
            external_labels = df_external.iloc[:, 0].values
            external_features = df_external.iloc[:, 1:].values

        n_samples = len(external_labels)
        n_features = external_features.shape[1]
        n_luc = sum(external_labels == 'LuC')
        n_normal = sum(external_labels == 'N')
        print(f"\nExternal validation dataset: {n_samples} samples, {n_features} features [LuC: {n_luc}, N: {n_normal}]")

        # 3. Run fine-tuning and validation multiple times
        all_results = []
        all_roc_data = []

        print("External validation in progress")

        for run_idx in range(self.n_repeats):
            # Use different random seed for fine-tuning
            random_seed = 42 + run_idx

            # Fine-tune parameters and train models
            fine_tuned_params, fine_tuned_models, validation_df = self.model_fine_tune(
                all_params, random_seed)

            # Validate external data
            result = self.validate_external_data(fine_tuned_models, validation_df)

            all_results.append(result)
            all_roc_data.append({
                'fpr': result['fpr'],
                'tpr': result['tpr'],
                'auc': result['auc']
            })

        # 4. Calculate metrics
        mean_results = self._calculate_mean_results(all_results)

        # 5. Save LOO results
        self._save_multiple_run_results(all_results, mean_results, all_roc_data)

        print(f"\nValidation completed!")
        print(f"Results: Accuracy={mean_results['mean_accuracy']:.3f}±{mean_results['std_accuracy']:.3f}, "
              f"F1={mean_results['mean_f1']:.3f}±{mean_results['std_f1']:.3f}, "
              f"Sensitivity={mean_results['mean_sensitivity']:.3f}±{mean_results['std_sensitivity']:.3f}, "
              f"Specificity={mean_results['mean_specificity']:.3f}±{mean_results['std_specificity']:.3f}, "
              f"AUC={mean_results['mean_auc']:.3f}±{mean_results['std_auc']:.3f}")

        return all_results, mean_results

    def _calculate_mean_results(self, all_results):

        metrics = ['accuracy', 'precision', 'recall', 'sensitivity', 'specificity', 'f1', 'auc']
        mean_results = {}

        for metric in metrics:
            values = [result[metric] for result in all_results]
            mean_results[f'mean_{metric}'] = np.mean(values)
            mean_results[f'std_{metric}'] = np.std(values)

        # Calculate mean confusion matrix
        confusion_matrices = [result['confusion_matrix'] for result in all_results]
        mean_cm = np.mean(confusion_matrices, axis=0)
        mean_results['mean_confusion_matrix'] = np.round(mean_cm).astype(int)

        return mean_results

    def _save_multiple_run_results(self, all_results, mean_results, all_roc_data):

        # Save detailed LOO results to Excel
        results_path = os.path.join(self.output_path, 'multiple_validation_results.xlsx')

        with pd.ExcelWriter(results_path) as writer:
            # Save individual run LOO results
            run_data = []
            for i, result in enumerate(all_results):
                run_data.append({
                    'Run': i + 1,
                    'Accuracy': result['accuracy'],
                    'Precision': result['precision'],
                    'Recall': result['recall'],
                    'Sensitivity': result['sensitivity'],
                    'Specificity': result['specificity'],
                    'F1': result['f1'],
                    'AUC': result['auc']
                })

            pd.DataFrame(run_data).to_excel(writer, sheet_name='Individual_Runs', index=False)

            mean_data = [{
                'Metric': 'Mean ± Std',
                'Accuracy': f"{mean_results['mean_accuracy']:.3f} ± {mean_results['std_accuracy']:.3f}",
                'Precision': f"{mean_results['mean_precision']:.3f} ± {mean_results['std_precision']:.3f}",
                'Recall': f"{mean_results['mean_recall']:.3f} ± {mean_results['std_recall']:.3f}",
                'Sensitivity': f"{mean_results['mean_sensitivity']:.3f} ± {mean_results['std_sensitivity']:.3f}",
                'Specificity': f"{mean_results['mean_specificity']:.3f} ± {mean_results['std_specificity']:.3f}",
                'F1': f"{mean_results['mean_f1']:.3f} ± {mean_results['std_f1']:.3f}",
                'AUC': f"{mean_results['mean_auc']:.3f} ± {mean_results['std_auc']:.3f}"
            }]

            pd.DataFrame(mean_data).to_excel(writer, sheet_name='Mean_Results', index=False)

            cm_df = pd.DataFrame(mean_results['mean_confusion_matrix'],
                                index=['N', 'LuC'], columns=['N', 'LuC'])
            cm_df.to_excel(writer, sheet_name='Mean_Confusion_Matrix')

        # Format Excel file
        self._format_excel_file(results_path)

        # Plot visualization charts
        self._plot_mean_confusion_matrix(mean_results['mean_confusion_matrix'])
        self._plot_mean_roc_curve(all_roc_data, mean_results)

    def _format_excel_file(self, file_path):
        """Format Excel file with Times New Roman font and bold headers"""
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font

            wb = load_workbook(file_path)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # Set all cells to Times New Roman, 11pt
                for row in ws.iter_rows():
                    for cell in row:
                        cell.font = Font(name='Times New Roman', size=11)

                # Set first row to bold
                for cell in ws[1]:
                    cell.font = Font(name='Times New Roman', size=11, bold=True)

            wb.save(file_path)
        except Exception as e:
            logging.error(f"Error formatting Excel file: {str(e)}")

    def _plot_mean_confusion_matrix(self, mean_cm):
        """Plot confusion matrix"""
        plt.rcParams['svg.fonttype'] = 'none'
        plt.figure(figsize=(10, 8))
        plt.rcParams['font.family'] = 'Arial'

        ax = plt.gca()
        for spine in ax.spines.values():
            spine.set_linewidth(5)

        sns.heatmap(mean_cm, annot=True, fmt='d', cmap='Blues',
                    xticklabels=['N', 'LuC'], yticklabels=['N', 'LuC'],
                    annot_kws={'size': 20, 'weight': 'bold'})

        plt.title('Mean Confusion Matrix (LuC Detection)', fontsize=20, pad=20)
        plt.ylabel('True Label', fontsize=24, labelpad=15)
        plt.xlabel('Predicted Label', fontsize=24, labelpad=15)
        ax.tick_params(axis='both', labelsize=24)

        output_path = os.path.join(self.output_path, 'mean_confusion_matrix.svg')
        plt.savefig(output_path, format='svg', bbox_inches='tight')
        plt.close()

    def _plot_mean_roc_curve(self, all_roc_data, mean_results):
        """Plot ROC curve"""
        plt.rcParams['svg.fonttype'] = 'none'
        plt.figure(figsize=(10, 8))
        plt.rcParams['font.family'] = 'Arial'

        ax = plt.gca()
        for spine in ax.spines.values():
            spine.set_linewidth(7)

        plt.title("ROC Curve for LuC Detection", fontsize=20, pad=15)
        plt.xlabel("False Positive Rate", fontsize=24, labelpad=10)
        plt.ylabel("True Positive Rate", fontsize=24, labelpad=10)

        # Plot diagonal line
        plt.plot([0, 1], [0, 1], 'k--', lw=2)

        # Create interpolation points
        mean_fpr = np.linspace(0, 1, 1000)
        tprs = []
        aucs = []

        for i, roc_data in enumerate(all_roc_data):
            fpr, tpr, auc_score = roc_data['fpr'], roc_data['tpr'], roc_data['auc']

            # Ensure complete start and end points
            if fpr[0] != 0:
                fpr = np.concatenate([[0], fpr])
                tpr = np.concatenate([[0], tpr])
            if fpr[-1] != 1:
                fpr = np.concatenate([fpr, [1]])
                tpr = np.concatenate([tpr, [tpr[-1]]])

            # Interpolation
            interp_tpr = np.interp(mean_fpr, fpr, tpr)
            interp_tpr[0] = 0.0
            tprs.append(interp_tpr)
            aucs.append(auc_score)

            plt.plot(fpr, tpr, alpha=0.3, lw=1, color='gray')

        mean_tpr = np.mean(tprs, axis=0)
        mean_auc = mean_results['mean_auc']
        std_auc = mean_results['std_auc']

        plt.plot(mean_fpr, mean_tpr, color='blue', lw=5,
                 label=f'Mean ROC (AUC = {mean_auc:.3f} ± {std_auc:.3f})')

        std_tpr = np.std(tprs, axis=0)
        tprs_upper = np.minimum(mean_tpr + std_tpr, 1)
        tprs_lower = np.maximum(mean_tpr - std_tpr, 0)
        plt.fill_between(mean_fpr, tprs_lower, tprs_upper, color='blue', alpha=0.2)

        ax.tick_params(axis='both', labelsize=48, width=9, length=20)
        plt.xticks(np.arange(0, 1.1, 0.2))
        plt.yticks(np.arange(0, 1.1, 0.2))

        plt.grid(True, alpha=0.3)
        plt.xlim([-0.02, 1.02])
        plt.ylim([-0.02, 1.02])

        legend = plt.legend(loc="lower right", fontsize=28, frameon=True)
        legend.get_frame().set_linewidth(3)

        output_path = os.path.join(self.output_path, 'mean_roc_curve.svg')
        plt.savefig(output_path, format='svg', bbox_inches='tight')
        plt.close()


def main():
    """Main function"""
    config = {
        # CV algorithms module path
        'rf_lda_ensemble_module_path': os.path.join(os.path.dirname(os.path.abspath(__file__)), 'RF+LDA_Early cancer detection_CV.py'),  # Path to the RF+LDA ensemble classifier Python module

        # Output path
        'output_path': os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'Results', 'Early cancer detection', 'External validation'),  # Directory path for saving output results

        # Training data path
        'training_data_path': os.path.join(os.path.dirname(os.path.abspath(__file__)), 'TS Train.xlsx'),  # Path to the training data Excel file
        'training_sheet_name': 'Sheet1',  # Name of the sheet in training data file (e.g., 'Sheet1')

        # External validation data path
        'external_data_path': os.path.join(os.path.dirname(os.path.abspath(__file__)), 'TS Test.xlsx'),  # Path to the external validation data Excel file
        'external_sheet_name': 'Sheet1',  # Name of the sheet in external data file (e.g., 'Sheet1')

        # Number of most similar parameters to select
        'n_select': 5,

        # Enhanced synthetic data generation parameters
        'synthetic_multiplier': 5,
        'synthetic_noise_range': (0.05, 0.20),
        'cv_folds': 5,
        'preserve_correlation': True,

        # Fine-tuning parameters
        'fine_tune_fraction': 0.138,
        'perturb_range': (0.1, 0.3),

        # Number of repeated runs
        'n_repeats': 1
    }

    try:
        # Create timestamped directory
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        config['output_path'] = os.path.join(config['output_path'], timestamp)
        os.makedirs(config['output_path'], exist_ok=True)

        validator = ExtendedEnsembleValidator(config)
        all_results, mean_results = validator.run_multiple_validation()


    except Exception as e:
        logging.error(f"Main program error: {str(e)}")
        import traceback
        logging.error(traceback.format_exc())


if __name__ == "__main__":
    main()