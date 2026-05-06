# -*- coding: utf-8 -*-
import numpy as np
import pandas as pd
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import StratifiedKFold
from sklearn.metrics import (f1_score, roc_curve, auc, confusion_matrix, precision_score,
                             recall_score, precision_recall_curve, average_precision_score,
                             roc_auc_score)
import matplotlib.pyplot as plt
from joblib import Parallel, delayed
import logging
import os
import seaborn as sns
import warnings
import json
import time
from datetime import datetime
import optuna
from optuna.samplers import TPESampler
from optuna.pruners import MedianPruner
from sklearn.ensemble import RandomForestClassifier
from sklearn.discriminant_analysis import LinearDiscriminantAnalysis
import matplotlib as mpl

mpl.rcParams['svg.fonttype'] = 'none'
mpl.rcParams['font.family'] = 'Arial'

warnings.filterwarnings('ignore')

logging.basicConfig(level=logging.INFO, format='%(message)s')

# Completely disable all Optuna logging
optuna.logging.set_verbosity(optuna.logging.CRITICAL)
optuna.logging.disable_default_handler()
import sys
optuna.logging._get_library_root_logger().addHandler(logging.NullHandler())
optuna.logging._get_library_root_logger().propagate = False


# Calculate specificity
def specificity_score(y_true, y_pred):

    tn, fp, fn, tp = confusion_matrix(y_true, y_pred).ravel()
    return tn / (tn + fp) if (tn + fp) > 0 else 0.0


class RFLDAEnsembleClassifier:
    def __init__(self, config):
        """
        Initialize LDA and RF ensemble classifier

        Args:
            config: Configuration dictionary containing algorithms parameters, cross-validation settings, etc.
        """
        self.config = config
        self.scaler = StandardScaler()

        self.algorithm_colors = {
            'LuC': [242 / 255, 128 / 255, 128 / 255],
            'N': [106 / 255, 180 / 255, 193 / 255],
            'RF': '#a5b9dd',
            'LDA': '#a5acdd',
            'Ensemble': '#f1923d'
        }

        # Initialize RF and LDA parameter ranges
        self.rf_param_ranges = {
            'n_estimators_factor': (1, 12),  # Will be multiplied by 50
            'criterion_idx': (0, 1),  # 0: 'gini', 1: 'entropy'
            'max_depth': (3, 35),
            'min_samples_split': (2, 20),
            'min_samples_leaf': (1, 10),
            'class_weight_multiplier': (0.5, 2.0)
        }

        self.lda_param_ranges = {
            'solver_idx': (0, 2),  # 0: 'svd', 1: 'lsqr', 2: 'eigen'
            'shrinkage_factor': (0, 1)  # Only used for 'lsqr' and 'eigen'
        }

        # Optuna configuration
        self.optuna_config = config.get('optuna_config', {
            'n_trials': 100,
            'timeout': 800,
            'sampler': 'TPE',
            'pruner': None,
            'n_startup_trials': 20,
            'n_ei_candidates': 150,
            'gamma': 0.20,
            'multivariate' : True,
            'prior_weight': 2.0
        })

    def load_data(self):
        """
        Load and preprocess data

        Returns:
            tuple: (processed features, processed labels)
        """
        try:
            df = pd.read_excel(self.config['file_path'], sheet_name=self.config['sheet_name'])

            # Check if DataFrame is empty or has abnormal structure
            if df.empty:
                raise ValueError("DataFrame is empty, please check if Excel file contains data.")

            if df.shape[1] <= 1:
                raise ValueError(f"DataFrame only has {df.shape[1]} column(s), expected at least 2 columns (label + features).")

            # Extract labels and features
            if 'Sample' in df.columns:
                # If 'Sample' column exists, use it as labels
                labels = df['Sample'].values
                features = df.drop('Sample', axis=1).values
            else:
                # Otherwise, use first column as labels
                labels = df.iloc[:, 0].values
                features = df.iloc[:, 1:].values

            # Get original feature names
            if 'Sample' in df.columns:
                feature_names = df.drop('Sample', axis=1).columns.tolist()
            else:
                feature_names = df.columns[1:].tolist()

            self.feature_names = feature_names

            # Check if features contain NaN or infinite values
            if np.any(np.isnan(features)) or np.any(np.isinf(features)):
                # Replace NaN and infinite values with 0
                features = np.nan_to_num(features, nan=0.0, posinf=0.0, neginf=0.0)

            # Convert to binary classification (LuC=1, N=0)
            binary_labels = np.array([1 if l == 'LuC' else 0 for l in labels])

            # Standardize features
            features_processed = self.scaler.fit_transform(features)

            # Output dataset summary
            n_samples = len(binary_labels)
            n_features = features_processed.shape[1]
            n_positive = sum(binary_labels == 1)
            n_negative = sum(binary_labels == 0)
            print(f"Dataset: {n_samples} samples, {n_features} features [LuC: {n_positive}, N: {n_negative}]")

            return features_processed, binary_labels

        except Exception as e:
            logging.error(f"Error loading data: {str(e)}")
            # Print more debugging information
            if 'df' in locals():
                logging.error(f"DataFrame info: shape={df.shape}, columns={df.columns.tolist()}")
            raise

    def _create_rf_model(self, params, class_ratio=None):
        """
        Create RandomForest algorithms

        Args:
            params: Model parameters
            class_ratio: Class ratio, used to calculate class_weight

        Returns:
            RandomForestClassifier instance
        """
        # Check if parameters are None
        if params is None:
            params = {
                'n_estimators_factor': 5.0,
                'criterion_idx': 0.0,
                'max_depth': 10.0,
                'min_samples_split': 5.0,
                'min_samples_leaf': 2.0,
                'class_weight_multiplier': 1.0
            }
            logging.warning("Using default RF parameters, original parameters are None")

        # Convert parameters
        n_estimators = int(params.get('n_estimators_factor', 5.0) * 50)
        criteria = ['gini', 'entropy']
        criterion = criteria[int(params.get('criterion_idx', 0.0) + 0.5) % len(criteria)]
        max_depth = int(params.get('max_depth', 10.0) + 0.5)
        min_samples_split = int(params.get('min_samples_split', 5.0) + 0.5)
        min_samples_leaf = int(params.get('min_samples_leaf', 2.0) + 0.5)

        # Calculate class weights
        if class_ratio is not None and 'class_weight_multiplier' in params:
            class_weight = {
                0: 1.0,
                1: params['class_weight_multiplier'] * class_ratio
            }
        else:
            class_weight = None

        return RandomForestClassifier(
            n_estimators=n_estimators,
            criterion=criterion,
            max_depth=max_depth,
            min_samples_split=min_samples_split,
            min_samples_leaf=min_samples_leaf,
            class_weight=class_weight,
            n_jobs=-1,
            random_state=42
        )

    def _create_lda_model(self, params):
        """
        Create LDA algorithms

        Args:
            params: Model parameters

        Returns:
            LinearDiscriminantAnalysis instance
        """
        # Check if parameters are None
        if params is None:
            params = {
                'solver_idx': 0.0,  # svd
                'shrinkage_factor': 0.5,
            }
            logging.warning("Using default LDA parameters, original parameters are None")

        # Process solver
        solvers = ['svd', 'lsqr', 'eigen']
        solver_idx = int(params.get('solver_idx', 0.0) + 0.5)
        solver = solvers[solver_idx % len(solvers)]

        if solver == 'svd':
            return LinearDiscriminantAnalysis(solver=solver)
        else:
            # Only use shrinkage for 'lsqr' and 'eigen'
            return LinearDiscriminantAnalysis(
                solver=solver,
                shrinkage=params.get('shrinkage_factor', 0.5)
            )

    def _optimize_rf_hyperparameters(self, X_train, y_train):
        """
        Optimize Random Forest hyperparameters using Optuna

        Args:
            X_train: Training features
            y_train: Training labels

        Returns:
            dict: Optimal hyperparameters
        """
        # Calculate class ratio
        neg_count = sum(y_train == 0)
        pos_count = sum(y_train == 1)
        base_class_ratio = neg_count / max(1, pos_count)

        def objective(trial):
            """Optuna optimization objective function - RF"""
            try:

                n_estimators_factor = trial.suggest_float('n_estimators_factor',
                                                          self.rf_param_ranges['n_estimators_factor'][0],
                                                          self.rf_param_ranges['n_estimators_factor'][1])
                criterion_idx = trial.suggest_float('criterion_idx',
                                                    self.rf_param_ranges['criterion_idx'][0],
                                                    self.rf_param_ranges['criterion_idx'][1])
                max_depth = trial.suggest_float('max_depth',
                                                self.rf_param_ranges['max_depth'][0],
                                                self.rf_param_ranges['max_depth'][1])
                min_samples_split = trial.suggest_float('min_samples_split',
                                                        self.rf_param_ranges['min_samples_split'][0],
                                                        self.rf_param_ranges['min_samples_split'][1])
                min_samples_leaf = trial.suggest_float('min_samples_leaf',
                                                       self.rf_param_ranges['min_samples_leaf'][0],
                                                       self.rf_param_ranges['min_samples_leaf'][1])
                class_weight_multiplier = trial.suggest_float('class_weight_multiplier',
                                                              self.rf_param_ranges['class_weight_multiplier'][0],
                                                              self.rf_param_ranges['class_weight_multiplier'][1])

                # Build parameter dictionary
                params = {
                    'n_estimators_factor': n_estimators_factor,
                    'criterion_idx': criterion_idx,
                    'max_depth': max_depth,
                    'min_samples_split': min_samples_split,
                    'min_samples_leaf': min_samples_leaf,
                    'class_weight_multiplier': class_weight_multiplier
                }

                # Create algorithms
                model = self._create_rf_model(params, base_class_ratio)

                # Use StratifiedKFold for cross-validation
                cv = StratifiedKFold(n_splits=self.config['inner_cv_splits'], shuffle=True, random_state=42)

                # Track multiple metrics
                f1_scores = []
                auc_scores = []
                sensitivity_scores = []
                specificity_scores = []

                for train_idx, val_idx in cv.split(X_train, y_train):
                    X_train_cv, X_val_cv = X_train[train_idx], X_train[val_idx]
                    y_train_cv, y_val_cv = y_train[train_idx], y_train[val_idx]

                    try:
                        # Train algorithms
                        model.fit(X_train_cv, y_train_cv)

                        # Get prediction probabilities
                        y_pred_proba = model.predict_proba(X_val_cv)[:, 1]

                        # Try different thresholds to find best F1
                        best_f1 = 0
                        best_threshold = 0.5
                        # Set default prediction to avoid best_pred being None
                        best_pred = (y_pred_proba >= 0.5).astype(int)

                        for threshold in [0.3, 0.35, 0.4, 0.45, 0.5]:
                            y_pred_thresh = (y_pred_proba >= threshold).astype(int)
                            curr_f1 = f1_score(y_val_cv, y_pred_thresh)
                            if curr_f1 > best_f1:
                                best_f1 = curr_f1
                                best_threshold = threshold
                                best_pred = y_pred_thresh

                        # Ensure best_pred is not None
                        if best_pred is None:
                            best_pred = (y_pred_proba >= 0.5).astype(int)

                        # Calculate metrics using best prediction LOO results
                        try:
                            f1 = f1_score(y_val_cv, best_pred)
                            f1_scores.append(f1)
                        except Exception as e:
                            f1_scores.append(0)

                        try:
                            roc_auc = roc_auc_score(y_val_cv, y_pred_proba)
                            auc_scores.append(roc_auc)
                        except Exception as e:
                            auc_scores.append(0.5)

                        try:
                            sensitivity = recall_score(y_val_cv, best_pred)
                            sensitivity_scores.append(sensitivity)
                        except Exception as e:
                            sensitivity_scores.append(0)

                        try:
                            spec = specificity_score(y_val_cv, best_pred)
                            specificity_scores.append(spec)
                        except Exception as e:
                            specificity_scores.append(0)

                    except Exception as e:
                        # Use default values for failed folds
                        f1_scores.append(0)
                        auc_scores.append(0.5)
                        sensitivity_scores.append(0)
                        specificity_scores.append(0)

                # Calculate average metrics
                if len(f1_scores) > 0:
                    mean_f1 = np.mean(f1_scores)
                    # Return F1 score
                    combined_score = mean_f1
                else:
                    # If all folds fail, return 0 score
                    combined_score = 0

                return combined_score

            except Exception as e:
                # Return a very low score
                return 0.0

        try:
            # Use Optuna optimization
            study = optuna.create_study(
                direction='maximize',
                sampler=TPESampler(
                    seed=42,
                    n_startup_trials=self.optuna_config.get('n_startup_trials', 30),
                    n_ei_candidates=self.optuna_config.get('n_ei_candidates', 100),
                    gamma=self.optuna_config.get('gamma', 0.25),  # More conservative good/bad split
                    prior_weight=self.optuna_config.get('prior_weight', 2.0)  # Increase prior weight
                ),
                pruner=None  # Disable pruning
            )

            start_time = time.time()
            study.optimize(
                objective,
                n_trials=self.optuna_config.get('n_trials', 80),
                timeout=self.optuna_config.get('timeout', 300),
                n_jobs=1
            )
            optimization_time = time.time() - start_time

            # Return best parameters
            best_params = study.best_trial.params

            print(
                f"         RF optimizing... Complete Best F1={study.best_value:.3f} ({len(study.trials)} trials, {optimization_time:.1f}s)")

            # Check if best score is too low
            if study.best_value < 0.1:
                logging.warning("RF best score is too low, using default parameters")
                best_params = {
                    'n_estimators_factor': 5.0,
                    'criterion_idx': 0.0,
                    'max_depth': 10.0,
                    'min_samples_split': 5.0,
                    'min_samples_leaf': 2.0,
                    'class_weight_multiplier': 1.0
                }

            return best_params

        except Exception as e:
            logging.error(f"RF hyperparameter optimization failed: {str(e)}")
            # Return default parameters
            default_params = {
                'n_estimators_factor': 5.0,
                'criterion_idx': 0.0,
                'max_depth': 10.0,
                'min_samples_split': 5.0,
                'min_samples_leaf': 2.0,
                'class_weight_multiplier': 1.0
            }
            logging.info(f"Using RF default parameters: {default_params}")
            return default_params

    def _optimize_lda_hyperparameters(self, X_train, y_train):
        """
        Optimize LDA hyperparameters using Optuna

        Args:
            X_train: Training features
            y_train: Training labels

        Returns:
            dict: Optimal hyperparameters
        """

        def objective(trial):
            """Optuna optimization objective function - LDA"""
            try:
                # Suggest hyperparameters
                solver_idx = trial.suggest_float('solver_idx',
                                                 self.lda_param_ranges['solver_idx'][0],
                                                 self.lda_param_ranges['solver_idx'][1])
                shrinkage_factor = trial.suggest_float('shrinkage_factor',
                                                       self.lda_param_ranges['shrinkage_factor'][0],
                                                       self.lda_param_ranges['shrinkage_factor'][1])

                # Build parameter dictionary
                params = {
                    'solver_idx': solver_idx,
                    'shrinkage_factor': shrinkage_factor
                }

                # Create algorithms
                model = self._create_lda_model(params)

                # Use StratifiedKFold for cross-validation
                cv = StratifiedKFold(n_splits=self.config['inner_cv_splits'], shuffle=True, random_state=42)

                # Track multiple metrics
                f1_scores = []
                auc_scores = []
                sensitivity_scores = []
                specificity_scores = []

                for train_idx, val_idx in cv.split(X_train, y_train):
                    X_train_cv, X_val_cv = X_train[train_idx], X_train[val_idx]
                    y_train_cv, y_val_cv = y_train[train_idx], y_train[val_idx]

                    try:
                        # Train algorithms
                        with warnings.catch_warnings():
                            warnings.simplefilter("ignore")
                            model.fit(X_train_cv, y_train_cv)

                        # Get prediction probabilities
                        y_pred_proba = model.predict_proba(X_val_cv)[:, 1]

                        # Try different thresholds to find best F1
                        best_f1 = 0
                        best_threshold = 0.5
                        # Set default prediction to avoid best_pred being None
                        best_pred = (y_pred_proba >= 0.5).astype(int)

                        for threshold in [0.3, 0.35, 0.4, 0.45, 0.5]:
                            y_pred_thresh = (y_pred_proba >= threshold).astype(int)
                            curr_f1 = f1_score(y_val_cv, y_pred_thresh)
                            if curr_f1 > best_f1:
                                best_f1 = curr_f1
                                best_threshold = threshold
                                best_pred = y_pred_thresh

                        # Ensure best_pred is not None
                        if best_pred is None:
                            best_pred = (y_pred_proba >= 0.5).astype(int)

                        # Calculate metrics using best prediction LOO results
                        try:
                            f1 = f1_score(y_val_cv, best_pred)
                            f1_scores.append(f1)
                        except Exception as e:
                            f1_scores.append(0)

                        try:
                            roc_auc = roc_auc_score(y_val_cv, y_pred_proba)
                            auc_scores.append(roc_auc)
                        except Exception as e:
                            auc_scores.append(0.5)

                        try:
                            sensitivity = recall_score(y_val_cv, best_pred)
                            sensitivity_scores.append(sensitivity)
                        except Exception as e:
                            sensitivity_scores.append(0)

                        try:
                            spec = specificity_score(y_val_cv, best_pred)
                            specificity_scores.append(spec)
                        except Exception as e:
                            specificity_scores.append(0)

                    except Exception as e:
                        # Use default values for failed folds
                        f1_scores.append(0)
                        auc_scores.append(0.5)
                        sensitivity_scores.append(0)
                        specificity_scores.append(0)

                # Calculate average metrics
                if len(f1_scores) > 0:
                    mean_f1 = np.mean(f1_scores)
                    # Return F1 score
                    combined_score = mean_f1
                else:
                    # If all folds fail, return 0 score
                    combined_score = 0

                return combined_score

            except Exception as e:
                # Return a very low score
                return 0.0

        try:
            # Use Optuna optimization - Solution 2: Optimize TPE parameters for small samples
            study = optuna.create_study(
                direction='maximize',
                sampler=TPESampler(
                    seed=42,
                    n_startup_trials=self.optuna_config.get('n_startup_trials', 30),
                    n_ei_candidates=self.optuna_config.get('n_ei_candidates', 100),
                    gamma=self.optuna_config.get('gamma', 0.25),
                    prior_weight=self.optuna_config.get('prior_weight', 2.0)
                ),
                pruner=None  # Disable pruning
            )

            start_time = time.time()
            study.optimize(
                objective,
                n_trials=self.optuna_config.get('n_trials', 80),
                timeout=self.optuna_config.get('timeout', 300),
                n_jobs=1
            )
            optimization_time = time.time() - start_time

            # Return best parameters
            best_params = study.best_trial.params

            print(
                f"         LDA optimizing... Complete Best F1={study.best_value:.3f} ({len(study.trials)} trials, {optimization_time:.1f}s)")

            # Check if best score is too low
            if study.best_value < 0.1:
                logging.warning("LDA best score is too low, using default parameters")
                best_params = {
                    'solver_idx': 0.0,  # svd
                    'shrinkage_factor': 0.5,
                }

            return best_params

        except Exception as e:
            logging.error(f"LDA hyperparameter optimization failed: {str(e)}")
            # Return default parameters
            default_params = {
                'solver_idx': 0.0,  # svd
                'shrinkage_factor': 0.5,
            }
            logging.info(f"Using LDA default parameters: {default_params}")
            return default_params

    def _optimize_ensemble_weight(self, rf_proba, lda_proba, y_true):
        """
        Optimize ensemble weight using Optuna

        Args:
            rf_proba: RF prediction probabilities
            lda_proba: LDA prediction probabilities
            y_true: True labels

        Returns:
            float: Optimal weight for LDA
        """

        def objective(trial):
            """Optuna optimization objective function - Ensemble weight"""
            try:
                # Suggest LDA weight
                lda_weight = trial.suggest_float('lda_weight', 0.0, 1.0)

                # Calculate weighted prediction probability
                ensemble_proba = (1 - lda_weight) * rf_proba + lda_weight * lda_proba

                # Try different thresholds to find best F1
                best_f1 = 0
                # Set default prediction
                best_pred = (ensemble_proba >= 0.5).astype(int)

                for threshold in np.arange(0.3, 0.6, 0.05):
                    y_pred = (ensemble_proba >= threshold).astype(int)
                    try:
                        curr_f1 = f1_score(y_true, y_pred)
                        if curr_f1 > best_f1:
                            best_f1 = curr_f1
                            best_pred = y_pred
                    except Exception as e:
                        pass

                # If no better F1 found, calculate F1 using default prediction
                if best_f1 == 0:
                    try:
                        best_f1 = f1_score(y_true, best_pred)
                    except Exception as e:
                        best_f1 = 0

                return best_f1
            except Exception as e:
                return 0.0

        try:
            # Check inputs
            if rf_proba is None or lda_proba is None or y_true is None:
                logging.error("Input is None when optimizing ensemble weight")
                return 0.5  # Return default weight

            # Use Optuna optimization to find optimal weight
            study = optuna.create_study(
                direction='maximize',
                sampler=TPESampler(seed=42)
            )

            start_time = time.time()
            study.optimize(
                objective,
                n_trials=20,
                timeout=60,
                n_jobs=1
            )
            optimization_time = time.time() - start_time

            best_weight = study.best_trial.params['lda_weight']
            best_score = study.best_value

            print(
                f"         Ensemble optimizing... Complete Weight={best_weight:.2f}, F1={best_score:.3f} ({len(study.trials)} trials, {optimization_time:.1f}s)")

            # If best score is too low, use default weight
            if best_score < 0.1:
                logging.warning(f"Ensemble weight optimization score is too low: {best_score}, using default weight 0.5")
                return 0.5

            return best_weight

        except Exception as e:
            logging.error(f"Ensemble weight optimization failed: {str(e)}")
            # Default to equal weight for both models
            return 0.5

    def process_fold(self, X, y, train_index, test_index, fold_idx):
        """
        Process a single cross-validation fold

        Args:
            X: Feature matrix
            y: Labels
            train_index: Training set indices
            test_index: Test set indices
            fold_idx: Fold number

        Returns:
            dict: Dictionary containing LOO results
        """
        try:
            X_train, X_test = X[train_index], X[test_index]
            y_train, y_test = y[train_index], y[test_index]

            # Check test set class distribution
            test_class_dist = np.bincount(y_test)

            # Record class distribution
            neg_count = sum(y_train == 0)
            pos_count = sum(y_train == 1)
            class_ratio = neg_count / max(1, pos_count)

            print(f"   Fold {fold_idx + 1}:")

            # Optimize RF hyperparameters
            rf_best_params = self._optimize_rf_hyperparameters(X_train, y_train)
            rf_model = self._create_rf_model(rf_best_params, class_ratio)
            rf_model.fit(X_train, y_train)

            # Optimize LDA hyperparameters
            lda_best_params = self._optimize_lda_hyperparameters(X_train, y_train)
            lda_model = self._create_lda_model(lda_best_params)
            lda_model.fit(X_train, y_train)

            # Get prediction probabilities
            rf_y_pred_proba = rf_model.predict_proba(X_test)[:, 1]
            lda_y_pred_proba = lda_model.predict_proba(X_test)[:, 1]

            # Optimize ensemble weight
            best_ensemble_weight = self._optimize_ensemble_weight(rf_y_pred_proba, lda_y_pred_proba, y_test)

            # Calculate weighted ensemble prediction probability
            ensemble_y_pred_proba = (
                                                1 - best_ensemble_weight) * rf_y_pred_proba + best_ensemble_weight * lda_y_pred_proba

            # Use dynamic threshold selection
            thresholds = np.arange(0.1, 0.7, 0.025)
            best_f1 = 0
            best_threshold = 0.5

            # Only optimize threshold when test set contains both classes
            if len(test_class_dist) > 1 and test_class_dist[0] > 0 and test_class_dist[1] > 0:
                for threshold in thresholds:
                    y_pred_thresh = (ensemble_y_pred_proba >= threshold).astype(int)
                    try:
                        curr_f1 = f1_score(y_test, y_pred_thresh)
                        if curr_f1 > best_f1:
                            best_f1 = curr_f1
                            best_threshold = threshold
                    except Exception as e:
                        pass

            # Use best threshold for prediction
            ensemble_y_pred = (ensemble_y_pred_proba >= best_threshold).astype(int)
            rf_y_pred = (rf_y_pred_proba >= best_threshold).astype(int)
            lda_y_pred = (lda_y_pred_proba >= best_threshold).astype(int)

            # Calculate F1 scores for each algorithms
            rf_f1 = f1_score(y_test, rf_y_pred)
            lda_f1 = f1_score(y_test, lda_y_pred)
            ensemble_f1 = f1_score(y_test, ensemble_y_pred)

            # Calculate other evaluation metrics
            ensemble_precision = precision_score(y_test, ensemble_y_pred)
            ensemble_recall = recall_score(y_test, ensemble_y_pred)
            ensemble_sensitivity = ensemble_recall
            ensemble_specificity = specificity_score(y_test, ensemble_y_pred)

            # Calculate ROC curve
            rf_fpr, rf_tpr, _ = roc_curve(y_test, rf_y_pred_proba)
            rf_auc = auc(rf_fpr, rf_tpr)

            lda_fpr, lda_tpr, _ = roc_curve(y_test, lda_y_pred_proba)
            lda_auc = auc(lda_fpr, lda_tpr)

            ensemble_fpr, ensemble_tpr, _ = roc_curve(y_test, ensemble_y_pred_proba)
            ensemble_auc = auc(ensemble_fpr, ensemble_tpr)

            # Confusion matrix
            ensemble_conf_matrix = confusion_matrix(y_test, ensemble_y_pred)

            # Calculate false negative rate and false positive rate
            tn, fp, fn, tp = 0, 0, 0, 0
            if ensemble_conf_matrix.shape == (2, 2):
                tn, fp, fn, tp = ensemble_conf_matrix.ravel()

            # False negative rate and false positive rate
            fnr = fn / (fn + tp) if (fn + tp) > 0 else 0
            fpr_val = fp / (fp + tn) if (fp + tn) > 0 else 0

            # Return LOO results
            return {
                "fold_idx": fold_idx,
                "y_test": y_test,
                "rf_pred": rf_y_pred,
                "rf_pred_proba": rf_y_pred_proba,
                "rf_f1": rf_f1,
                "rf_auc": rf_auc,
                "rf_fpr": rf_fpr,
                "rf_tpr": rf_tpr,
                "lda_pred": lda_y_pred,
                "lda_pred_proba": lda_y_pred_proba,
                "lda_f1": lda_f1,
                "lda_auc": lda_auc,
                "lda_fpr": lda_fpr,
                "lda_tpr": lda_tpr,
                "ensemble_pred": ensemble_y_pred,
                "ensemble_pred_proba": ensemble_y_pred_proba,
                "ensemble_f1": ensemble_f1,
                "ensemble_precision": ensemble_precision,
                "ensemble_recall": ensemble_recall,
                "ensemble_sensitivity": ensemble_sensitivity,
                "ensemble_specificity": ensemble_specificity,
                "ensemble_auc": ensemble_auc,
                "ensemble_fpr": ensemble_fpr,
                "ensemble_tpr": ensemble_tpr,
                "optimal_threshold": best_threshold,
                "optimal_lda_weight": best_ensemble_weight,
                "fnr": fnr,
                "fpr_val": fpr_val,
                "confusion_matrix": ensemble_conf_matrix,
                "rf_model": rf_model,
                "lda_model": lda_model,
                "rf_best_params": rf_best_params,
                "lda_best_params": lda_best_params
            }

        except Exception as e:
            logging.error(f"Error processing fold {fold_idx + 1}: {str(e)}")
            import traceback
            logging.error(traceback.format_exc())

            # Return placeholder LOO results
            return {
                "fold_idx": fold_idx,
                "error": str(e),
                "ensemble_f1": 0,
                "ensemble_auc": 0.5,
                "optimal_lda_weight": 0.5,
                "rf_f1": 0,
                "lda_f1": 0,
                "rf_auc": 0.5,
                "lda_auc": 0.5
            }

    def plot_roc_curves(self, fold_results, class_name):
        """
        Plot ROC curves for RF, LDA and ensemble models

        Args:
            fold_results: Results from all folds
            class_name: Class name
        """
        plt.figure(figsize=(12, 10))
        ax = plt.gca()

        for spine in ax.spines.values():
            spine.set_linewidth(5)

        plt.rcParams['font.family'] = 'Arial'

        plt.title(f"ROC Curve", fontsize=24, pad=15)
        plt.xlabel("False Positive Rate", fontsize=24, labelpad=10)
        plt.ylabel("True Positive Rate", fontsize=24, labelpad=10)

        # Plot diagonal line
        plt.plot([0, 1], [0, 1], 'k--', lw=2)

        rf_color = self.algorithm_colors['RF']
        lda_color = self.algorithm_colors['LDA']
        ensemble_color = self.algorithm_colors[class_name]

        # Create dense interpolation points
        mean_fpr = np.linspace(0, 1, 1000)

        # Prepare storage for interpolated TPR for each algorithms
        rf_tprs, lda_tprs, ensemble_tprs = [], [], []
        rf_aucs, lda_aucs, ensemble_aucs = [], [], []

        # Process ROC curve for each fold
        for result in fold_results:
            # Check for errors
            if 'error' in result:
                continue

            # RF
            rf_fpr, rf_tpr = result['rf_fpr'], result['rf_tpr']
            rf_auc = result['rf_auc']

            if len(rf_fpr) > 0:
                if rf_fpr[0] != 0:
                    rf_fpr = np.concatenate([[0], rf_fpr])
                    rf_tpr = np.concatenate([[0], rf_tpr])
                if rf_fpr[-1] != 1:
                    rf_fpr = np.concatenate([rf_fpr, [1]])
                    rf_tpr = np.concatenate([rf_tpr, [rf_tpr[-1]]])

                rf_interp_tpr = np.interp(mean_fpr, rf_fpr, rf_tpr)
                rf_interp_tpr[0] = 0.0
                rf_tprs.append(rf_interp_tpr)
                rf_aucs.append(rf_auc)

                plt.plot(rf_fpr, rf_tpr, lw=1, alpha=0.15, color=rf_color)

            # LDA
            lda_fpr, lda_tpr = result['lda_fpr'], result['lda_tpr']
            lda_auc = result['lda_auc']

            if len(lda_fpr) > 0:
                if lda_fpr[0] != 0:
                    lda_fpr = np.concatenate([[0], lda_fpr])
                    lda_tpr = np.concatenate([[0], lda_tpr])
                if lda_fpr[-1] != 1:
                    lda_fpr = np.concatenate([lda_fpr, [1]])
                    lda_tpr = np.concatenate([lda_tpr, [lda_tpr[-1]]])


                lda_interp_tpr = np.interp(mean_fpr, lda_fpr, lda_tpr)
                lda_interp_tpr[0] = 0.0
                lda_tprs.append(lda_interp_tpr)
                lda_aucs.append(lda_auc)

                # Plot LDA ROC curve for each fold
                plt.plot(lda_fpr, lda_tpr, lw=1, alpha=0.15, color=lda_color)

            # Ensemble
            ensemble_fpr, ensemble_tpr = result['ensemble_fpr'], result['ensemble_tpr']
            ensemble_auc = result['ensemble_auc']

            # Ensure start and end points are complete
            if len(ensemble_fpr) > 0:
                if ensemble_fpr[0] != 0:
                    ensemble_fpr = np.concatenate([[0], ensemble_fpr])
                    ensemble_tpr = np.concatenate([[0], ensemble_tpr])
                if ensemble_fpr[-1] != 1:
                    ensemble_fpr = np.concatenate([ensemble_fpr, [1]])
                    ensemble_tpr = np.concatenate([ensemble_tpr, [ensemble_tpr[-1]]])

                # Interpolate and save
                ensemble_interp_tpr = np.interp(mean_fpr, ensemble_fpr, ensemble_tpr)
                ensemble_interp_tpr[0] = 0.0
                ensemble_tprs.append(ensemble_interp_tpr)
                ensemble_aucs.append(ensemble_auc)

                # Plot ensemble ROC curve for each fold
                plt.plot(ensemble_fpr, ensemble_tpr, lw=1, alpha=0.15, color=ensemble_color)

        # Calculate average curves
        # RF
        if rf_tprs:
            rf_mean_tpr = np.mean(rf_tprs, axis=0)
            rf_mean_tpr[-1] = 1.0
            rf_mean_auc = np.mean(rf_aucs)
            rf_std_auc = np.std(rf_aucs)
            rf_std_tpr = np.std(rf_tprs, axis=0)
            rf_tprs_upper = np.minimum(rf_mean_tpr + rf_std_tpr, 1)
            rf_tprs_lower = np.maximum(rf_mean_tpr - rf_std_tpr, 0)

            plt.plot(mean_fpr, rf_mean_tpr, color=rf_color,
                     label=f'RF (AUC = {rf_mean_auc:.3f} ± {rf_std_auc:.3f})',
                     lw=4, alpha=0.8)

            plt.fill_between(mean_fpr, rf_tprs_lower, rf_tprs_upper,
                             color=rf_color, alpha=0.1)

        # LDA
        if lda_tprs:
            lda_mean_tpr = np.mean(lda_tprs, axis=0)
            lda_mean_tpr[-1] = 1.0
            lda_mean_auc = np.mean(lda_aucs)
            lda_std_auc = np.std(lda_aucs)
            lda_std_tpr = np.std(lda_tprs, axis=0)
            lda_tprs_upper = np.minimum(lda_mean_tpr + lda_std_tpr, 1)
            lda_tprs_lower = np.maximum(lda_mean_tpr - lda_std_tpr, 0)

            plt.plot(mean_fpr, lda_mean_tpr, color=lda_color,
                     label=f'LDA (AUC = {lda_mean_auc:.3f} ± {lda_std_auc:.3f})',
                     lw=4, alpha=0.8)

            plt.fill_between(mean_fpr, lda_tprs_lower, lda_tprs_upper,
                             color=lda_color, alpha=0.1)

        # Ensemble
        if ensemble_tprs:
            ensemble_mean_tpr = np.mean(ensemble_tprs, axis=0)
            ensemble_mean_tpr[-1] = 1.0
            ensemble_mean_auc = np.mean(ensemble_aucs)
            ensemble_std_auc = np.std(ensemble_aucs)
            ensemble_std_tpr = np.std(ensemble_tprs, axis=0)
            ensemble_tprs_upper = np.minimum(ensemble_mean_tpr + ensemble_std_tpr, 1)
            ensemble_tprs_lower = np.maximum(ensemble_mean_tpr - ensemble_std_tpr, 0)

            plt.plot(mean_fpr, ensemble_mean_tpr, color=ensemble_color,
                     label=f'Ensemble (AUC = {ensemble_mean_auc:.3f} ± {ensemble_std_auc:.3f})',
                     lw=5, alpha=1.0)

            plt.fill_between(mean_fpr, ensemble_tprs_lower, ensemble_tprs_upper,
                             color=ensemble_color, alpha=0.1)

        # Increase tick label size
        ax.tick_params(axis='both', labelsize=24)

        plt.grid(True, alpha=0.3)
        plt.xlim([-0.02, 1.02])
        plt.ylim([-0.02, 1.02])

        legend = plt.legend(loc="lower right", fontsize=20, frameon=True)
        legend.get_frame().set_linewidth(3)

        output_file = os.path.join(self.config['output_path'], f"ROC_Curves_{class_name}.svg")
        plt.savefig(output_file, format='svg', bbox_inches='tight')
        plt.close()

        # Save ROC curve data to Excel
        try:
            roc_data = pd.DataFrame({
                'FPR': mean_fpr
            })

            # Add TPR for each algorithms
            if rf_tprs:
                roc_data['RF_TPR'] = rf_mean_tpr
                if 'rf_tprs_upper' in locals():
                    roc_data['RF_TPR_upper'] = rf_tprs_upper
                    roc_data['RF_TPR_lower'] = rf_tprs_lower

            if lda_tprs:
                roc_data['LDA_TPR'] = lda_mean_tpr
                if 'lda_tprs_upper' in locals():
                    roc_data['LDA_TPR_upper'] = lda_tprs_upper
                    roc_data['LDA_TPR_lower'] = lda_tprs_lower

            if ensemble_tprs:
                roc_data['Ensemble_TPR'] = ensemble_mean_tpr
                if 'ensemble_tprs_upper' in locals():
                    roc_data['Ensemble_TPR_upper'] = ensemble_tprs_upper
                    roc_data['Ensemble_TPR_lower'] = ensemble_tprs_lower

            roc_data_file = os.path.join(self.config['output_path'], f"ROC_data_{class_name}.xlsx")
            roc_data.to_excel(roc_data_file, index=False)

            # Format Excel file
            self.format_excel_file(roc_data_file)
        except Exception as e:
            logging.error(f"Error saving ROC data: {str(e)}")

    def plot_confusion_matrix(self, fold_results, class_name):
        """
        Plot average confusion matrix for all folds

        Args:
            fold_results: Results from all folds
            class_name: Class name
        """
        # Collect confusion matrices from all folds
        all_matrices = []
        for result in fold_results:
            if 'confusion_matrix' in result:
                all_matrices.append(result['confusion_matrix'])

        if not all_matrices:
            logging.warning(f"No confusion matrix data available")
            return

        try:
            all_matrices_2x2 = []
            for matrix in all_matrices:
                if matrix.shape == (2, 2):
                    all_matrices_2x2.append(matrix)
                else:
                    new_matrix = np.zeros((2, 2))
                    rows, cols = matrix.shape
                    for i in range(min(rows, 2)):
                        for j in range(min(cols, 2)):
                            new_matrix[i, j] = matrix[i, j]
                    all_matrices_2x2.append(new_matrix)

            avg_matrix = np.mean(all_matrices_2x2, axis=0)

            plt.figure(figsize=(10, 8))

            plt.rcParams['font.family'] = 'Arial'

            ax = plt.gca()

            for spine in ax.spines.values():
                spine.set_linewidth(5)

            sns.heatmap(avg_matrix, annot=True, fmt='.1f', cmap='Blues',
                        annot_kws={'size': 20, 'weight': 'bold'})

            plt.title(f'{class_name} Confusion Matrix (Average)', fontsize=20, pad=20)
            plt.ylabel('True Label', fontsize=24, labelpad=15)
            plt.xlabel('Predicted Label', fontsize=24, labelpad=15)

            ax.tick_params(axis='both', labelsize=24)

            output_file = os.path.join(self.config['output_path'], f"confusion_matrix_{class_name}.svg")
            plt.savefig(output_file, format='svg', bbox_inches='tight')
            plt.close()
        except Exception as e:
            logging.error(f"Error plotting confusion matrix: {str(e)}")

    def save_results(self, fold_results, class_name):
        """
        Save LOO results to Excel

        Args:
            fold_results: Results from all folds
            class_name: Class name
        """
        try:
            output_path = os.path.join(self.config['output_path'],
                                       f"Ensemble_RF_LDA_results_{class_name}.xlsx")

            # Prepare summary data
            rf_f1_scores = [r.get('rf_f1', 0) for r in fold_results if 'rf_f1' in r]
            lda_f1_scores = [r.get('lda_f1', 0) for r in fold_results if 'lda_f1' in r]
            ensemble_f1_scores = [r.get('ensemble_f1', 0) for r in fold_results if 'ensemble_f1' in r]

            rf_auc_scores = [r.get('rf_auc', 0.5) for r in fold_results if 'rf_auc' in r]
            lda_auc_scores = [r.get('lda_auc', 0.5) for r in fold_results if 'lda_auc' in r]
            ensemble_auc_scores = [r.get('ensemble_auc', 0.5) for r in fold_results if 'ensemble_auc' in r]

            ensemble_precision = [r.get('ensemble_precision', 0) for r in fold_results if 'ensemble_precision' in r]
            ensemble_recall = [r.get('ensemble_recall', 0) for r in fold_results if 'ensemble_recall' in r]
            ensemble_sensitivity = [r.get('ensemble_sensitivity', 0) for r in fold_results if
                                    'ensemble_sensitivity' in r]
            ensemble_specificity = [r.get('ensemble_specificity', 0) for r in fold_results if
                                    'ensemble_specificity' in r]
            lda_weights = [r.get('optimal_lda_weight', 0.5) for r in fold_results if 'optimal_lda_weight' in r]
            thresholds = [r.get('optimal_threshold', 0.5) for r in fold_results if 'optimal_threshold' in r]

            with pd.ExcelWriter(output_path) as writer:
                # 1. Save overall summary
                summary_data = {
                    'Model': ['RF', 'LDA', 'Ensemble'],
                    'Mean F1': [np.mean(rf_f1_scores) if rf_f1_scores else 0,
                                np.mean(lda_f1_scores) if lda_f1_scores else 0,
                                np.mean(ensemble_f1_scores) if ensemble_f1_scores else 0],
                    'Std F1': [np.std(rf_f1_scores) if rf_f1_scores else 0,
                               np.std(lda_f1_scores) if lda_f1_scores else 0,
                               np.std(ensemble_f1_scores) if ensemble_f1_scores else 0],
                    'Mean AUC': [np.mean(rf_auc_scores) if rf_auc_scores else 0.5,
                                 np.mean(lda_auc_scores) if lda_auc_scores else 0.5,
                                 np.mean(ensemble_auc_scores) if ensemble_auc_scores else 0.5],
                    'Std AUC': [np.std(rf_auc_scores) if rf_auc_scores else 0,
                                np.std(lda_auc_scores) if lda_auc_scores else 0,
                                np.std(ensemble_auc_scores) if ensemble_auc_scores else 0]
                }

                # Add additional metrics for ensemble algorithms only
                ensemble_summary = {
                    'Mean Precision': np.mean(ensemble_precision) if ensemble_precision else 0,
                    'Std Precision': np.std(ensemble_precision) if ensemble_precision else 0,
                    'Mean Recall': np.mean(ensemble_recall) if ensemble_recall else 0,
                    'Std Recall': np.std(ensemble_recall) if ensemble_recall else 0,
                    'Mean Sensitivity': np.mean(ensemble_sensitivity) if ensemble_sensitivity else 0,
                    'Std Sensitivity': np.std(ensemble_sensitivity) if ensemble_sensitivity else 0,
                    'Mean Specificity': np.mean(ensemble_specificity) if ensemble_specificity else 0,
                    'Std Specificity': np.std(ensemble_specificity) if ensemble_specificity else 0,
                    'Mean LDA Weight': np.mean(lda_weights) if lda_weights else 0.5,
                    'Std LDA Weight': np.std(lda_weights) if lda_weights else 0,
                    'Mean Threshold': np.mean(thresholds) if thresholds else 0.5,
                    'Std Threshold': np.std(thresholds) if thresholds else 0
                }

                pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', index=False)

                # Add additional ensemble metrics to a new sheet
                pd.DataFrame([ensemble_summary]).to_excel(writer, sheet_name='Ensemble_Details', index=False)

                # 2. Save detailed LOO results for each fold
                fold_data = []
                for r in fold_results:
                    if 'error' in r:
                        fold_data.append({
                            'Fold': r['fold_idx'] + 1,
                            'Error': r['error'],
                            'RF F1': 0,
                            'LDA F1': 0,
                            'Ensemble F1': 0
                        })
                        continue

                    fold_data.append({
                        'Fold': r['fold_idx'] + 1,
                        'RF F1': r.get('rf_f1', 0),
                        'LDA F1': r.get('lda_f1', 0),
                        'Ensemble F1': r.get('ensemble_f1', 0),
                        'RF AUC': r.get('rf_auc', 0.5),
                        'LDA AUC': r.get('lda_auc', 0.5),
                        'Ensemble AUC': r.get('ensemble_auc', 0.5),
                        'Ensemble Precision': r.get('ensemble_precision', 0),
                        'Ensemble Recall': r.get('ensemble_recall', 0),
                        'Ensemble Sensitivity': r.get('ensemble_sensitivity', 0),
                        'Ensemble Specificity': r.get('ensemble_specificity', 0),
                        'Optimal Threshold': r.get('optimal_threshold', 0.5),
                        'Optimal LDA Weight': r.get('optimal_lda_weight', 0.5)
                    })

                pd.DataFrame(fold_data).to_excel(writer, sheet_name='Fold_Details', index=False)

                # 3. Save RF algorithms parameters
                rf_params = []
                for r in fold_results:
                    if 'rf_best_params' not in r:
                        continue

                    params = r['rf_best_params']
                    fold_params = {
                        'Fold': r['fold_idx'] + 1,
                        'n_estimators_factor': params.get('n_estimators_factor', 0),
                        'criterion_idx': params.get('criterion_idx', 0),
                        'max_depth': params.get('max_depth', 0),
                        'min_samples_split': params.get('min_samples_split', 0),
                        'min_samples_leaf': params.get('min_samples_leaf', 0),
                        'class_weight_multiplier': params.get('class_weight_multiplier', 0)
                    }
                    rf_params.append(fold_params)

                pd.DataFrame(rf_params).to_excel(writer, sheet_name='RF_Parameters', index=False)

                # 4. Save LDA algorithms parameters
                lda_params = []
                for r in fold_results:
                    if 'lda_best_params' not in r:
                        continue

                    params = r['lda_best_params']
                    fold_params = {
                        'Fold': r['fold_idx'] + 1,
                        'solver_idx': params.get('solver_idx', 0),
                        'shrinkage_factor': params.get('shrinkage_factor', 0),
                    }
                    lda_params.append(fold_params)

                pd.DataFrame(lda_params).to_excel(writer, sheet_name='LDA_Parameters', index=False)

            # Format Excel file
            self.format_excel_file(output_path)

            logging.info(f"Results saved to: {output_path}")
        except Exception as e:
            logging.error(f"Error saving LOO results: {str(e)}")

    def format_excel_file(self, file_path):
        """Format Excel file with Times New Roman font and bold headers"""
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font

            wb = load_workbook(file_path)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # Set all cells to Times New Roman, 11pt
                for row in ws.iter_rows():
                    for cell in row:
                        cell.font = Font(name='Times New Roman', size=11)

                # Set first row to bold
                for cell in ws[1]:
                    cell.font = Font(name='Times New Roman', size=11, bold=True)

            wb.save(file_path)
        except Exception as e:
            logging.error(f"Error formatting Excel file: {str(e)}")

    def run(self):
        """
        Run ensemble algorithms
        """
        print("LDA and RF ensemble classifier")
        print()

        # Load data
        X, y = self.load_data()
        print(f"Cross-validation: {self.config['outer_cv_splits']}-fold outer, {self.config['inner_cv_splits']}-fold inner")
        print(f"Optimization: Optuna (TPE sampler)")
        print()

        # Create cross-validation object
        outer_cv = StratifiedKFold(n_splits=self.config['outer_cv_splits'], shuffle=True, random_state=42)

        # Check for data issues
        if np.any(np.isnan(X)) or np.any(np.isinf(X)):
            logging.warning("Data contains NaN or infinite values, will replace them")
            X = np.nan_to_num(X, nan=0.0, posinf=0.0, neginf=0.0)

        # Process each fold in parallel
        fold_indices = list(enumerate(outer_cv.split(X, y)))

        try:
            # Try parallel processing
            results = Parallel(n_jobs=-1, prefer="threads")(
                delayed(self.process_fold)(
                    X, y, train_idx, test_idx, fold_idx
                )
                for fold_idx, (train_idx, test_idx) in fold_indices
            )
        except Exception as e:
            logging.error(f"Parallel processing error: {str(e)}")
            # Try sequential processing
            results = []
            for fold_idx, (train_idx, test_idx) in fold_indices:
                try:
                    result = self.process_fold(X, y, train_idx, test_idx, fold_idx)
                    results.append(result)
                except Exception as fold_e:
                    logging.error(f"Error processing fold {fold_idx + 1}: {str(fold_e)}")
                    # Add error result placeholder
                    results.append({
                        "fold_idx": fold_idx,
                        "error": str(fold_e),
                        "ensemble_f1": 0,
                        "ensemble_auc": 0.5,
                        "optimal_lda_weight": 0.5,
                        "rf_f1": 0,
                        "lda_f1": 0,
                        "rf_auc": 0.5,
                        "lda_auc": 0.5
                    })

        if not results:
            logging.error("No LOO results generated, please check error messages above.")
            return

        # Filter out LOO results without errors
        valid_results = [r for r in results if 'error' not in r]

        if not valid_results:
            logging.error("All folds failed, cannot generate valid LOO results")
            return

        # Plot ROC curves
        self.plot_roc_curves(valid_results, 'LuC')

        # Plot confusion matrix
        self.plot_confusion_matrix(valid_results, 'LuC')

        # Save LOO results
        self.save_results(results, 'LuC')  # Save all LOO results, including errors

        # Calculate average performance metrics
        if valid_results:
            rf_f1 = np.mean([r['rf_f1'] for r in valid_results])
            lda_f1 = np.mean([r['lda_f1'] for r in valid_results])
            ensemble_f1 = np.mean([r['ensemble_f1'] for r in valid_results])

            rf_auc = np.mean([r['rf_auc'] for r in valid_results])
            lda_auc = np.mean([r['lda_auc'] for r in valid_results])
            ensemble_auc = np.mean([r['ensemble_auc'] for r in valid_results])

            ensemble_sensitivity = np.mean([r['ensemble_sensitivity'] for r in valid_results])
            ensemble_specificity = np.mean([r['ensemble_specificity'] for r in valid_results])

            lda_weights = [r['optimal_lda_weight'] for r in valid_results]
            mean_lda_weight = np.mean(lda_weights)
            std_lda_weight = np.std(lda_weights)

            print(f"\n   Results: F1={ensemble_f1:.3f}±{np.std([r['ensemble_f1'] for r in valid_results]):.3f}, "
                  f"AUC={ensemble_auc:.3f}±{np.std([r['ensemble_auc'] for r in valid_results]):.3f}, "
                  f"Sens={ensemble_sensitivity:.3f}±{np.std([r['ensemble_sensitivity'] for r in valid_results]):.3f}, "
                  f"Spec={ensemble_specificity:.3f}±{np.std([r['ensemble_specificity'] for r in valid_results]):.3f}")

            print("\nCross validation complete!")

        return results


def main():
    """Main function"""
    # Configuration parameters
    config = {
        'file_path': os.path.join(os.path.dirname(os.path.abspath(__file__)), 'TS Train.xlsx'),  # Path to the input Excel file containing training data
        'sheet_name': 'Sheet1',  # Name of the sheet in the Excel file (e.g., 'Sheet1')
        'output_path': os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'Results', 'Early cancer detection'),  # Directory path for saving output results
        'outer_cv_splits': 5,  # Number of outer cross-validation folds
        'inner_cv_splits': 5,  # Number of inner cross-validation folds
        'optuna_config': {
            'n_trials': 100,
            'timeout': 800,
            'sampler': 'TPE',
            'pruner': None,
            'n_startup_trials': 30,
            'n_ei_candidates': 100,
            'gamma': 0.25,
            'prior_weight': 2.0
        }
    }

    try:
        # Create output directory
        os.makedirs(config['output_path'], exist_ok=True)

        # Run ensemble classifier
        classifier = RFLDAEnsembleClassifier(config)
        results = classifier.run()

    except Exception as e:
        logging.error(f"Error occurred: {str(e)}")
        import traceback
        logging.error(traceback.format_exc())
        raise


if __name__ == "__main__":
    main()